import pandas as pd
import numpy as np
import json
import os
import sys
from datetime import datetime

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter

# ---------- 0. Paths ----------
CSV_PATH = "microbiome-responses.csv"      # change if needed
JSON_PATH = "microbiome-json.json"         # change if needed

coef_out_path_T = "coefficients_R45_by_panelist.csv"
coef_out_path_B = "coefficients_R12_by_panelist.csv"
coef_out_path_R = "coefficients_R_by_panelist.csv"

T_INTERCEPTS_CSV = "T_Intercepts.csv"
B_INTERCEPTS_CSV = "B_Intercepts.csv"
R_INTERCEPTS_CSV = "R_Intercepts.csv"

PANEL_COL  = "Panelist"
RATING_COL = "Rating"
RESPONSE_TIME_COL = "ResponseTime"
GENDER_COL = "Gender"
AGE_COL    = "Age"
TASK_COL   = "Task"     # last column of classification block

# global age bins
AGE_BINS = [
    "13-17",
    "18-24",
    "25-34",
    "35-44",
    "45-54",
    "55-64",
    "65+",
]

# ---------- 1. Load data ----------
# We use keep_default_na=False to prevent "None" from being interpreted as NaN
# We manually specify other common NA values
NA_VALUES = ['', '#N/A', '#N/A N/A', '#NA', '-1.#IND', '-1.#QNAN', '-NaN', '-nan', 
             '1.#IND', '1.#QNAN', '<NA>', 'N/A', 'NA', 'NULL', 'NaN', 'n/a', 'nan', 'null']
df = pd.read_csv(CSV_PATH, keep_default_na=False, na_values=NA_VALUES)

# ---------- 2. Load JSON ----------
with open(JSON_PATH, "r") as f:
    study = json.load(f)

elements_json = study.get("elements", [])
categories_json = study.get("categories", [])

# ---------- 2a. Excel filename from title ----------
def clean_title_for_filename(title: str) -> str:
    allowed_extra = {" ", "-", "_"}
    cleaned_chars = []
    for c in title:
        if c.isalnum() or c in allowed_extra:
            cleaned_chars.append(c)
    cleaned = "".join(cleaned_chars)
    cleaned = " ".join(cleaned.split())
    return cleaned or "study"

raw_title = str(study.get("title", "study"))
safe_title = clean_title_for_filename(raw_title)
excel_path = f"{safe_title}.xlsx"

# ---------- 2b. Element metadata ----------
element_meta = []
for el in elements_json:
    cat_obj = el.get("category", {})
    cat_name = cat_obj.get("name")
    el_name = el.get("name")
    if not cat_name or not el_name:
        continue
    col_name = f"{cat_name}_{el_name}"
    element_meta.append({
        "csv_col": col_name,
        "category_name": cat_name,
        "element_name": el_name,
        "category_order": cat_obj.get("order", 0),
    })

# Keep only columns that exist in CSV, preserving JSON order
element_cols = [m["csv_col"] for m in element_meta if m["csv_col"] in df.columns]
print("Detected element columns:", element_cols)
print("Total elements:", len(element_cols))

# Maps
col_to_catname = {
    m["csv_col"]: m["category_name"]
    for m in element_meta if m["csv_col"] in element_cols
}
col_to_eltname = {
    m["csv_col"]: m["element_name"]
    for m in element_meta if m["csv_col"] in element_cols
}

# Category ordering
cat_order = {}
for m in element_meta:
    col = m["csv_col"]
    if col not in element_cols:
        continue
    name = m["category_name"]
    order = m["category_order"]
    if name not in cat_order or order < cat_order[name]:
        cat_order[name] = order

sorted_categories = sorted(cat_order.keys(), key=lambda c: cat_order[c])

# ---------- 2c. Classification columns from CSV ----------
classification_cols = []
if GENDER_COL in df.columns and TASK_COL in df.columns:
    g_idx = df.columns.get_loc(GENDER_COL)
    t_idx = df.columns.get_loc(TASK_COL)
    if t_idx - g_idx > 1:
        classification_cols = list(df.columns[g_idx + 1:t_idx])

print("Detected classification columns:", classification_cols)

# ---------- 3. Gender helpers ----------
def normalize_gender(val):
    if not isinstance(val, str):
        return np.nan
    v = val.strip().lower()
    if v.startswith("m"):
        return "Male"
    if v.startswith("f"):
        return "Female"
    return np.nan

gender_map = None
if GENDER_COL in df.columns:
    gender_map = (
        df.dropna(subset=[GENDER_COL])
          .groupby(PANEL_COL)[GENDER_COL]
          .agg(lambda s: s.iloc[0])
    )

def build_gender_groups_from_coef(coef_table):
    groups = {}
    if gender_map is None:
        return groups

    coef_with_gender = coef_table.merge(
        gender_map.rename("Gender"),
        left_on="Panelist",
        right_index=True,
        how="left"
    )
    coef_with_gender["Gender_norm"] = coef_with_gender["Gender"].apply(normalize_gender)

    for g_name in ["Male", "Female"]:
        sub = coef_with_gender[coef_with_gender["Gender_norm"] == g_name]
        if not sub.empty:
            groups[g_name] = {
                "base": int(sub["Panelist"].nunique()),
                "means": sub[element_cols].mean(axis=0).round().astype(int) if "R2_Response" not in sub.columns else sub[element_cols].mean(axis=0)
            }
    return groups

# ---------- 3b. Age helpers ----------
def _age_string_to_bin(s: str):
    txt = s.strip()
    base = txt.replace(" ", "")
    for b in AGE_BINS:
        if base == b.replace(" ", ""):
            return b
    digits = "".join(ch if ch.isdigit() else " " for ch in txt)
    parts = [p for p in digits.split() if p]
    if parts:
        try:
            age_num = int(parts[0])
            return _age_number_to_bin(age_num)
        except ValueError:
            return np.nan
    return np.nan

def _age_number_to_bin(age: int):
    if age is None or np.isnan(age):
        return np.nan
    if 13 <= age <= 17:
        return "13-17"
    if 18 <= age <= 24:
        return "18-24"
    if 25 <= age <= 34:
        return "25-34"
    if 35 <= age <= 44:
        return "35-44"
    if 45 <= age <= 54:
        return "45-54"
    if 55 <= age <= 64:
        return "55-64"
    if age >= 65:
        return "65+"
    return np.nan

def normalize_age_to_bin(val):
    if pd.isna(val):
        return np.nan
    if isinstance(val, (int, float, np.integer, np.floating)):
        return _age_number_to_bin(int(val))
    if isinstance(val, str):
        return _age_string_to_bin(val)
    return np.nan

age_bin_map = None
if AGE_COL in df.columns:
    tmp_age = (
        df.dropna(subset=[AGE_COL])
          .groupby(PANEL_COL)[AGE_COL]
          .agg(lambda s: s.iloc[0])
    )
    age_bin_map = tmp_age.apply(normalize_age_to_bin)
    age_bin_map = age_bin_map.dropna()

def build_age_groups_from_coef(coef_table):
    groups = {}
    if age_bin_map is None or age_bin_map.empty:
        return groups

    coef_with_age = coef_table.merge(
        age_bin_map.rename("AgeBin"),
        left_on="Panelist",
        right_index=True,
        how="left"
    )

    for bin_label in AGE_BINS:
        sub = coef_with_age[coef_with_age["AgeBin"] == bin_label]
        if not sub.empty:
            groups[bin_label] = {
                "base": int(sub["Panelist"].nunique()),
                "means": sub[element_cols].mean(axis=0).round().astype(int) if "R2_Response" not in sub.columns else sub[element_cols].mean(axis=0)
            }
    return groups

# ---------- 3c. Classification helpers ----------
def build_class_groups_from_coef(coef_table):
    """
    Returns:
      {
        question_col_name: {
           "question_text": str,
           "answer_labels": [answer1, answer2, ...],
           "segments": {
               answer: {"base": int, "means": Series[int]}
           }
        },
        ...
      }
    """
    groups = {}
    if not classification_cols:
        return groups

    import sys
    for col_name in classification_cols:
        # print(f"DEBUG: Processing column: '{col_name}'", file=sys.stderr)
        if col_name not in df.columns:
            print(f"[WARN] Classification column not found in CSV: {col_name}")
            continue

        qtext = col_name  # column header is the question text

        ans_series = (
            df[[PANEL_COL, col_name]]
            .dropna(subset=[col_name])
            .groupby(PANEL_COL)[col_name]
            .agg(lambda s: s.iloc[0])
        )

        coef_with_ans = coef_table.merge(
            ans_series.rename("Answer"),
            left_on="Panelist",
            right_index=True,
            how="left"
        )

        segs = {}
        answer_labels = []

        unique_opts = ans_series.dropna().unique()
        # if "scientific issues" in col_name:
        #     print(f"DEBUG: {col_name} options: {unique_opts}", file=sys.stderr)
            
        #     if "None" in unique_opts:
        #         none_panelists = ans_series[ans_series == "None"].index
        #         coef_panelists = coef_table["Panelist"].unique()
        #         common = [p for p in none_panelists if p in coef_panelists]
        #         print(f"DEBUG: 'None' panelists: {len(none_panelists)}, In coef_table: {len(common)}", file=sys.stderr)

        for opt in unique_opts:
            sub = coef_with_ans[coef_with_ans["Answer"] == opt]
            if not sub.empty:
                segs[opt] = {
                    "base": int(sub["Panelist"].nunique()),
                    "means": sub[element_cols].mean(axis=0).round().astype(int) if "R2_Response" not in sub.columns else sub[element_cols].mean(axis=0)
                }
                answer_labels.append(opt)
            else:
                if "scientific issues" in col_name:
                    print(f"DEBUG: Option '{opt}' has empty subset in coef_table")

        if segs:
            groups[col_name] = {
                "question_text": qtext,
                "answer_labels": answer_labels,
                "segments": segs,
            }

    return groups

def flatten_class_groups(class_groups):
    """
    Optional: flatten to {answer_label -> segment}.
    (Used mainly for debugging.)
    """
    flat_labels = []
    flat_segments = {}

    for q_col in classification_cols:
        if q_col not in class_groups:
            continue
        info = class_groups[q_col]
        for ans in info["answer_labels"]:
            seg = info["segments"][ans]
            label = ans
            if label not in flat_segments:
                flat_labels.append(label)
                flat_segments[label] = seg
    return flat_labels, flat_segments

# ---------- 4. Regression helpers ----------
rng = np.random.default_rng(123)

def regress_one_panel_top(panel_df):
    X = panel_df[element_cols].to_numpy(dtype=float)
    ratings = panel_df[RATING_COL].to_numpy()
    Y = np.where(ratings >= 4, 100.0, 0.0)
    Y = Y + rng.uniform(-0.5, 0.5, size=Y.shape) * 1e-5
    beta, _, _, _ = np.linalg.lstsq(X, Y, rcond=None)
    Y_hat = X @ beta
    sse = float(np.sum((Y - Y_hat) ** 2))
    sst = float(np.sum(Y ** 2))
    r2 = np.nan if sst == 0 else 1.0 - sse / sst
    return beta, r2

def regress_one_panel_bottom(panel_df):
    X = panel_df[element_cols].to_numpy(dtype=float)
    ratings = panel_df[RATING_COL].to_numpy()
    Y = np.where(ratings <= 2, 100.0, 0.0)
    Y = Y + rng.uniform(-0.5, 0.5, size=Y.shape) * 1e-5
    beta, _, _, _ = np.linalg.lstsq(X, Y, rcond=None)
    Y_hat = X @ beta
    sse = float(np.sum((Y - Y_hat) ** 2))
    sst = float(np.sum(Y ** 2))
    r2 = np.nan if sst == 0 else 1.0 - sse / sst
    return beta, r2

def regress_one_panel_response_time(panel_df):
    X = panel_df[element_cols].to_numpy(dtype=float)
    # Cap at 7s, multiply by 100
    rt = panel_df[RESPONSE_TIME_COL].clip(upper=7.0).to_numpy()
    Y = rt
    beta, _, _, _ = np.linalg.lstsq(X, Y, rcond=None)
    Y_hat = X @ beta
    sse = float(np.sum((Y - Y_hat) ** 2))
    sst = float(np.sum(Y ** 2))
    r2 = np.nan if sst == 0 else 1.0 - sse / sst
    return beta, r2

def regress_no_intercept(X, Y):
    beta, _, _, _ = np.linalg.lstsq(X, Y, rcond=None)
    Y_hat = X @ beta
    sse = float(np.sum((Y - Y_hat) ** 2))
    sst = float(np.sum(Y ** 2))
    r2 = np.nan if sst == 0 else 1.0 - sse / sst
    return beta, r2

def regress_with_intercept_t(X, Y):
    n, p = X.shape
    X_design = np.column_stack([np.ones(n), X])
    beta_full, _, _, _ = np.linalg.lstsq(X_design, Y, rcond=None)

    Y_hat = X_design @ beta_full
    e = Y - Y_hat
    sse = float(np.sum(e ** 2))
    sst = float(np.sum((Y - Y.mean()) ** 2))
    r2 = np.nan if sst == 0 else 1.0 - sse / sst

    dof = n - X_design.shape[1]
    sigma2 = sse / dof
    XtX_inv = np.linalg.inv(X_design.T @ X_design)
    se = np.sqrt(np.diag(sigma2 * XtX_inv))
    t_vals = beta_full / se
    return beta_full, t_vals, r2

# ---------- 4b. Custom K-Means (Pearson Distance) ----------
def custom_kmeans_pearson(data, k, max_iters=100, seed=42):
    """
    K-Means minimizing 1 - Pearson Correlation.
    Equivalent to Spherical K-Means on centered data.
    """
    # 1. Preprocess: Center and Normalize rows
    # Center
    row_means = data.mean(axis=1, keepdims=True)
    centered = data - row_means
    # Normalize
    norms = np.linalg.norm(centered, axis=1, keepdims=True)
    norms[norms == 0] = 1e-9 # avoid div by zero
    normalized = centered / norms

    n_samples, n_features = normalized.shape
    rng = np.random.default_rng(seed)

    # 2. Init centroids (random points)
    indices = rng.choice(n_samples, k, replace=False)
    centroids = normalized[indices]

    labels = np.zeros(n_samples, dtype=int)
    
    for _ in range(max_iters):
        # 3. Assign clusters (Max Dot Product = Min Cosine Dist)
        # centroids are also unit vectors (maintained below)
        dots = normalized @ centroids.T
        new_labels = np.argmax(dots, axis=1)

        if np.array_equal(new_labels, labels):
            break
        labels = new_labels

        # 4. Update centroids
        new_centroids = np.zeros_like(centroids)
        for i in range(k):
            points = normalized[labels == i]
            if len(points) > 0:
                # Mean of points
                mean_vec = points.mean(axis=0)
                # Normalize (Spherical K-Means step)
                norm = np.linalg.norm(mean_vec)
                if norm > 1e-9:
                    new_centroids[i] = mean_vec / norm
                else:
                    # re-init if collapsed
                    new_centroids[i] = normalized[rng.choice(n_samples)]
            else:
                 # re-init empty cluster
                new_centroids[i] = normalized[rng.choice(n_samples)]
        centroids = new_centroids

    return labels

# ---------- 5. Panel-level coefficients: TOP ----------
rows_T = []
for pid, g in df.groupby(PANEL_COL):
    beta, r2 = regress_one_panel_top(g)
    row = {"Panelist": pid, "R2_R45": r2}
    for col_name, b in zip(element_cols, beta):
        row[col_name] = b
    rows_T.append(row)

coef_table_T = pd.DataFrame(rows_T)
coef_table_T.to_csv(coef_out_path_T, index=False)
print("Saved TOP coefficients:", coef_out_path_T)

# ---------- 6. Panel-level coefficients: BOTTOM ----------
rows_B = []
for pid, g in df.groupby(PANEL_COL):
    beta, r2 = regress_one_panel_bottom(g)
    row = {"Panelist": pid, "R2_R12": r2}
    for col_name, b in zip(element_cols, beta):
        row[col_name] = b
    rows_B.append(row)

coef_table_B = pd.DataFrame(rows_B)
coef_table_B.to_csv(coef_out_path_B, index=False)
print("Saved BOTTOM coefficients:", coef_out_path_B)

# ---------- 6b. Panel-level coefficients: RESPONSE TIME ----------
rows_R = []
for pid, g in df.groupby(PANEL_COL):
    beta, r2 = regress_one_panel_response_time(g)
    row = {"Panelist": pid, "R2_Response": r2}
    for col_name, b in zip(element_cols, beta):
        row[col_name] = b
    rows_R.append(row)

coef_table_R = pd.DataFrame(rows_R)
coef_table_R.to_csv(coef_out_path_R, index=False)
print("Saved RESPONSE coefficients:", coef_out_path_R)

# ---------- 7. Base size & means ----------
base_size = df[PANEL_COL].nunique()

element_means_T = coef_table_T[element_cols].mean(axis=0).round().astype(int)
element_means_B = coef_table_B[element_cols].mean(axis=0).round().astype(int)

gender_groups_T = build_gender_groups_from_coef(coef_table_T)
gender_groups_B = build_gender_groups_from_coef(coef_table_B)
age_groups_T    = build_age_groups_from_coef(coef_table_T)
age_groups_B    = build_age_groups_from_coef(coef_table_B)
class_groups_T  = build_class_groups_from_coef(coef_table_T)
class_groups_B  = build_class_groups_from_coef(coef_table_B)

element_means_R = coef_table_R[element_cols].mean(axis=0)
gender_groups_R = build_gender_groups_from_coef(coef_table_R)
age_groups_R    = build_age_groups_from_coef(coef_table_R)
class_groups_R  = build_class_groups_from_coef(coef_table_R)

class_labels_T, class_segments_T = flatten_class_groups(class_groups_T)
class_labels_B, class_segments_B = flatten_class_groups(class_groups_B)

print("Gender TOP:", list(gender_groups_T.keys()))
print("Gender BOTTOM:", list(gender_groups_B.keys()))
print("Age TOP bins:", list(age_groups_T.keys()))
print("Age BOTTOM bins:", list(age_groups_B.keys()))
print("Class TOP questions:", list(class_groups_T.keys()))
print("Class BOTTOM questions:", list(class_groups_B.keys()))

# ---------- 8. Pooled regressions for Intercepts ----------
X_all = df[element_cols].to_numpy(dtype=float)
ratings_all = df[RATING_COL].to_numpy()

# TOP (R45)
Y_all_T = np.where(ratings_all >= 4, 100.0, 0.0)
Y_all_T = Y_all_T + np.random.default_rng(123).uniform(-0.5, 0.5, size=Y_all_T.shape) * 1e-5

beta_no_T, _ = regress_no_intercept(X_all, Y_all_T)
beta_full_T, t_vals_T, _ = regress_with_intercept_t(X_all, Y_all_T)
beta_with_T = beta_full_T[1:]
t_elements_T = t_vals_T[1:]

pooled_df_T = pd.DataFrame({
    "element": element_cols,
    "beta_no_intercept": beta_no_T,
    "beta_with_intercept": beta_with_T,
    "t_with_intercept": t_elements_T
})

mask_T = t_elements_T >= 2.0
coef_threshold_T = None if not np.any(mask_T) else float(np.min(beta_no_T[mask_T]))
print("TOP coefficient threshold (t>=2) =", coef_threshold_T)

# BOTTOM (R12)
Y_all_B = np.where(ratings_all <= 2, 100.0, 0.0)
Y_all_B = Y_all_B + np.random.default_rng(456).uniform(-0.5, 0.5, size=Y_all_B.shape) * 1e-5

beta_no_B, _ = regress_no_intercept(X_all, Y_all_B)
beta_full_B, t_vals_B, _ = regress_with_intercept_t(X_all, Y_all_B)
beta_with_B = beta_full_B[1:]
t_elements_B = t_vals_B[1:]

pooled_df_B = pd.DataFrame({
    "element": element_cols,
    "beta_no_intercept": beta_no_B,
    "beta_with_intercept": beta_with_B,
    "t_with_intercept": t_elements_B
})

mask_B = t_elements_B >= 2.0
coef_threshold_B = None if not np.any(mask_B) else float(np.min(beta_no_B[mask_B]))
print("BOTTOM coefficient threshold (t>=2) =", coef_threshold_B)

# RESPONSE TIME (R)
rt_all = df[RESPONSE_TIME_COL].clip(upper=7.0).to_numpy()
Y_all_R = rt_all

beta_no_R, _ = regress_no_intercept(X_all, Y_all_R)
beta_full_R, t_vals_R, _ = regress_with_intercept_t(X_all, Y_all_R)
beta_with_R = beta_full_R[1:]
t_elements_R = t_vals_R[1:]

pooled_df_R = pd.DataFrame({
    "element": element_cols,
    "beta_no_intercept": beta_no_R,
    "beta_with_intercept": beta_with_R,
    "t_with_intercept": t_elements_R
})

mask_R = t_elements_R >= 2.0
coef_threshold_R = None if not np.any(mask_R) else float(np.min(beta_no_R[mask_R]))
print("RESPONSE coefficient threshold (t>=2) =", coef_threshold_R)

pooled_df_T.to_csv(T_INTERCEPTS_CSV, index=False)
pooled_df_B.to_csv(B_INTERCEPTS_CSV, index=False)
pooled_df_R.to_csv(R_INTERCEPTS_CSV, index=False)
print("Saved intercept CSVs:", T_INTERCEPTS_CSV, B_INTERCEPTS_CSV, R_INTERCEPTS_CSV)

# ---------- 9. Write intercept sheets ----------
with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
    pooled_df_T.to_excel(writer, sheet_name="(T) Intercepts", index=False)
    pooled_df_B.to_excel(writer, sheet_name="(B) Intercepts", index=False)
    pooled_df_R.to_excel(writer, sheet_name="(R) Intercepts", index=False)

print("Base Excel (intercepts) written:", excel_path)

# ---------- 10. Build formatted sheets ----------
wb = load_workbook(excel_path)

bold_font   = Font(bold=True)
header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
green_fill  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill    = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
blue_fill   = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

MAX_WIDTH = 45
letters = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"

def wrap_col_B(ws):
    max_len = 0
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=2).value
        if v is not None:
            max_len = max(max_len, len(str(v)))
    ws.column_dimensions["B"].width = min(max_len + 2, MAX_WIDTH)
    for r in range(1, ws.max_row + 1):
        cell = ws.cell(row=r, column=2)
        cell.alignment = Alignment(
            horizontal=cell.alignment.horizontal or "general",
            vertical=cell.alignment.vertical or "top",
            wrap_text=True
        )

def autofit_all_cols(ws, max_width=30):
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        if max_len > 0:
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)

def set_classification_col_widths(ws, start_col=4, max_width=20):
    """Set width of all columns from start_col to end to match B (capped)."""
    width_B = ws.column_dimensions["B"].width or max_width
    col_width = min(width_B, max_width)
    for col_idx in range(start_col, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = col_width
        for r in range(1, ws.max_row + 1):
            cell = ws.cell(row=r, column=col_idx)
            cell.alignment = Alignment(
                horizontal=cell.alignment.horizontal or "general",
                vertical=cell.alignment.vertical or "top",
                wrap_text=True
            )

def create_front_page(wb, json_path):
    with open(json_path, "r") as f:
        data = json.load(f)
    
    ws = wb.create_sheet("Front Page", 0)
    
    # Data extraction
    title = data.get("title", "")
    background = data.get("background", "")
    language = data.get("language", "")
    launched_at_raw = data.get("launched_at", "")
    
    # Format date: YYYY-MM-DD... -> MM.DD.YYYY
    launched_at = ""
    if launched_at_raw:
        try:
            date_part = launched_at_raw.split("T")[0]
            yyyy, mm, dd = date_part.split("-")
            launched_at = f"{mm}.{dd}.{yyyy}"
        except:
            launched_at = launched_at_raw

    # Layout
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 50
    
    labels = ["Title:", "Background:", "Language:", "Launched At:"]
    values = [title, background, language, launched_at]
    
    thick = Side(border_style="thick", color="000000")
    thin = Side(border_style="thin", color="000000")
    
    start_row = 2
    for i, (lab, val) in enumerate(zip(labels, values)):
        r = start_row + i
        c_lab = ws.cell(row=r, column=2, value=lab)
        c_val = ws.cell(row=r, column=3, value=val)
        c_lab.font = Font(bold=True)
        
        # Determine borders
        top = thick if i == 0 else thin
        bottom = thick if i == len(labels) - 1 else thin
        
        c_lab.border = Border(top=top, left=thick, right=thin, bottom=bottom)
        c_val.border = Border(top=top, left=thin, right=thick, bottom=bottom)

def create_info_block(wb, json_path):
    with open(json_path, "r") as f:
        data = json.load(f)
    
    ws = wb.create_sheet("Information Block", 1)
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 60
    
    # --- Helper to write rows ---
    current_row = 1
    
    def write_row(label, value, bold_label=True):
        nonlocal current_row
        c_lab = ws.cell(row=current_row, column=1, value=label)
        c_val = ws.cell(row=current_row, column=2, value=value)
        if bold_label:
            c_lab.font = Font(bold=True)
        c_val.alignment = Alignment(wrap_text=True)
        current_row += 1

    # --- 1. General Info ---
    # Study title: title
    write_row("Study Title", data.get("title", ""))
    
    # Study Date: launched_at
    # if country === US use mm-dd-yyyy format else dd-mm-yyyy
    aud = data.get("audience_segmentation", {})
    country = aud.get("country", "")
    launched_at_raw = data.get("launched_at", "")
    launched_at_fmt = ""
    if launched_at_raw:
        try:
            # Parse ISO format
            dt = datetime.fromisoformat(launched_at_raw.replace("Z", "+00:00"))
            if country == "US":
                launched_at_fmt = dt.strftime("%m-%d-%Y")
            else:
                launched_at_fmt = dt.strftime("%d-%m-%Y")
        except:
            launched_at_fmt = launched_at_raw
    write_row("Study Date", launched_at_fmt)
    
    # Number of Respondents
    # Make sure value for Number of Respondents is aligned left
    resp_val = aud.get("number_of_respondents", "")
    c_lab = ws.cell(row=current_row, column=1, value="Number of Respondents")
    c_val = ws.cell(row=current_row, column=2, value=resp_val)
    c_lab.font = Font(bold=True)
    c_val.alignment = Alignment(horizontal='left', wrap_text=True)
    current_row += 1
    
    # Country
    write_row("Country", country)
    
    # Gender Distribution
    gender_dist = aud.get("gender_distribution", {})
    g_dist_str = ", ".join([f"{k}: {v}%" for k, v in gender_dist.items() if v > 0])
    write_row("Gender Distribution", g_dist_str)

    # Age Distribution
    age_dist = aud.get("age_distribution", {})
    a_dist_str = ", ".join([f"{k}: {v}%" for k, v in age_dist.items() if v > 0])
    write_row("Age Distribution", a_dist_str)
    
    # Study Background
    write_row("Study Background", data.get("background", ""))
    
    # Study Orientation
    write_row("Study Orientation", data.get("orientation_text", ""))
    
    # Study Type
    write_row("Study Type", data.get("study_type", ""))
    
    write_row("", "") # Spacer
    
    # --- 2. Study Categories ---
    # Categories: categories.name
    # Elements: elements.name, elements.description -> element.content
    
    categories = data.get("categories", [])
    elements = data.get("elements", [])
    
    # Map category_id to elements
    cat_elements = {}
    for el in elements:
        cid = el.get("category_id")
        if cid not in cat_elements:
            cat_elements[cid] = []
        cat_elements[cid].append(el)
        
    for cat in categories:
        cat_name = cat.get("name", "")
        cat_id = cat.get("category_id") or cat.get("id") # Try both
        
        write_row("Study Category", cat_name)
        
        # Find elements for this category
        # Note: JSON structure shows elements have "category_id" matching category "id" (not "category_id" inside category object? Let's check JSON again.
        # Category object: "category_id": "...", "id": "..."
        # Element object: "category_id": "..."
        # In the JSON view:
        # Category 0: "id": "a231b615...", "category_id": "5f6889b2..."
        # Element 0: "category_id": "a231b615..." (Matches Category "id")
        
        # So we match element["category_id"] to category["id"]
        
        c_elements = [e for e in elements if e.get("category_id") == cat.get("id")]
        
        for el in c_elements:
            el_name = el.get("name", "")
            el_content = el.get("content", "") # User mapped "elements.description" label to "element.content" value
            
            write_row(f"{cat_name} element", el_name)
            write_row("Element Content", el_content)
            
        write_row("", "") # Spacer

    # --- 3. Classification Questions ---
    class_qs = data.get("classification_questions", [])
    for q in class_qs:
        q_text = q.get("question_text", "")
        ans_opts = q.get("answer_options", [])
        
        write_row("Classification Question", q_text)
        
        # Format answers
        # User screenshot shows "1=..., 2=..."
        # Or just list them? User said "Possible answers: answer_options"
        # I'll join them with newlines or commas. Newlines might be cleaner in a cell.
        # Or maybe "1=Option1, 2=Option2..." if they are indexed.
        # The JSON is a list of strings.
        # Let's try "1=Option1\n2=Option2"
        ans_str = "\n".join([f"{i+1}={opt}" for i, opt in enumerate(ans_opts)])
        write_row("Possible answers", ans_str)
        
    write_row("", "") # Spacer

    # --- 4. Rating Question ---
    # Rating question: main_question
    main_q = data.get("main_question", "")
    write_row("Rating question", main_q)
    
    # Ratings
    # rating_scale.min_value = rating_scale.min_label
    # rating_scale.middle_label
    # rating_scale.max_value = rating_scale.max_label
    
    rs = data.get("rating_scale", {})
    min_v = rs.get("min_value", 1)
    max_v = rs.get("max_value", 5)
    min_l = rs.get("min_label", "")
    max_l = rs.get("max_label", "")
    mid_l = rs.get("middle_label", "")
    
    # Construct the rating string
    # "1=min_label\n... middle ...\n5=max_label"
    # If it's 1 to 5.
    # We don't have labels for 2, 4 usually in the JSON unless implied.
    # User said:
    # ratings: rating_scale.min_value = rating_scale.min_label
    # rating_scale.middle_label
    # rating_scale.max_value = rating_scale.max_label
    
    # I will format it as:
    # {min_value}={min_label}
    # Middle: {middle_label}
    # {max_value}={max_label}
    
    rating_lines = []
    rating_lines.append(f"{min_v}={min_l}")
    if mid_l:
        rating_lines.append(f"Middle={mid_l}")
    rating_lines.append(f"{max_v}={max_l}")
    
    write_row("Rating Scale", "\n".join(rating_lines))


# ---------- 10a. Intercepts formatting ----------
ws_T_int = wb["(T) Intercepts"]
ws_B_int = wb["(B) Intercepts"]
ws_R_int = wb["(R) Intercepts"]

for ws, coef_thr in [(ws_T_int, coef_threshold_T), (ws_B_int, coef_threshold_B), (ws_R_int, coef_threshold_R)]:
    if coef_thr is not None:
        last = ws.max_row
        thr_row = last + 2
        cell_label = ws.cell(row=thr_row, column=1, value="Threshold")
        cell_label.font = bold_font
        ws.cell(row=thr_row, column=2, value=coef_thr)
        ws.cell(row=thr_row, column=4, value=2.0)

    autofit_all_cols(ws, max_width=45)

last_T_raw = ws_T_int.max_row
rule_t_T = CellIsRule(
    operator="greaterThanOrEqual",
    formula=["2.0"],
    fill=green_fill,
)
ws_T_int.conditional_formatting.add(f"D2:D{last_T_raw}", rule_t_T)

last_B_raw = ws_B_int.max_row
rule_t_B = CellIsRule(
    operator="greaterThanOrEqual",
    formula=["2.0"],
    fill=red_fill,
)
ws_B_int.conditional_formatting.add(f"D2:D{last_B_raw}", rule_t_B)

last_R_raw = ws_R_int.max_row
rule_t_R = CellIsRule(
    operator="greaterThanOrEqual",
    formula=["2.0"],
    fill=blue_fill,
)
ws_R_int.conditional_formatting.add(f"D2:D{last_R_raw}", rule_t_R)

# ---------- (T) Overall ----------
ws_T_overall = wb.create_sheet("(T) Overall", 0)

ws_T_overall["B1"] = "Group (Binary Ratings)"; ws_T_overall["B1"].font = bold_font
ws_T_overall["D1"] = "Total";                 ws_T_overall["D1"].font = bold_font
ws_T_overall["B2"] = "Base Size";             ws_T_overall["B2"].font = bold_font
ws_T_overall["D2"] = base_size

row = 5
first_val_T = last_val_T = None

for cat_idx, cat_name in enumerate(sorted_categories):
    letter = letters[cat_idx]
    ws_T_overall.cell(row=row, column=2, value=f"{letter}. {cat_name}")
    cat_cell = ws_T_overall.cell(row=row, column=2)
    cat_cell.font = bold_font
    cat_cell.fill = header_fill
    row += 1

    cols_in_cat = [c for c in element_cols if col_to_catname.get(c) == cat_name]
    for elt_idx, col in enumerate(cols_in_cat, start=1):
        code = f"{letter}{elt_idx}"
        elt_name = col_to_eltname.get(col, col)
        val = int(element_means_T[col])

        ws_T_overall.cell(row=row, column=1, value=code)
        ws_T_overall.cell(row=row, column=2, value=elt_name)
        ws_T_overall.cell(row=row, column=4, value=val)

        if first_val_T is None:
            first_val_T = row
        last_val_T = row
        row += 1

    row += 1

wrap_col_B(ws_T_overall)

if coef_threshold_T is not None and first_val_T is not None:
    formula_T = f'AND(D{first_val_T}<>"",D{first_val_T}>={coef_threshold_T})'
    rule_T = FormulaRule(
        formula=[formula_T],
        fill=green_fill,
    )
    ws_T_overall.conditional_formatting.add(
        f"D{first_val_T}:D{last_val_T}",
        rule_T
    )

# ---------- (B) Overall ----------
ws_B_overall = wb.create_sheet("(B) Overall", 1)

ws_B_overall["B1"] = "Group (Binary Ratings)"; ws_B_overall["B1"].font = bold_font
ws_B_overall["D1"] = "Total";                 ws_B_overall["D1"].font = bold_font
ws_B_overall["B2"] = "Base Size";             ws_B_overall["B2"].font = bold_font
ws_B_overall["D2"] = base_size

row = 5
first_val_B = last_val_B = None

for cat_idx, cat_name in enumerate(sorted_categories):
    letter = letters[cat_idx]
    ws_B_overall.cell(row=row, column=2, value=f"{letter}. {cat_name}")
    cat_cell = ws_B_overall.cell(row=row, column=2)
    cat_cell.font = bold_font
    cat_cell.fill = header_fill
    row += 1

    cols_in_cat = [c for c in element_cols if col_to_catname.get(c) == cat_name]
    for elt_idx, col in enumerate(cols_in_cat, start=1):
        code = f"{letter}{elt_idx}"
        elt_name = col_to_eltname.get(col, col)
        val = int(element_means_B[col])

        ws_B_overall.cell(row=row, column=1, value=code)
        ws_B_overall.cell(row=row, column=2, value=elt_name)
        ws_B_overall.cell(row=row, column=4, value=val)

        if first_val_B is None:
            first_val_B = row
        last_val_B = row
        row += 1

    row += 1

wrap_col_B(ws_B_overall)

if coef_threshold_B is not None and first_val_B is not None:
    formula_B = f'AND(D{first_val_B}<>"",D{first_val_B}>={coef_threshold_B})'
    rule_B = FormulaRule(
        formula=[formula_B],
        fill=red_fill,
    )
    ws_B_overall.conditional_formatting.add(
        f"D{first_val_B}:D{last_val_B}",
        rule_B
    )

# ---------- (T) Gender ----------
if gender_groups_T:
    ws_T_gender = wb.create_sheet("(T) Gender", 2)
    ws_T_gender["B1"] = "Group (Binary Ratings)"; ws_T_gender["B1"].font = bold_font
    ws_T_gender["B2"] = "Base Size";             ws_T_gender["B2"].font = bold_font

    col_map_T = {}
    col_idx = 4  # D
    for g_name in ["Male", "Female"]:
        if g_name in gender_groups_T:
            ws_T_gender.cell(row=1, column=col_idx, value=g_name)
            ws_T_gender.cell(row=1, column=col_idx).font = bold_font
            ws_T_gender.cell(row=2, column=col_idx, value=gender_groups_T[g_name]["base"])
            col_map_T[g_name] = col_idx
            col_idx += 1

    row = 5
    first_val_gT = last_val_gT = None

    for cat_idx, cat_name in enumerate(sorted_categories):
        letter = letters[cat_idx]
        ws_T_gender.cell(row=row, column=2, value=f"{letter}. {cat_name}")
        cat_cell = ws_T_gender.cell(row=row, column=2)
        cat_cell.font = bold_font
        cat_cell.fill = header_fill
        row += 1

        cols_in_cat = [c for c in element_cols if col_to_catname.get(c) == cat_name]
        for elt_idx, col in enumerate(cols_in_cat, start=1):
            code = f"{letter}{elt_idx}"
            elt_name = col_to_eltname.get(col, col)

            ws_T_gender.cell(row=row, column=1, value=code)
            ws_T_gender.cell(row=row, column=2, value=elt_name)

            for g_name, info in gender_groups_T.items():
                c = col_map_T[g_name]
                ws_T_gender.cell(row=row, column=c, value=info["means"][col])

            if first_val_gT is None:
                first_val_gT = row
            last_val_gT = row
            row += 1
        row += 1

    wrap_col_B(ws_T_gender)

    if coef_threshold_T is not None and first_val_gT is not None:
        for g_name, c_idx in col_map_T.items():
            col_letter = ws_T_gender.cell(row=1, column=c_idx).column_letter
            formula = f'AND({col_letter}{first_val_gT}<>"",{col_letter}{first_val_gT}>={coef_threshold_T})'
            rule = FormulaRule(
                formula=[formula],
                fill=green_fill,
            )
            ws_T_gender.conditional_formatting.add(
                f"{col_letter}{first_val_gT}:{col_letter}{last_val_gT}",
                rule
            )

# ---------- (B) Gender ----------
if gender_groups_B:
    ws_B_gender = wb.create_sheet("(B) Gender", 3)
    ws_B_gender["B1"] = "Group (Binary Ratings)"; ws_B_gender["B1"].font = bold_font
    ws_B_gender["B2"] = "Base Size";             ws_B_gender["B2"].font = bold_font

    col_map_B = {}
    col_idx = 4  # D
    for g_name in ["Male", "Female"]:
        if g_name in gender_groups_B:
            ws_B_gender.cell(row=1, column=col_idx, value=g_name)
            ws_B_gender.cell(row=1, column=col_idx).font = bold_font
            ws_B_gender.cell(row=2, column=col_idx, value=gender_groups_B[g_name]["base"])
            col_map_B[g_name] = col_idx
            col_idx += 1

    row = 5
    first_val_gB = last_val_gB = None

    for cat_idx, cat_name in enumerate(sorted_categories):
        letter = letters[cat_idx]
        ws_B_gender.cell(row=row, column=2, value=f"{letter}. {cat_name}")
        cat_cell = ws_B_gender.cell(row=row, column=2)
        cat_cell.font = bold_font
        cat_cell.fill = header_fill
        row += 1

        cols_in_cat = [c for c in element_cols if col_to_catname.get(c) == cat_name]
        for elt_idx, col in enumerate(cols_in_cat, start=1):
            code = f"{letter}{elt_idx}"
            elt_name = col_to_eltname.get(col, col)

            ws_B_gender.cell(row=row, column=1, value=code)
            ws_B_gender.cell(row=row, column=2, value=elt_name)

            for g_name, info in gender_groups_B.items():
                c = col_map_B[g_name]
                val = int(info["means"][col])
                ws_B_gender.cell(row=row, column=c, value=val)

            if first_val_gB is None:
                first_val_gB = row
            last_val_gB = row
            row += 1

        row += 1

    wrap_col_B(ws_B_gender)

    if coef_threshold_B is not None and first_val_gB is not None:
        for g_name, c_idx in col_map_B.items():
            col_letter = ws_B_gender.cell(row=1, column=c_idx).column_letter
            formula = f'AND({col_letter}{first_val_gB}<>"",{col_letter}{first_val_gB}>={coef_threshold_B})'
            rule = FormulaRule(
                formula=[formula],
                fill=red_fill,
            )
            ws_B_gender.conditional_formatting.add(
                f"{col_letter}{first_val_gB}:{col_letter}{last_val_gB}",
                rule
            )

# ---------- (T) Age ----------
if age_groups_T:
    ws_T_age = wb.create_sheet("(T) Age")
    ws_T_age["B1"] = "Group (Binary Ratings)"; ws_T_age["B1"].font = bold_font
    ws_T_age["B2"] = "Base Size";             ws_T_age["B2"].font = bold_font

    col_map_age_T = {}
    col_idx = 4  # D
    for age_label in AGE_BINS:
        if age_label in age_groups_T:
            ws_T_age.cell(row=1, column=col_idx, value=age_label)
            ws_T_age.cell(row=1, column=col_idx).font = bold_font
            ws_T_age.cell(row=2, column=col_idx, value=age_groups_T[age_label]["base"])
            col_map_age_T[age_label] = col_idx
            col_idx += 1

    row = 5
    first_val_aT = last_val_aT = None

    for cat_idx, cat_name in enumerate(sorted_categories):
        letter = letters[cat_idx]
        ws_T_age.cell(row=row, column=2, value=f"{letter}. {cat_name}")
        cat_cell = ws_T_age.cell(row=row, column=2)
        cat_cell.font = bold_font
        cat_cell.fill = header_fill
        row += 1

        cols_in_cat = [c for c in element_cols if col_to_catname.get(c) == cat_name]
        for elt_idx, col in enumerate(cols_in_cat, start=1):
            code = f"{letter}{elt_idx}"
            elt_name = col_to_eltname.get(col, col)

            ws_T_age.cell(row=row, column=1, value=code)
            ws_T_age.cell(row=row, column=2, value=elt_name)

            for age_label, info in age_groups_T.items():
                c = col_map_age_T.get(age_label)
                if c is not None:
                    ws_T_age.cell(row=row, column=c, value=info["means"][col])

            if first_val_aT is None:
                first_val_aT = row
            last_val_aT = row
            row += 1
        row += 1

    wrap_col_B(ws_T_age)

    if coef_threshold_T is not None and first_val_aT is not None:
        for age_label, c_idx in col_map_age_T.items():
            col_letter = ws_T_age.cell(row=1, column=c_idx).column_letter
            formula = f'AND({col_letter}{first_val_aT}<>"",{col_letter}{first_val_aT}>={coef_threshold_T})'
            rule = FormulaRule(
                formula=[formula],
                fill=green_fill,
            )
            ws_T_age.conditional_formatting.add(
                f"{col_letter}{first_val_aT}:{col_letter}{last_val_aT}",
                rule
            )

# ---------- (B) Age ----------
if age_groups_B:
    ws_B_age = wb.create_sheet("(B) Age")
    ws_B_age["B1"] = "Group (Binary Ratings)"; ws_B_age["B1"].font = bold_font
    ws_B_age["B2"] = "Base Size";             ws_B_age["B2"].font = bold_font

    col_map_age_B = {}
    col_idx = 4  # D
    for age_label in AGE_BINS:
        if age_label in age_groups_B:
            ws_B_age.cell(row=1, column=col_idx, value=age_label)
            ws_B_age.cell(row=1, column=col_idx).font = bold_font
            ws_B_age.cell(row=2, column=col_idx, value=age_groups_B[age_label]["base"])
            col_map_age_B[age_label] = col_idx
            col_idx += 1

    row = 5
    first_val_aB = last_val_aB = None

    for cat_idx, cat_name in enumerate(sorted_categories):
        letter = letters[cat_idx]
        ws_B_age.cell(row=row, column=2, value=f"{letter}. {cat_name}")
        cat_cell = ws_B_age.cell(row=row, column=2)
        cat_cell.font = bold_font
        cat_cell.fill = header_fill
        row += 1

        cols_in_cat = [c for c in element_cols if col_to_catname.get(c) == cat_name]
        for elt_idx, col in enumerate(cols_in_cat, start=1):
            code = f"{letter}{elt_idx}"
            elt_name = col_to_eltname.get(col, col)

            ws_B_age.cell(row=row, column=1, value=code)
            ws_B_age.cell(row=row, column=2, value=elt_name)

            for age_label, info in age_groups_B.items():
                c = col_map_age_B.get(age_label)
                if c is not None:
                    val = int(info["means"][col])
                    ws_B_age.cell(row=row, column=c, value=val)

            if first_val_aB is None:
                first_val_aB = row
            last_val_aB = row
            row += 1

        row += 1

    wrap_col_B(ws_B_age)

    if coef_threshold_B is not None and first_val_aB is not None:
        for age_label, c_idx in col_map_age_B.items():
            col_letter = ws_B_age.cell(row=1, column=c_idx).column_letter
            formula = f'AND({col_letter}{first_val_aB}<>"",{col_letter}{first_val_aB}>={coef_threshold_B})'
            rule = FormulaRule(
                formula=[formula],
                fill=red_fill,
            )
            ws_B_age.conditional_formatting.add(
                f"{col_letter}{first_val_aB}:{col_letter}{last_val_aB}",
                rule
            )

# ---------- (T) Classification Questions ----------
if class_groups_T:
    ws_T_cls = wb.create_sheet("(T) Classification Questions")
    ws_T_cls["B1"] = "Group (Binary Ratings)"; ws_T_cls["B1"].font = bold_font
    ws_T_cls["B2"] = "Base Size";             ws_T_cls["B2"].font = bold_font

    # Map (question_col, answer) -> column index
    col_map_cls_T = {}
    col_idx = 4  # D

    # Header rows: questions + answer columns
    for q_col in classification_cols:
        info = class_groups_T.get(q_col)
        if not info:
            continue
        qtext = info["question_text"]
        answer_labels = info["answer_labels"]

        # Question column
        q_cell = ws_T_cls.cell(row=1, column=col_idx, value=qtext)
        
        q_cell.font = bold_font
        q_cell.fill = header_fill
        q_cell.alignment = Alignment(wrap_text=True)

        # Answer columns for this question
        for ans in answer_labels:
            col_idx += 1
            header_cell = ws_T_cls.cell(row=1, column=col_idx, value=ans)
            header_cell.font = bold_font
            header_cell.alignment = Alignment(wrap_text=True)
            seg = info["segments"][ans]
            ws_T_cls.cell(row=2, column=col_idx, value=seg["base"])
            col_map_cls_T[(q_col, ans)] = col_idx

        # Blank spacer column before next question
        col_idx += 2

    row = 5
    first_val_cT = last_val_cT = None

    # Body rows (elements)
    for cat_idx, cat_name in enumerate(sorted_categories):
        letter = letters[cat_idx]
        cat_cell = ws_T_cls.cell(row=row, column=2, value=f"{letter}. {cat_name}")
        cat_cell.font = bold_font
        cat_cell.fill = header_fill
        row += 1

        cols_in_cat = [c for c in element_cols if col_to_catname.get(c) == cat_name]
        for elt_idx, col in enumerate(cols_in_cat, start=1):
            code = f"{letter}{elt_idx}"
            elt_name = col_to_eltname.get(col, col)

            ws_T_cls.cell(row=row, column=1, value=code)
            ws_T_cls.cell(row=row, column=2, value=elt_name)

            # Write coefficients for each answer segment
            for q_col in classification_cols:
                info = class_groups_T.get(q_col)
                if not info:
                    continue
                for ans in info["answer_labels"]:
                    c_idx_ans = col_map_cls_T.get((q_col, ans))
                    if c_idx_ans is None:
                        continue
                    seg = info["segments"][ans]
                    val = int(seg["means"][col])
                    ws_T_cls.cell(row=row, column=c_idx_ans, value=val)

            if first_val_cT is None:
                first_val_cT = row
            last_val_cT = row
            row += 1

        row += 1

    wrap_col_B(ws_T_cls)
    set_classification_col_widths(ws_T_cls, start_col=4, max_width=20)

    # Conditional formatting per answer column (green)
    if coef_threshold_T is not None and first_val_cT is not None:
        for (q_col, ans), c_idx in col_map_cls_T.items():
            col_letter = get_column_letter(c_idx)
            formula = f'AND({col_letter}{first_val_cT}<>"",{col_letter}{first_val_cT}>={coef_threshold_T})'
            rule = FormulaRule(
                formula=[formula],
                fill=green_fill,
            )
            ws_T_cls.conditional_formatting.add(
                f"{col_letter}{first_val_cT}:{col_letter}{last_val_cT}",
                rule
            )

# ---------- (B) Classification Questions ----------
if class_groups_B:
    ws_B_cls = wb.create_sheet("(B) Classification Questions")
    ws_B_cls["B1"] = "Group (Binary Ratings)"; ws_B_cls["B1"].font = bold_font
    ws_B_cls["B2"] = "Base Size";             ws_B_cls["B2"].font = bold_font

    col_map_cls_B = {}
    col_idx = 4  # D

    for q_col in classification_cols:
        info = class_groups_B.get(q_col)
        if not info:
            continue
        qtext = info["question_text"]
        answer_labels = info["answer_labels"]

        # Question column
        q_cell = ws_B_cls.cell(row=1, column=col_idx, value=qtext)
        q_cell.font = bold_font
        q_cell.fill = header_fill
        q_cell.alignment = Alignment(wrap_text=True)

        # Answer columns
        for ans in answer_labels:
            col_idx += 1
            header_cell = ws_B_cls.cell(row=1, column=col_idx, value=ans)
            header_cell.font = bold_font
            header_cell.alignment = Alignment(wrap_text=True)
            seg = info["segments"][ans]
            ws_B_cls.cell(row=2, column=col_idx, value=seg["base"])
            col_map_cls_B[(q_col, ans)] = col_idx

        # Spacer column
        col_idx += 2

    row = 5
    first_val_cB = last_val_cB = None

    for cat_idx, cat_name in enumerate(sorted_categories):
        letter = letters[cat_idx]
        cat_cell = ws_B_cls.cell(row=row, column=2, value=f"{letter}. {cat_name}")
        cat_cell.font = bold_font
        cat_cell.fill = header_fill
        row += 1

        cols_in_cat = [c for c in element_cols if col_to_catname.get(c) == cat_name]
        for elt_idx, col in enumerate(cols_in_cat, start=1):
            code = f"{letter}{elt_idx}"
            elt_name = col_to_eltname.get(col, col)

            ws_B_cls.cell(row=row, column=1, value=code)
            ws_B_cls.cell(row=row, column=2, value=elt_name)

            for q_col in classification_cols:
                info = class_groups_B.get(q_col)
                if not info:
                    continue
                for ans in info["answer_labels"]:
                    c_idx_ans = col_map_cls_B.get((q_col, ans))
                    if c_idx_ans is None:
                        continue
                    seg = info["segments"][ans]
                    val = int(seg["means"][col])
                    ws_B_cls.cell(row=row, column=c_idx_ans, value=val)

            if first_val_cB is None:
                first_val_cB = row
            last_val_cB = row
            row += 1

        row += 1

    wrap_col_B(ws_B_cls)
    set_classification_col_widths(ws_B_cls, start_col=4, max_width=20)

    if coef_threshold_B is not None and first_val_cB is not None:
        for (q_col, ans), c_idx in col_map_cls_B.items():
            col_letter = get_column_letter(c_idx)
            formula = f'AND({col_letter}{first_val_cB}<>"",{col_letter}{first_val_cB}>={coef_threshold_B})'
            rule = FormulaRule(
                formula=[formula],
                fill=red_fill,
            )
            ws_B_cls.conditional_formatting.add(
                f"{col_letter}{first_val_cB}:{col_letter}{last_val_cB}",
                rule
            )

# ---------- (T) Combined ----------
ws_T_comb = wb.create_sheet("(T) Combined")
ws_T_comb["B1"] = "Group (Binary Ratings)"; ws_T_comb["B1"].font = bold_font
ws_T_comb["B2"] = "Base Size";             ws_T_comb["B2"].font = bold_font

segment_cols_T = {}
col_idx = 4  # D

# Overall
h = ws_T_comb.cell(row=1, column=col_idx, value="Overall")
h.font = bold_font
ws_T_comb.cell(row=2, column=col_idx, value=base_size)
segment_cols_T[("overall", None)] = col_idx
col_idx += 1

# Gender
for g_name in ["Male", "Female"]:
    if g_name in gender_groups_T:
        h = ws_T_comb.cell(row=1, column=col_idx, value=g_name)
        h.font = bold_font
        ws_T_comb.cell(row=2, column=col_idx, value=gender_groups_T[g_name]["base"])
        segment_cols_T[("gender", g_name)] = col_idx
        col_idx += 1

# Age
for age_label in AGE_BINS:
    if age_label in age_groups_T:
        h = ws_T_comb.cell(row=1, column=col_idx, value=age_label)
        h.font = bold_font
        ws_T_comb.cell(row=2, column=col_idx, value=age_groups_T[age_label]["base"])
        segment_cols_T[("age", age_label)] = col_idx
        col_idx += 1

# Classification – make it like the separate sheet:
# Question column (grey) then answer columns, then a spacer column.
for q_col in classification_cols:
    info = class_groups_T.get(q_col)
    if not info:
        continue
    qtext = info["question_text"]
    answer_labels = info["answer_labels"]

    # Question column
    q_cell = ws_T_comb.cell(row=1, column=col_idx, value=qtext)
    q_cell.font = bold_font
    q_cell.fill = header_fill
    q_cell.alignment = Alignment(wrap_text=True)
    # Base row left blank for the question column
    col_idx += 1

    # Answer columns
    for ans in answer_labels:
        seg = info["segments"][ans]
        h = ws_T_comb.cell(row=1, column=col_idx, value=ans)
        h.font = bold_font
        h.alignment = Alignment(wrap_text=True)
        ws_T_comb.cell(row=2, column=col_idx, value=seg["base"])
        segment_cols_T[("class", q_col, ans)] = col_idx
        col_idx += 1

    # Spacer column before next question
    col_idx += 1

row = 5
first_val_comb_T = last_val_comb_T = None

for cat_idx, cat_name in enumerate(sorted_categories):
    letter = letters[cat_idx]
    cat_cell = ws_T_comb.cell(row=row, column=2, value=f"{letter}. {cat_name}")
    cat_cell.font = bold_font
    cat_cell.fill = header_fill
    row += 1

    cols_in_cat = [c for c in element_cols if col_to_catname.get(c) == cat_name]
    for elt_idx, col in enumerate(cols_in_cat, start=1):
        code = f"{letter}{elt_idx}"
        elt_name = col_to_eltname.get(col, col)

        ws_T_comb.cell(row=row, column=1, value=code)
        ws_T_comb.cell(row=row, column=2, value=elt_name)

        # Overall
        c_idx = segment_cols_T.get(("overall", None))
        if c_idx is not None:
            ws_T_comb.cell(row=row, column=c_idx, value=int(element_means_T[col]))

        # Gender
        for g_name in ["Male", "Female"]:
            if g_name in gender_groups_T:
                c_idx = segment_cols_T.get(("gender", g_name))
                if c_idx is not None:
                    ws_T_comb.cell(
                        row=row,
                        column=c_idx,
                        value=int(gender_groups_T[g_name]["means"][col])
                    )

        # Age
        for age_label in AGE_BINS:
            if age_label in age_groups_T:
                c_idx = segment_cols_T.get(("age", age_label))
                if c_idx is not None:
                    ws_T_comb.cell(
                        row=row,
                        column=c_idx,
                        value=int(age_groups_T[age_label]["means"][col])
                    )

        # Classification
        for q_col in classification_cols:
            info = class_groups_T.get(q_col)
            if not info:
                continue
            for ans in info["answer_labels"]:
                key = ("class", q_col, ans)
                c_idx = segment_cols_T.get(key)
                if c_idx is None:
                    continue
                seg = info["segments"][ans]
                ws_T_comb.cell(
                    row=row,
                    column=c_idx,
                    value=seg["means"][col]
                )

        if first_val_comb_T is None:
            first_val_comb_T = row
        last_val_comb_T = row
        row += 1

    row += 1

wrap_col_B(ws_T_comb)
set_classification_col_widths(ws_T_comb, start_col=4, max_width=20)

if coef_threshold_T is not None and first_val_comb_T is not None:
    for key, c_idx in segment_cols_T.items():
        col_letter = get_column_letter(c_idx)
        formula = f'AND({col_letter}{first_val_comb_T}<>"",{col_letter}{first_val_comb_T}>={coef_threshold_T})'
        rule = FormulaRule(formula=[formula], fill=green_fill)
        ws_T_comb.conditional_formatting.add(
            f"{col_letter}{first_val_comb_T}:{col_letter}{last_val_comb_T}",
            rule
        )

# ---------- (B) Combined ----------
ws_B_comb = wb.create_sheet("(B) Combined")
ws_B_comb["B1"] = "Group (Binary Ratings)"; ws_B_comb["B1"].font = bold_font
ws_B_comb["B2"] = "Base Size";             ws_B_comb["B2"].font = bold_font

segment_cols_B = {}
col_idx = 4  # D

# Overall
h = ws_B_comb.cell(row=1, column=col_idx, value="Overall")
h.font = bold_font
ws_B_comb.cell(row=2, column=col_idx, value=base_size)
segment_cols_B[("overall", None)] = col_idx
col_idx += 1

# Gender
for g_name in ["Male", "Female"]:
    if g_name in gender_groups_B:
        h = ws_B_comb.cell(row=1, column=col_idx, value=g_name)
        h.font = bold_font
        ws_B_comb.cell(row=2, column=col_idx, value=gender_groups_B[g_name]["base"])
        segment_cols_B[("gender", g_name)] = col_idx
        col_idx += 1

# Age
for age_label in AGE_BINS:
    if age_label in age_groups_B:
        h = ws_B_comb.cell(row=1, column=col_idx, value=age_label)
        h.font = bold_font
        ws_B_comb.cell(row=2, column=col_idx, value=age_groups_B[age_label]["base"])
        segment_cols_B[("age", age_label)] = col_idx
        col_idx += 1

# Classification (match T Combined: question column + answers + spacer)
for q_col in classification_cols:
    info = class_groups_B.get(q_col)
    if not info:
        continue
    qtext = info["question_text"]
    answer_labels = info["answer_labels"]

    # Question column
    q_cell = ws_B_comb.cell(row=1, column=col_idx, value=qtext)
    q_cell.font = bold_font
    q_cell.fill = header_fill
    q_cell.alignment = Alignment(wrap_text=True)
    # Base row left blank for the question column
    col_idx += 1

    # Answer columns
    for ans in answer_labels:
        seg = info["segments"][ans]
        h = ws_B_comb.cell(row=1, column=col_idx, value=ans)
        h.font = bold_font
        h.alignment = Alignment(wrap_text=True)
        ws_B_comb.cell(row=2, column=col_idx, value=seg["base"])
        segment_cols_B[("class", q_col, ans)] = col_idx
        col_idx += 1

    # Spacer column before next question
    col_idx += 1

row = 5
first_val_comb_B = last_val_comb_B = None

for cat_idx, cat_name in enumerate(sorted_categories):
    letter = letters[cat_idx]
    cat_cell = ws_B_comb.cell(row=row, column=2, value=f"{letter}. {cat_name}")
    cat_cell.font = bold_font
    cat_cell.fill = header_fill
    row += 1

    cols_in_cat = [c for c in element_cols if col_to_catname.get(c) == cat_name]
    for elt_idx, col in enumerate(cols_in_cat, start=1):
        code = f"{letter}{elt_idx}"
        elt_name = col_to_eltname.get(col, col)

        ws_B_comb.cell(row=row, column=1, value=code)
        ws_B_comb.cell(row=row, column=2, value=elt_name)

        # Overall
        c_idx = segment_cols_B.get(("overall", None))
        if c_idx is not None:
            ws_B_comb.cell(row=row, column=c_idx, value=int(element_means_B[col]))

        # Gender
        for g_name in ["Male", "Female"]:
            if g_name in gender_groups_B:
                c_idx = segment_cols_B.get(("gender", g_name))
                if c_idx is not None:
                    ws_B_comb.cell(
                        row=row,
                        column=c_idx,
                        value=int(gender_groups_B[g_name]["means"][col])
                    )

        # Age
        for age_label in AGE_BINS:
            if age_label in age_groups_B:
                c_idx = segment_cols_B.get(("age", age_label))
                if c_idx is not None:
                    ws_B_comb.cell(
                        row=row,
                        column=c_idx,
                        value=int(age_groups_B[age_label]["means"][col])
                    )

        # Classification
        for q_col in classification_cols:
            info = class_groups_B.get(q_col)
            if not info:
                continue
            for ans in info["answer_labels"]:
                key = ("class", q_col, ans)
                c_idx = segment_cols_B.get(key)
                if c_idx is None:
                    continue
                seg = info["segments"][ans]
                ws_B_comb.cell(
                    row=row,
                    column=c_idx,
                    value=seg["means"][col]
                )

        if first_val_comb_B is None:
            first_val_comb_B = row
        last_val_comb_B = row
        row += 1

    row += 1

wrap_col_B(ws_B_comb)
set_classification_col_widths(ws_B_comb, start_col=4, max_width=20)

if coef_threshold_B is not None and first_val_comb_B is not None:
    for key, c_idx in segment_cols_B.items():
        col_letter = get_column_letter(c_idx)
        formula = f'AND({col_letter}{first_val_comb_B}<>"",{col_letter}{first_val_comb_B}>={coef_threshold_B})'
        rule = FormulaRule(formula=[formula], fill=red_fill)
        ws_B_comb.conditional_formatting.add(
            f"{col_letter}{first_val_comb_B}:{col_letter}{last_val_comb_B}",
            rule
        )

# ---------- Mindsets Generation Helper ----------
def create_mindset_sheet(wb, sheet_name, coef_df, element_cols, base_size, threshold, fill_color, labels_2, labels_3):
    ws = wb.create_sheet(sheet_name)
    ws["B1"] = "Group (Binary Ratings)"; ws["B1"].font = bold_font
    ws["B2"] = "Base Size";             ws["B2"].font = bold_font
    
    # Organize columns
    # Total | Mindset 1 of 2 | Mindset 2 of 2 | Mindset 1 of 3 | Mindset 2 of 3 | Mindset 3 of 3
    
    col_idx = 4 # D
    
    # Total
    ws.cell(row=1, column=col_idx, value="Total").font = bold_font
    ws.cell(row=2, column=col_idx, value=base_size)
    col_map = {("Total", None): col_idx}
    col_idx += 2 # Spacer
    
    # 2 Clusters
    counts_2 = np.bincount(labels_2, minlength=2)
    for i in range(2):
        name = f"Mindset {i+1} of 2"
        ws.cell(row=1, column=col_idx, value=name).font = bold_font
        ws.cell(row=1, column=col_idx).alignment = Alignment(wrap_text=True)
        ws.cell(row=2, column=col_idx, value=counts_2[i])
        col_map[("2", i)] = col_idx
        col_idx += 1
    col_idx += 1 # Spacer
    
    # 3 Clusters
    counts_3 = np.bincount(labels_3, minlength=3)
    for i in range(3):
        name = f"Mindset {i+1} of 3"
        ws.cell(row=1, column=col_idx, value=name).font = bold_font
        ws.cell(row=1, column=col_idx).alignment = Alignment(wrap_text=True)
        ws.cell(row=2, column=col_idx, value=counts_3[i])
        col_map[("3", i)] = col_idx
        col_idx += 1
        
    # Calculate means for each group
    # Total
    means_total = coef_df[element_cols].mean(axis=0)
    
    # 2 Clusters
    means_2 = []
    for i in range(2):
        sub = coef_df.iloc[labels_2 == i]
        if not sub.empty:
            means_2.append(sub[element_cols].mean(axis=0))
        else:
            means_2.append(pd.Series(0, index=element_cols))
            
    # 3 Clusters
    means_3 = []
    for i in range(3):
        sub = coef_df.iloc[labels_3 == i]
        if not sub.empty:
            means_3.append(sub[element_cols].mean(axis=0))
        else:
            means_3.append(pd.Series(0, index=element_cols))

    # Write Rows
    row = 5
    first_val = last_val = None
    
    for cat_idx, cat_name in enumerate(sorted_categories):
        letter = letters[cat_idx]
        ws.cell(row=row, column=2, value=f"{letter}. {cat_name}")
        cat_cell = ws.cell(row=row, column=2)
        cat_cell.font = bold_font
        cat_cell.fill = header_fill
        row += 1

        cols_in_cat = [c for c in element_cols if col_to_catname.get(c) == cat_name]
        for elt_idx, col in enumerate(cols_in_cat, start=1):
            code = f"{letter}{elt_idx}"
            elt_name = col_to_eltname.get(col, col)

            ws.cell(row=row, column=1, value=code)
            ws.cell(row=row, column=2, value=elt_name)
            
            # Determine if we should round (for T and B sheets, not R)
            should_round = sheet_name.startswith("(T)") or sheet_name.startswith("(B)")
            
            # Total
            c_idx = col_map[("Total", None)]
            val = means_total[col]
            ws.cell(row=row, column=c_idx, value=round(val) if should_round else val)
            
            # 2 Clusters
            for i in range(2):
                c_idx = col_map[("2", i)]
                val = means_2[i][col]
                ws.cell(row=row, column=c_idx, value=round(val) if should_round else val)
                
            # 3 Clusters
            for i in range(3):
                c_idx = col_map[("3", i)]
                val = means_3[i][col]
                ws.cell(row=row, column=c_idx, value=round(val) if should_round else val)

            if first_val is None:
                first_val = row
            last_val = row
            row += 1
        row += 1
        
    wrap_col_B(ws)
    set_classification_col_widths(ws, start_col=4, max_width=15)
    
    if threshold is not None and first_val is not None:
        # Apply formatting to all data columns
        for c_idx in col_map.values():
            col_letter = get_column_letter(c_idx)
            formula = f'AND({col_letter}{first_val}<>"",{col_letter}{first_val}>={threshold})'
            rule = FormulaRule(formula=[formula], fill=fill_color)
            ws.conditional_formatting.add(
                f"{col_letter}{first_val}:{col_letter}{last_val}",
                rule
            )

# ---------- Generate Mindset Sheets ----------
# 1. Run clustering on T coefficients
X_T = coef_table_T[element_cols].to_numpy(dtype=float)
labels_2_T = custom_kmeans_pearson(X_T, k=2, seed=101)
labels_3_T = custom_kmeans_pearson(X_T, k=3, seed=202)

# 2. Generate sheets using T labels
create_mindset_sheet(wb, "(T) Mindsets", coef_table_T, element_cols, base_size, coef_threshold_T, green_fill, labels_2_T, labels_3_T)
create_mindset_sheet(wb, "(B) Mindsets", coef_table_B, element_cols, base_size, coef_threshold_B, red_fill, labels_2_T, labels_3_T)
create_mindset_sheet(wb, "(R) Mindsets", coef_table_R, element_cols, base_size, coef_threshold_R, blue_fill, labels_2_T, labels_3_T)

# ---------- (R) Overall ----------
ws_R_overall = wb.create_sheet("(R) Overall")
ws_R_overall["B1"] = "Group (RTSeconds)"; ws_R_overall["B1"].font = bold_font
ws_R_overall["D1"] = "Total";                 ws_R_overall["D1"].font = bold_font
ws_R_overall["B2"] = "Base Size";             ws_R_overall["B2"].font = bold_font
ws_R_overall["D2"] = base_size

row = 5
first_val_R = last_val_R = None

for cat_idx, cat_name in enumerate(sorted_categories):
    letter = letters[cat_idx]
    ws_R_overall.cell(row=row, column=2, value=f"{letter}. {cat_name}")
    cat_cell = ws_R_overall.cell(row=row, column=2)
    cat_cell.font = bold_font
    cat_cell.fill = header_fill
    row += 1

    cols_in_cat = [c for c in element_cols if col_to_catname.get(c) == cat_name]
    for elt_idx, col in enumerate(cols_in_cat, start=1):
        code = f"{letter}{elt_idx}"
        elt_name = col_to_eltname.get(col, col)
        val = element_means_R[col]

        ws_R_overall.cell(row=row, column=1, value=code)
        ws_R_overall.cell(row=row, column=2, value=elt_name)
        ws_R_overall.cell(row=row, column=4, value=val)

        if first_val_R is None:
            first_val_R = row
        last_val_R = row
        row += 1
    row += 1

wrap_col_B(ws_R_overall)

if coef_threshold_R is not None and first_val_R is not None:
    formula_R = f'AND(D{first_val_R}<>"",D{first_val_R}>={coef_threshold_R})'
    rule_R = FormulaRule(formula=[formula_R], fill=blue_fill)
    ws_R_overall.conditional_formatting.add(f"D{first_val_R}:D{last_val_R}", rule_R)

# ---------- (R) Gender ----------
if gender_groups_R:
    ws_R_gender = wb.create_sheet("(R) Gender")
    ws_R_gender["B1"] = "Group (RTSeconds)"; ws_R_gender["B1"].font = bold_font
    ws_R_gender["B2"] = "Base Size";             ws_R_gender["B2"].font = bold_font

    col_map_R = {}
    col_idx = 4
    for g_name in ["Male", "Female"]:
        if g_name in gender_groups_R:
            ws_R_gender.cell(row=1, column=col_idx, value=g_name)
            ws_R_gender.cell(row=1, column=col_idx).font = bold_font
            ws_R_gender.cell(row=2, column=col_idx, value=gender_groups_R[g_name]["base"])
            col_map_R[g_name] = col_idx
            col_idx += 1

    row = 5
    first_val_gR = last_val_gR = None

    for cat_idx, cat_name in enumerate(sorted_categories):
        letter = letters[cat_idx]
        ws_R_gender.cell(row=row, column=2, value=f"{letter}. {cat_name}")
        cat_cell = ws_R_gender.cell(row=row, column=2)
        cat_cell.font = bold_font
        cat_cell.fill = header_fill
        row += 1

        cols_in_cat = [c for c in element_cols if col_to_catname.get(c) == cat_name]
        for elt_idx, col in enumerate(cols_in_cat, start=1):
            code = f"{letter}{elt_idx}"
            elt_name = col_to_eltname.get(col, col)

            ws_R_gender.cell(row=row, column=1, value=code)
            ws_R_gender.cell(row=row, column=2, value=elt_name)

            for g_name, info in gender_groups_R.items():
                c = col_map_R[g_name]
                ws_R_gender.cell(row=row, column=c, value=info["means"][col])

            if first_val_gR is None:
                first_val_gR = row
            last_val_gR = row
            row += 1
        row += 1

    wrap_col_B(ws_R_gender)

    if coef_threshold_R is not None and first_val_gR is not None:
        for g_name, c_idx in col_map_R.items():
            col_letter = ws_R_gender.cell(row=1, column=c_idx).column_letter
            formula = f'AND({col_letter}{first_val_gR}<>"",{col_letter}{first_val_gR}>={coef_threshold_R})'
            rule = FormulaRule(formula=[formula], fill=blue_fill)
            ws_R_gender.conditional_formatting.add(f"{col_letter}{first_val_gR}:{col_letter}{last_val_gR}", rule)

# ---------- (R) Age ----------
if age_groups_R:
    ws_R_age = wb.create_sheet("(R) Age")
    ws_R_age["B1"] = "Group (RTSeconds)"; ws_R_age["B1"].font = bold_font
    ws_R_age["B2"] = "Base Size";             ws_R_age["B2"].font = bold_font

    col_map_age_R = {}
    col_idx = 4
    for age_label in AGE_BINS:
        if age_label in age_groups_R:
            ws_R_age.cell(row=1, column=col_idx, value=age_label)
            ws_R_age.cell(row=1, column=col_idx).font = bold_font
            ws_R_age.cell(row=2, column=col_idx, value=age_groups_R[age_label]["base"])
            col_map_age_R[age_label] = col_idx
            col_idx += 1

    row = 5
    first_val_aR = last_val_aR = None

    for cat_idx, cat_name in enumerate(sorted_categories):
        letter = letters[cat_idx]
        ws_R_age.cell(row=row, column=2, value=f"{letter}. {cat_name}")
        cat_cell = ws_R_age.cell(row=row, column=2)
        cat_cell.font = bold_font
        cat_cell.fill = header_fill
        row += 1

        cols_in_cat = [c for c in element_cols if col_to_catname.get(c) == cat_name]
        for elt_idx, col in enumerate(cols_in_cat, start=1):
            code = f"{letter}{elt_idx}"
            elt_name = col_to_eltname.get(col, col)

            ws_R_age.cell(row=row, column=1, value=code)
            ws_R_age.cell(row=row, column=2, value=elt_name)

            for age_label, info in age_groups_R.items():
                c = col_map_age_R.get(age_label)
                if c is not None:
                    ws_R_age.cell(row=row, column=c, value=info["means"][col])

            if first_val_aR is None:
                first_val_aR = row
            last_val_aR = row
            row += 1
        row += 1

    wrap_col_B(ws_R_age)

    if coef_threshold_R is not None and first_val_aR is not None:
        for age_label, c_idx in col_map_age_R.items():
            col_letter = ws_R_age.cell(row=1, column=c_idx).column_letter
            formula = f'AND({col_letter}{first_val_aR}<>"",{col_letter}{first_val_aR}>={coef_threshold_R})'
            rule = FormulaRule(formula=[formula], fill=blue_fill)
            ws_R_age.conditional_formatting.add(f"{col_letter}{first_val_aR}:{col_letter}{last_val_aR}", rule)

# ---------- (R) Classification Questions ----------
if class_groups_R:
    ws_R_cls = wb.create_sheet("(R) Classification Questions")
    ws_R_cls["B1"] = "Group (RTSeconds)"; ws_R_cls["B1"].font = bold_font
    ws_R_cls["B2"] = "Base Size";             ws_R_cls["B2"].font = bold_font

    col_map_cls_R = {}
    col_idx = 4

    for q_col in classification_cols:
        info = class_groups_R.get(q_col)
        if not info:
            continue
        qtext = info["question_text"]
        answer_labels = info["answer_labels"]

        # Question column
        q_cell = ws_R_cls.cell(row=1, column=col_idx, value=qtext)
        q_cell.font = bold_font
        q_cell.fill = header_fill
        q_cell.alignment = Alignment(wrap_text=True)
        col_idx += 1

        # Answer columns
        for ans in answer_labels:
            header_cell = ws_R_cls.cell(row=1, column=col_idx, value=ans)
            header_cell.font = bold_font
            header_cell.alignment = Alignment(wrap_text=True)
            seg = info["segments"][ans]
            ws_R_cls.cell(row=2, column=col_idx, value=seg["base"])
            col_map_cls_R[(q_col, ans)] = col_idx
            col_idx += 1

        col_idx += 1

    row = 5
    first_val_cR = last_val_cR = None

    for cat_idx, cat_name in enumerate(sorted_categories):
        letter = letters[cat_idx]
        cat_cell = ws_R_cls.cell(row=row, column=2, value=f"{letter}. {cat_name}")
        cat_cell.font = bold_font
        cat_cell.fill = header_fill
        row += 1

        cols_in_cat = [c for c in element_cols if col_to_catname.get(c) == cat_name]
        for elt_idx, col in enumerate(cols_in_cat, start=1):
            code = f"{letter}{elt_idx}"
            elt_name = col_to_eltname.get(col, col)

            ws_R_cls.cell(row=row, column=1, value=code)
            ws_R_cls.cell(row=row, column=2, value=elt_name)

            for q_col in classification_cols:
                info = class_groups_R.get(q_col)
                if not info:
                    continue
                for ans in info["answer_labels"]:
                    c_idx_ans = col_map_cls_R.get((q_col, ans))
                    if c_idx_ans is None:
                        continue
                    seg = info["segments"][ans]
                    val = seg["means"][col]
                    ws_R_cls.cell(row=row, column=c_idx_ans, value=val)

            if first_val_cR is None:
                first_val_cR = row
            last_val_cR = row
            row += 1
        row += 1

    wrap_col_B(ws_R_cls)
    set_classification_col_widths(ws_R_cls, start_col=4, max_width=20)

    if coef_threshold_R is not None and first_val_cR is not None:
        for (q_col, ans), c_idx in col_map_cls_R.items():
            col_letter = get_column_letter(c_idx)
            formula = f'AND({col_letter}{first_val_cR}<>"",{col_letter}{first_val_cR}>={coef_threshold_R})'
            rule = FormulaRule(formula=[formula], fill=blue_fill)
            ws_R_cls.conditional_formatting.add(f"{col_letter}{first_val_cR}:{col_letter}{last_val_cR}", rule)

# ---------- (R) Combined ----------
ws_R_comb = wb.create_sheet("(R) Combined")
ws_R_comb["B1"] = "Group (RTSeconds)"; ws_R_comb["B1"].font = bold_font
ws_R_comb["B2"] = "Base Size";             ws_R_comb["B2"].font = bold_font

segment_cols_R = {}
col_idx = 4

# Overall
h = ws_R_comb.cell(row=1, column=col_idx, value="Overall")
h.font = bold_font
ws_R_comb.cell(row=2, column=col_idx, value=base_size)
segment_cols_R[("overall", None)] = col_idx
col_idx += 1

# Gender
for g_name in ["Male", "Female"]:
    if g_name in gender_groups_R:
        h = ws_R_comb.cell(row=1, column=col_idx, value=g_name)
        h.font = bold_font
        ws_R_comb.cell(row=2, column=col_idx, value=gender_groups_R[g_name]["base"])
        segment_cols_R[("gender", g_name)] = col_idx
        col_idx += 1

# Age
for age_label in AGE_BINS:
    if age_label in age_groups_R:
        h = ws_R_comb.cell(row=1, column=col_idx, value=age_label)
        h.font = bold_font
        ws_R_comb.cell(row=2, column=col_idx, value=age_groups_R[age_label]["base"])
        segment_cols_R[("age", age_label)] = col_idx
        col_idx += 1

# Classification
for q_col in classification_cols:
    info = class_groups_R.get(q_col)
    if not info:
        continue
    qtext = info["question_text"]
    answer_labels = info["answer_labels"]

    # Question column
    q_cell = ws_R_comb.cell(row=1, column=col_idx, value=qtext)
    q_cell.font = bold_font
    q_cell.fill = header_fill
    q_cell.alignment = Alignment(wrap_text=True)
    col_idx += 1

    # Answer columns
    for ans in answer_labels:
        seg = info["segments"][ans]
        h = ws_R_comb.cell(row=1, column=col_idx, value=ans)
        h.font = bold_font
        h.alignment = Alignment(wrap_text=True)
        ws_R_comb.cell(row=2, column=col_idx, value=seg["base"])
        segment_cols_R[("class", q_col, ans)] = col_idx
        col_idx += 1

    # Spacer
    col_idx += 1

row = 5
first_val_comb_R = last_val_comb_R = None

for cat_idx, cat_name in enumerate(sorted_categories):
    letter = letters[cat_idx]
    cat_cell = ws_R_comb.cell(row=row, column=2, value=f"{letter}. {cat_name}")
    cat_cell.font = bold_font
    cat_cell.fill = header_fill
    row += 1

    cols_in_cat = [c for c in element_cols if col_to_catname.get(c) == cat_name]
    for elt_idx, col in enumerate(cols_in_cat, start=1):
        code = f"{letter}{elt_idx}"
        elt_name = col_to_eltname.get(col, col)

        ws_R_comb.cell(row=row, column=1, value=code)
        ws_R_comb.cell(row=row, column=2, value=elt_name)

        # Overall
        c_idx = segment_cols_R.get(("overall", None))
        if c_idx is not None:
            ws_R_comb.cell(row=row, column=c_idx, value=element_means_R[col])

        # Gender
        for g_name in ["Male", "Female"]:
            if g_name in gender_groups_R:
                c_idx = segment_cols_R.get(("gender", g_name))
                if c_idx is not None:
                    ws_R_comb.cell(row=row, column=c_idx, value=gender_groups_R[g_name]["means"][col])

        # Age
        for age_label in AGE_BINS:
            if age_label in age_groups_R:
                c_idx = segment_cols_R.get(("age", age_label))
                if c_idx is not None:
                    ws_R_comb.cell(row=row, column=c_idx, value=age_groups_R[age_label]["means"][col])

        # Classification
        for q_col in classification_cols:
            info = class_groups_R.get(q_col)
            if not info:
                continue
            for ans in info["answer_labels"]:
                key = ("class", q_col, ans)
                c_idx = segment_cols_R.get(key)
                if c_idx is None:
                    continue
                seg = info["segments"][ans]
                ws_R_comb.cell(row=row, column=c_idx, value=seg["means"][col])

        if first_val_comb_R is None:
            first_val_comb_R = row
        last_val_comb_R = row
        row += 1
    row += 1

wrap_col_B(ws_R_comb)
if coef_threshold_R is not None and first_val_comb_R is not None:
    for key, c_idx in segment_cols_R.items():
        col_letter = get_column_letter(c_idx)
        formula = f'AND({col_letter}{first_val_comb_R}<>"",{col_letter}{first_val_comb_R}>={coef_threshold_R})'
        rule = FormulaRule(formula=[formula], fill=blue_fill)
        ws_R_comb.conditional_formatting.add(f"{col_letter}{first_val_comb_R}:{col_letter}{last_val_comb_R}", rule)


def create_raw_data_sheet(wb, csv_path):
    """
    Creates a 'RawData' sheet and pastes the exact content of the CSV.
    """
    if "RawData" in wb.sheetnames:
        del wb["RawData"]

    ws = wb.create_sheet("RawData")

    import csv
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        for row_idx, row in enumerate(reader, 1):
            for col_idx, value in enumerate(row, 1):
                # Try to convert to number if possible
                try:
                    # Try int first
                    if '.' not in value:
                        numeric_value = int(value)
                    else:
                        numeric_value = float(value)
                    ws.cell(row=row_idx, column=col_idx, value=numeric_value)
                except (ValueError, AttributeError):
                    # If conversion fails, keep as string
                    ws.cell(row=row_idx, column=col_idx, value=value)


# ---------- 11. Front Page ----------
create_front_page(wb, JSON_PATH)

# ---------- 12. Information Block ----------
create_info_block(wb, JSON_PATH)

# ---------- 13. Raw Data Sheet ----------
create_raw_data_sheet(wb, CSV_PATH)

# ---------- 14. Reorder Sheets ----------
# Exact order as specified
desired_order = [
    "Front Page",
    "Information Block",
    "RawData",
    "(T) Overall",
    "(B) Overall",
    "(R) Overall",
    "(T) Mindsets",
    "(B) Mindsets",
    "(R) Mindsets",
    "(T) Gender",
    "(B) Gender",
    "(R) Gender",
    "(T) Age",
    "(B) Age",
    "(R) Age",
    "(T) Classification Questions",
    "(B) Classification Questions",
    "(R) Classification Questions",
    "(T) Combined",
    "(B) Combined",
    "(R) Combined",
    "(T) Intercepts",
    "(B) Intercepts",
    "(R) Intercepts",
]

# Apply order
wb._sheets = [wb[s] for s in desired_order if s in wb.sheetnames]

wb.save(excel_path)
print(f"Saved Excel: {excel_path}")
