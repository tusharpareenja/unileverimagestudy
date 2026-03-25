"""
Celery jobs for background task processing.
- celery_job.generate_tasks: Task generation (grid, layer, hybrid)
- celery_job.simulate_synthetic_respondents: AI/random respondent simulation
"""
import logging
import random
import time
from typing import Dict, Any, Optional
from datetime import datetime

from app.celery_app import celery_app

# Ensure all SQLAlchemy models are loaded before DB operations
import app.models.user_model  # noqa: F401
import app.models.project_model  # noqa: F401
import app.models.study_model  # noqa: F401
import app.models.response_model  # noqa: F401
import app.models.job_model  # noqa: F401

logger = logging.getLogger(__name__)


def _update_job_progress(job_id: str, progress: float, message: str) -> None:
    """Update job progress in DB and notify WebSocket subscribers via Redis."""
    from app.db.session import SessionLocal
    from app.models.job_model import Job
    from app.websocket.job_notifier import job_progress_notifier

    db = SessionLocal()
    try:
        job = db.query(Job).filter(Job.job_id == job_id).first()
        if job and abs(job.progress - progress) > 1.0:
            job.progress = progress
            job.message = message
            db.commit()

            # Notify WebSocket subscribers (Redis pub/sub for cross-worker)
            job_progress_notifier.notify(
                job_id,
                {"type": "progress", "progress": progress, "message": message},
            )
    except Exception as e:
        logger.warning(f"Progress update failed (non-fatal): {e}")
        db.rollback()
    finally:
        try:
            db.close()
        except Exception:
            pass


def _generate_grid_tasks(
    job_id: str,
    payload: Dict,
    phase_type: Optional[str] = None,
    progress_range: tuple = (20.0, 90.0),
) -> Dict[str, Any]:
    """Generate grid-style tasks for a phase."""
    from app.services.task_generation_core import generate_grid_tasks_v2

    categories_data = []
    payload_categories = payload.get("categories") or []
    payload_elements = payload.get("elements") or []

    filtered_categories = payload_categories
    if phase_type:
        filtered_categories = [c for c in payload_categories if c.get("phase_type") == phase_type]

    if not filtered_categories:
        return {"tasks": {}, "metadata": {}}

    for cat in filtered_categories:
        cat_elements = [e for e in payload_elements if e.get("category_id") == cat.get("category_id")]
        categories_data.append(
            {
                "category_name": cat.get("name"),
                "elements": [
                    {
                        "element_id": str(el.get("element_id")),
                        "name": el.get("name"),
                        "content": el.get("content"),
                        "alt_text": el.get("alt_text") or el.get("name"),
                        "element_type": el.get("element_type"),
                    }
                    for el in cat_elements
                ],
            }
        )

    p_start, p_end = progress_range

    def on_progress(done: int, N: int) -> None:
        try:
            frac = max(0.0, min(1.0, done / max(1, N)))
            current_progress = p_start + ((p_end - p_start) * frac)
            phase_msg = f" ({phase_type})" if phase_type else ""
            _update_job_progress(job_id, current_progress, f"Building respondents{phase_msg} {done}/{N}...")
        except Exception:
            pass

    try:
        result = generate_grid_tasks_v2(
            categories_data=categories_data,
            number_of_respondents=payload.get("audience_segmentation", {}).get("number_of_respondents", 0),
            exposure_tolerance_cv=payload.get("exposure_tolerance_cv", 1.0),
            seed=payload.get("seed"),
            progress_callback=on_progress,
        )
        return result
    except RuntimeError as e:
        if "Preflight failed" in str(e) or "consider more elements" in str(e):
            raise RuntimeError(
                f"Study configuration error{' in ' + phase_type + ' phase' if phase_type else ''}: {str(e)}."
            )
        raise


def _generate_layer_tasks(job_id: str, payload: Dict) -> Dict[str, Any]:
    """Generate layer-style tasks."""
    from app.db.session import SessionLocal
    from app.models.study_model import Study
    from app.services.task_generation_core import generate_layer_tasks_v2

    _update_job_progress(job_id, 10.0, "Planning layer task generation...")

    layers = []
    for layer in payload.get("study_layers", []):
        layer_obj = {
            "name": layer.get("name", ""),
            "z_index": layer.get("z_index", 0),
            "order": layer.get("order", 0),
            "images": [{"name": img.get("name", ""), "url": img.get("url", "")} for img in layer.get("images", [])],
        }
        layers.append(layer_obj)

    _update_job_progress(job_id, 20.0, "Starting layer task generation...")

    p_start, p_end = 20.0, 90.0

    def on_progress(done: int, N: int) -> None:
        try:
            frac = max(0.0, min(1.0, done / max(1, N)))
            current_progress = p_start + ((p_end - p_start) * frac)
            _update_job_progress(job_id, current_progress, f"Building respondents {done}/{N}...")
        except Exception:
            pass

    try:
        result = generate_layer_tasks_v2(
            layers_data=layers,
            number_of_respondents=payload.get("audience_segmentation", {}).get("number_of_respondents", 0),
            exposure_tolerance_pct=payload.get("exposure_tolerance_pct", 2.0),
            seed=payload.get("seed"),
            progress_callback=on_progress,
        )
    except RuntimeError as e:
        if "Preflight failed" in str(e) or "consider more elements" in str(e):
            raise RuntimeError(f"Study configuration error: {str(e)}.")
        raise

    _update_job_progress(job_id, 90.0, "Saving generated tasks...")

    # Attach background image URL
    if isinstance(result, dict):
        try:
            study_id = payload.get("study_id")
            if study_id:
                db = SessionLocal()
                try:
                    from uuid import UUID
                    study = db.query(Study).filter(Study.id == UUID(str(study_id))).first()
                    if study and hasattr(study, "background_image_url"):
                        bg_url = getattr(study, "background_image_url", None)
                        if bg_url is not None:
                            meta = result.get("metadata", {})
                            if isinstance(meta, dict):
                                meta["background_image_url"] = bg_url
                                result["metadata"] = meta
                finally:
                    try:
                        db.close()
                    except Exception:
                        pass
        except Exception as e:
            logger.warning(f"Could not attach background image URL: {e}")

    return result


def _generate_hybrid_tasks(job_id: str, payload: Dict) -> Dict[str, Any]:
    """Generate hybrid (multi-phase) tasks."""
    phase_order = payload.get("phase_order", [])
    if not phase_order:
        raise ValueError("Hybrid study requires phase_order")

    _update_job_progress(job_id, 10.0, "Initializing hybrid task generation...")

    combined_tasks = {}
    combined_metadata = {"phases": []}

    is_mix = "mix" in phase_order
    payload_categories = payload.get("categories") or []

    target_phases = []
    if is_mix:
        target_phases = list(set(c.get("phase_type") for c in payload_categories if c.get("phase_type")))
    else:
        target_phases = phase_order

    num_phases = len(target_phases)
    tasks_per_respondent_mix = {}

    for i, phase_type in enumerate(target_phases):
        p_start = 10.0 + (i * (80.0 / num_phases))
        p_end = 10.0 + ((i + 1) * (80.0 / num_phases))

        _update_job_progress(job_id, p_start, f"Starting generation for phase: {phase_type}...")

        phase_result = _generate_grid_tasks(
            job_id, payload, phase_type=phase_type, progress_range=(p_start, p_end)
        )

        phase_tasks = phase_result.get("tasks", {})

        if is_mix:
            for resp_id, t_list in phase_tasks.items():
                if resp_id not in tasks_per_respondent_mix:
                    tasks_per_respondent_mix[resp_id] = []
                for t in t_list:
                    t["phase_type"] = phase_type
                    tasks_per_respondent_mix[resp_id].append(t)
        else:
            for resp_id, tasks in phase_tasks.items():
                if resp_id not in combined_tasks:
                    combined_tasks[resp_id] = []

                offset = len(combined_tasks[resp_id])
                for t in tasks:
                    t["task_index"] = int(t.get("task_index", 0)) + offset
                    t["phase_type"] = phase_type

                combined_tasks[resp_id].extend(tasks)

        combined_metadata["phases"].append(
            {"phase_type": phase_type, "metadata": phase_result.get("metadata", {})}
        )

    if is_mix:
        _update_job_progress(job_id, 85.0, "Randomizing mixed tasks...")
        for resp_id, t_list in tasks_per_respondent_mix.items():
            random.shuffle(t_list)
            for idx, t in enumerate(t_list):
                t["task_index"] = idx
            combined_tasks[resp_id] = t_list

    _update_job_progress(job_id, 90.0, "Combining all phases...")

    return {"tasks": combined_tasks, "metadata": combined_metadata}


def _save_results(job_id: str, study_id: str, user_id: str, payload: Dict, result: Dict[str, Any]) -> None:
    """Save generated tasks to the study."""
    from uuid import UUID
    from app.db.session import SessionLocal
    from app.models.study_model import Study, StudyMember

    db = SessionLocal()
    try:
        study = (
            db.query(Study)
            .filter(Study.id == UUID(study_id), Study.creator_id == UUID(user_id))
            .first()
        )

        if not study:
            stmt_member = db.query(Study).join(StudyMember).filter(
                Study.id == UUID(study_id), StudyMember.user_id == UUID(user_id)
            )
            study = stmt_member.first()

        if not study:
            raise ValueError(f"Study {study_id} not found or access denied")

        study.tasks = result.get("tasks", {})

        last_step_val = payload.get("last_step")
        if last_step_val is not None and isinstance(last_step_val, int):
            current_step = getattr(study, "last_step", 1) or 1
            if last_step_val > current_step:
                study.last_step = last_step_val

        db.commit()
        logger.info(f"Saved tasks for study {study_id}")
    finally:
        try:
            db.close()
        except Exception:
            pass


def _launch_study(job_id: str, study_id: str, user_id: str, payload: Dict) -> None:
    """Launch study: update title, background, etc. from payload."""
    try:
        from uuid import UUID
        from app.db.session import SessionLocal
        from app.services.study import update_study
        from app.schemas.study_schema import StudyUpdate

        study_updates = {}
        for field in [
            "title",
            "background",
            "language",
            "main_question",
            "orientation_text",
            "background_image_url",
            "rating_scale",
            "audience_segmentation",
            "phase_order",
        ]:
            if payload.get(field) is not None:
                study_updates[field] = payload.get(field)

        if study_updates:
            db = SessionLocal()
            try:
                update_payload = StudyUpdate(**study_updates)
                update_study(
                    db=db,
                    study_id=UUID(study_id),
                    owner_id=UUID(user_id),
                    payload=update_payload,
                )
            finally:
                try:
                    db.close()
                except Exception:
                    pass
    except Exception as e:
        logger.error(f"Failed to launch study {study_id}: {e}")


@celery_app.task(bind=True, name="celery_job.generate_tasks")
def generate_tasks_celery(
    self, job_id: str, study_id: str, user_id: str, payload: Dict[str, Any]
) -> Dict[str, Any]:
    """
    Celery job for generating study tasks in background.
    Runs in a separate worker process, freeing web workers.
    """
    if payload.get("seed") is None:
        payload = {**payload, "seed": int(time.time() * 1000) % (2**31)}

    # Ensure study_id is in payload for layer/background URL lookup
    payload = {**payload, "study_id": study_id}

    from app.db.session import SessionLocal
    from app.models.job_model import Job, JobStatus
    from app.websocket.job_notifier import job_progress_notifier

    db = SessionLocal()
    try:
        job = db.query(Job).filter(Job.job_id == job_id).first()
        if not job:
            logger.error(f"Job {job_id} not found")
            return {"success": False, "error": "Job not found"}

        job.status = JobStatus.PROCESSING
        job.message = "Generating tasks..."
        job.started_at = datetime.utcnow()
        db.commit()
        try:
            db.close()
        except Exception:
            pass

        study_type = payload.get("study_type")
        result = None

        if study_type in ("grid", "text"):
            result = _generate_grid_tasks(job_id, payload)
        elif study_type == "layer":
            result = _generate_layer_tasks(job_id, payload)
        elif study_type == "hybrid":
            result = _generate_hybrid_tasks(job_id, payload)
        else:
            raise ValueError(f"Unsupported study type: {study_type}")

        _save_results(job_id, study_id, user_id, payload, result)
        _launch_study(job_id, study_id, user_id, payload)

        db = SessionLocal()
        try:
            job = db.query(Job).filter(Job.job_id == job_id).first()
            if job:
                job.status = JobStatus.COMPLETED
                job.completed_at = datetime.utcnow()
                job.progress = 100.0
                job.message = "Task generation completed successfully"
                job.result = None
                db.commit()

            job_progress_notifier.notify(
                job_id,
                {
                    "type": "completed",
                    "progress": 100.0,
                    "message": "Task generation completed successfully",
                },
            )
            logger.info(f"Job {job_id} completed successfully")
        finally:
            try:
                db.close()
            except Exception:
                pass

        return {"success": True, "job_id": job_id}

    except RuntimeError as e:
        db = SessionLocal()
        try:
            job = db.query(Job).filter(Job.job_id == job_id).first()
            if job:
                job.status = JobStatus.FAILED
                job.completed_at = datetime.utcnow()
                job.error = str(e)
                job.message = f"Study configuration error: {str(e)}"
                db.commit()

            job_progress_notifier.notify(
                job_id,
                {"type": "failed", "error": str(e), "message": f"Study configuration error: {str(e)}"},
            )
            logger.error(f"Job {job_id} failed due to study configuration: {e}")
        finally:
            try:
                db.close()
            except Exception:
                pass
        return {"success": False, "error": str(e)}

    except Exception as e:
        import traceback

        db = SessionLocal()
        try:
            job = db.query(Job).filter(Job.job_id == job_id).first()
            if job:
                job.status = JobStatus.FAILED
                job.completed_at = datetime.utcnow()
                job.error = str(e)
                job.message = f"Task generation failed: {str(e)}"
                db.commit()

            job_progress_notifier.notify(
                job_id,
                {"type": "failed", "error": str(e), "message": f"Task generation failed: {str(e)}"},
            )
            logger.error(f"Job {job_id} failed: {e}")
            logger.error(f"Traceback: {traceback.format_exc()}")
        finally:
            try:
                db.close()
            except Exception:
                pass
        return {"success": False, "error": str(e)}


@celery_app.task(bind=True, name="celery_job.simulate_synthetic_respondents")
def simulate_synthetic_respondents_celery(
    self, job_id: str, study_id: str, user_id: str, payload: Dict[str, Any]
) -> Dict[str, Any]:
    """
    Celery job for simulating synthetic (AI or randomized) respondents.
    Runs in a separate worker process, freeing web workers.
    """
    from uuid import UUID
    from app.db.session import SessionLocal
    from app.models.job_model import Job, JobStatus
    from app.services.synthetic_simulation_service import run_simulation
    from app.websocket.job_notifier import job_progress_notifier

    db = SessionLocal()
    try:
        job = db.query(Job).filter(Job.job_id == job_id).first()
        if not job:
            logger.error(f"Job {job_id} not found")
            return {"success": False, "error": "Job not found"}

        job.status = JobStatus.PROCESSING
        job.message = "Simulating AI respondents..."
        job.started_at = datetime.utcnow()
        db.commit()
    finally:
        try:
            db.close()
        except Exception:
            pass

    def progress_callback(done: int, total: int, msg: str) -> None:
        try:
            pct = (100.0 * done / total) if total else 0
            _update_job_progress(job_id, pct, msg)
            logger.info(f"[Simulate AI respondents] {msg} — progress: {pct:.0f}% ({done}/{total})")

            job_progress_notifier.notify(
                job_id,
                {
                    "type": "progress",
                    "progress": pct,
                    "message": msg,
                    "respondents_completed": done,
                    "respondents_requested": total,
                },
            )
        except Exception as e:
            logger.warning(f"Progress callback error (non-fatal): {e}")

    try:
        db_sim = SessionLocal()
        try:
            result = run_simulation(
                db=db_sim,
                study_id=UUID(study_id),
                max_respondents=payload.get("max_respondents"),
                progress_callback=progress_callback,
                max_panelist_workers=payload.get("max_panelist_workers", 10),
                is_special_creator=payload.get("is_special_creator", False),
                randomize=payload.get("randomize", False),
            )
        finally:
            try:
                db_sim.close()
            except Exception:
                pass

        db = SessionLocal()
        try:
            job = db.query(Job).filter(Job.job_id == job_id).first()
            if job:
                job.completed_at = datetime.utcnow()
                job.progress = 100.0
                if result.get("success"):
                    job.status = JobStatus.COMPLETED
                    job.message = result.get("message", "Simulation completed.")
                    job.error = None
                else:
                    job.status = JobStatus.FAILED
                    job.message = result.get("message", "Simulation failed.")
                    job.error = result.get("error")
                db.commit()

                max_respondents = payload.get("max_respondents", 0)
                if result.get("success"):
                    job_progress_notifier.notify(
                        job_id,
                        {
                            "type": "completed",
                            "progress": 100.0,
                            "message": job.message,
                            "respondents_completed": max_respondents,
                            "respondents_requested": max_respondents,
                        },
                    )
                else:
                    job_progress_notifier.notify(
                        job_id,
                        {
                            "type": "failed",
                            "error": job.error,
                            "message": job.message,
                            "respondents_completed": 0,
                            "respondents_requested": max_respondents,
                        },
                    )

            logger.info(
                f"Job {job_id} simulate_synthetic_respondents finished: success={result.get('success')}, "
                f"message={result.get('message', '')}"
            )
        finally:
            try:
                db.close()
            except Exception:
                pass

        return {
            "success": result.get("success", False),
            "job_id": job_id,
            "respondents_simulated": result.get("respondents_simulated", 0),
            "message": result.get("message", ""),
        }

    except Exception as e:
        import traceback

        logger.error(f"Job {job_id} failed: {e}")
        logger.error(f"Traceback: {traceback.format_exc()}")

        db = SessionLocal()
        try:
            job = db.query(Job).filter(Job.job_id == job_id).first()
            if job:
                job.status = JobStatus.FAILED
                job.completed_at = datetime.utcnow()
                job.error = str(e)
                job.message = f"Simulation failed: {str(e)}"
                db.commit()

            job_progress_notifier.notify(
                job_id,
                {
                    "type": "failed",
                    "error": str(e),
                    "message": f"Simulation failed: {str(e)}",
                    "respondents_completed": 0,
                    "respondents_requested": payload.get("max_respondents", 0),
                },
            )
        finally:
            try:
                db.close()
            except Exception:
                pass
        return {"success": False, "error": str(e)}


@celery_app.task(bind=True, name="celery_job.export_project_zip")
def export_project_zip_celery(
    self, job_id: str, project_id: str, user_id: str
) -> Dict[str, Any]:
    """
    Celery job for exporting project as ZIP with per-study Excel reports.
    Uploads the ZIP to Azure Blob Storage and returns the download URL.
    """
    import io
    import zipfile
    import pandas as pd
    from uuid import UUID
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from app.db.session import SessionLocal
    from app.models.job_model import Job, JobStatus
    from app.services import project_service
    from app.services.response import StudyResponseService
    from app.services.analysis import StudyAnalysisService
    from app.api.v1.project import build_study_data_for_analysis
    from app.websocket.job_notifier import job_progress_notifier
    from app.core.config import settings
    from azure.storage.blob import BlobServiceClient, ContentSettings

    db = SessionLocal()
    try:
        job = db.query(Job).filter(Job.job_id == job_id).first()
        if not job:
            logger.error(f"Job {job_id} not found")
            return {"success": False, "error": "Job not found"}

        job.status = JobStatus.PROCESSING
        job.message = "Loading project studies..."
        job.started_at = datetime.utcnow()
        db.commit()
        
        job_progress_notifier.notify(
            job_id,
            {"type": "progress", "progress": 5, "message": "Loading project studies..."},
        )
    finally:
        try:
            db.close()
        except Exception:
            pass

    def _round6(val):
        if val is None or val == "":
            return ""
        try:
            return round(float(val), 6)
        except (TypeError, ValueError):
            return ""

    def _format_cell(v):
        if v is None or v == "":
            return ""
        if isinstance(v, (int, float)):
            return round(float(v), 6)
        try:
            return round(float(v), 6)
        except (TypeError, ValueError):
            return v

    try:
        db_main = SessionLocal()
        try:
            studies = project_service.get_project_studies_for_export(
                db=db_main,
                project_id=UUID(project_id),
                user_id=UUID(user_id),
            )
            
            # Build canonical headers
            all_key_names = []
            seen_keys = set()
            for s in studies:
                for k in (getattr(s, "product_keys", None) or []):
                    if isinstance(k, dict) and k.get("name") and k["name"] not in seen_keys:
                        seen_keys.add(k["name"])
                        all_key_names.append(k["name"])

            class_q_names = []
            if studies and getattr(studies[0], "classification_questions", None):
                for q in sorted(studies[0].classification_questions, key=lambda x: getattr(x, "order", 0)):
                    class_q_names.append(q.question_text or "")

            element_cols_ordered = []
            seen_el = set()
            for s in studies:
                study_data = build_study_data_for_analysis(s)
                for el in study_data.get("elements", []):
                    cat = el.get("category", {}) or {}
                    cat_name = cat.get("name", "")
                    el_name = el.get("name", "")
                    if not cat_name or not el_name:
                        continue
                    col = f"{cat_name}-{el_name}".replace("_", "-").replace(" ", "-")
                    if col not in seen_el:
                        seen_el.add(col)
                        element_cols_ordered.append(col)

            product_header = ["product_id"] + all_key_names + class_q_names + element_cols_ordered + ["Rating", "ResponseTime"]
            
            # Prepare study data for workers
            studies_data = []
            for i, s in enumerate(studies):
                study_data_dict = build_study_data_for_analysis(s)
                product_keys_list = getattr(s, "product_keys", None) or []
                if not isinstance(product_keys_list, list):
                    product_keys_list = []
                studies_data.append({
                    "study_id": s.id,
                    "product_id": getattr(s, "product_id", None),
                    "product_keys": product_keys_list,
                    "study_data": study_data_dict,
                    "index": i,
                })
        finally:
            try:
                db_main.close()
            except Exception:
                pass

        _update_job_progress(job_id, 10, f"Processing {len(studies)} studies...")
        job_progress_notifier.notify(
            job_id,
            {"type": "progress", "progress": 10, "message": f"Processing {len(studies)} studies..."},
        )

        def _worker(study_info: Dict[str, Any]) -> Dict[str, Any]:
            """Worker: process one study and return results."""
            study_id = study_info["study_id"]
            product_id_val = study_info["product_id"]
            product_keys_list = study_info["product_keys"]
            study_data_dict = study_info["study_data"]
            index = study_info["index"]
            
            xlsx_bytes = b""
            product_row = {"product_id": product_id_val or ""}
            for kn in all_key_names:
                product_row[kn] = ""
            for qn in class_q_names:
                product_row[qn] = ""
            for ec in element_cols_ordered:
                product_row[ec] = ""
            product_row["Rating"] = ""
            product_row["ResponseTime"] = ""
            raw_df = pd.DataFrame()
            db_session = None

            try:
                db_session = SessionLocal()
                response_svc = StudyResponseService(db_session)
                df = response_svc.get_study_dataframe(study_id, unilever_format=True)

                analysis_svc = StudyAnalysisService()
                try:
                    excel_io = analysis_svc.generate_report(df, study_data_dict)
                    xlsx_bytes = excel_io.getvalue() if excel_io else b""
                except Exception:
                    from openpyxl import Workbook
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Raw Data"
                    ws.append(list(df.columns) if not df.empty else ["product_id"])
                    buf = io.BytesIO()
                    wb.save(buf)
                    xlsx_bytes = buf.getvalue()

                t_scores = analysis_svc.get_t_overall_scores(df, study_data_dict) if not df.empty else {}

                key_map = {}
                for k in product_keys_list or []:
                    if isinstance(k, dict) and k.get("name") is not None:
                        pct = k.get("percentage")
                        key_map[k["name"]] = _round6(pct) if pct is not None else ""
                for kn in all_key_names:
                    product_row[kn] = key_map.get(kn, "")

                for qn in class_q_names:
                    if qn and not df.empty and qn in df.columns:
                        try:
                            product_row[qn] = _round6(float(df[qn].mean()))
                        except (TypeError, ValueError):
                            pass

                for ec in element_cols_ordered:
                    product_row[ec] = _round6(t_scores.get(ec, ""))

                if not df.empty and "Rating" in df.columns:
                    try:
                        product_row["Rating"] = _round6(float(df["Rating"].mean()))
                    except (TypeError, ValueError):
                        pass
                if not df.empty and "ResponseTime" in df.columns:
                    try:
                        product_row["ResponseTime"] = _round6(float(df["ResponseTime"].mean()))
                    except (TypeError, ValueError):
                        pass

                pid = product_id_val or "unknown"
                raw_df = df.copy()
                if "Product ID" in raw_df.columns:
                    raw_df = raw_df.drop(columns=["Product ID"])
                raw_df.insert(0, "product_id", pid)
            except Exception as e:
                logger.warning(f"Worker error for study {study_id}: {e}")
            finally:
                if db_session is not None:
                    try:
                        db_session.close()
                    except Exception:
                        pass

            return {
                "index": index,
                "study_id": str(study_id),
                "product_id": product_id_val or "",
                "xlsx_bytes": xlsx_bytes,
                "product_row": product_row,
                "raw_df": raw_df,
            }

        # Process studies with ThreadPoolExecutor
        if not studies_data:
            results = []
            raw_chunks = []
        else:
            max_workers = min(10, max(1, len(studies_data)))
            results = []
            raw_chunks = []
            BATCH_SIZE = 20
            current_batch = []
            completed = 0
            total_studies = len(studies_data)

            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                futures = {executor.submit(_worker, sd): sd["index"] for sd in studies_data}
                for fut in as_completed(futures):
                    try:
                        result = fut.result()
                        results.append(result)
                        raw_df = result.get("raw_df")
                        if raw_df is not None and not raw_df.empty:
                            current_batch.append(raw_df)
                            if len(current_batch) >= BATCH_SIZE:
                                raw_chunks.append(pd.concat(current_batch, join="outer", ignore_index=True))
                                current_batch = []
                        
                        completed += 1
                        progress = 10 + int(70 * completed / total_studies)
                        _update_job_progress(job_id, progress, f"Processed {completed}/{total_studies} studies...")
                        job_progress_notifier.notify(
                            job_id,
                            {"type": "progress", "progress": progress, "message": f"Processed {completed}/{total_studies} studies..."},
                        )
                    except Exception as e:
                        logger.warning(f"Future error: {e}")

                if current_batch:
                    raw_chunks.append(pd.concat(current_batch, join="outer", ignore_index=True))

        # Sort by index
        results.sort(key=lambda r: r["index"])

        _update_job_progress(job_id, 85, "Building ZIP file...")
        job_progress_notifier.notify(
            job_id,
            {"type": "progress", "progress": 85, "message": "Building ZIP file..."},
        )

        # Build filenames
        def _safe_filename(s: str) -> str:
            safe = "".join(c if c.isalnum() or c in "._-" else "_" for c in (s or ""))
            return safe.strip() or "unknown"

        pid_counts = {}
        index_to_filename = {}
        for r in results:
            base = _safe_filename(r["product_id"] or "unknown")
            pid_counts[base] = pid_counts.get(base, 0) + 1
            n = pid_counts[base]
            filename = f"{base}.xlsx" if n == 1 else f"{base}_{n}.xlsx"
            index_to_filename[r["index"]] = filename

        # Build ZIP
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for r in results:
                zf.writestr(index_to_filename[r["index"]], r["xlsx_bytes"])

            # Mega-sheet
            raw_combined = pd.concat(raw_chunks, join="outer", ignore_index=True) if raw_chunks else pd.DataFrame(columns=["product_id"])
            if not raw_combined.empty and "product_id" in raw_combined.columns:
                raw_combined["_sort_key"] = raw_combined["product_id"].fillna("").astype(str).str.strip()
                raw_combined.loc[raw_combined["_sort_key"] == "", "_sort_key"] = "zzzz"
                raw_combined["_order"] = range(len(raw_combined))
                raw_combined = raw_combined.sort_values(by=["_sort_key", "_order"], kind="mergesort").drop(
                    columns=["_sort_key", "_order"]
                ).reset_index(drop=True)

            from openpyxl import Workbook
            from openpyxl.utils.dataframe import dataframe_to_rows

            wb = Workbook()
            ws_raw = wb.active
            ws_raw.title = "Raw Data"
            if not raw_combined.empty:
                for row in dataframe_to_rows(raw_combined, index=False, header=True):
                    ws_raw.append(row)
            else:
                ws_raw.append(["product_id"])

            ws_product = wb.create_sheet("Product Data")
            ws_product.append(product_header)
            for r in results:
                row_dict = r["product_row"]
                cells = [_format_cell(row_dict.get(h, "")) for h in product_header]
                ws_product.append(cells)

            ws_range = wb.create_sheet("Range Sheet")
            ws_range.append(["Dependent Variable", "Low", "High", "Average"])
            product_rows = [r["product_row"] for r in results]
            range_cols = class_q_names + element_cols_ordered
            for col in range_cols:
                vals = []
                for row in product_rows:
                    v = row.get(col, "")
                    if v is None or v == "":
                        continue
                    try:
                        vals.append(float(v))
                    except (TypeError, ValueError):
                        continue
                if vals:
                    ws_range.append([col, round(min(vals), 6), round(max(vals), 6), round(sum(vals) / len(vals), 6)])
                else:
                    ws_range.append([col, "", "", ""])

            xl_buf = io.BytesIO()
            wb.save(xl_buf)
            xl_buf.seek(0)
            zf.writestr("mega_sheet.xlsx", xl_buf.getvalue())

        buf.seek(0)
        zip_bytes = buf.getvalue()

        _update_job_progress(job_id, 90, "Uploading to cloud storage...")
        job_progress_notifier.notify(
            job_id,
            {"type": "progress", "progress": 90, "message": "Uploading to cloud storage..."},
        )

        # Upload to Azure Blob Storage
        conn = settings.AZURE_STORAGE_CONNECTION_STRING
        if not conn:
            if settings.AZURE_STORAGE_ACCOUNT_NAME and settings.AZURE_STORAGE_ACCOUNT_KEY:
                endpoint = f"https://{settings.AZURE_STORAGE_ACCOUNT_NAME}.blob.core.windows.net"
                blob_service = BlobServiceClient(account_url=endpoint, credential=settings.AZURE_STORAGE_ACCOUNT_KEY)
            else:
                raise ValueError("Azure Storage not configured")
        else:
            blob_service = BlobServiceClient.from_connection_string(conn)

        container = settings.AZURE_STORAGE_CONTAINER or "exports"
        blob_name = f"exports/project_{project_id}_{job_id}.zip"
        content_settings = ContentSettings(content_type="application/zip")
        
        try:
            blob_client = blob_service.get_blob_client(container=container, blob=blob_name)
            blob_client.upload_blob(zip_bytes, overwrite=True, content_settings=content_settings)
        except Exception as e:
            logger.error(f"Failed to upload ZIP to Azure: {e}")
            raise ValueError(f"Failed to upload ZIP: {e}")

        # Build download URL
        account_name = settings.AZURE_STORAGE_ACCOUNT_NAME
        if not account_name and conn:
            import re
            m = re.search(r"AccountName=([^;]+)", conn)
            if m:
                account_name = m.group(1)
        
        download_url = f"https://{account_name}.blob.core.windows.net/{container}/{blob_name}"
        if settings.AZURE_STORAGE_SAS_TOKEN:
            token = settings.AZURE_STORAGE_SAS_TOKEN.lstrip('?')
            download_url = f"{download_url}?{token}"

        # Update job as completed and send email notification
        db_final = SessionLocal()
        try:
            from app.models.user_model import User
            from app.models.project_model import Project
            from app.services.email_service import email_service
            
            job = db_final.query(Job).filter(Job.job_id == job_id).first()
            if job:
                job.status = JobStatus.COMPLETED
                job.progress = 100.0
                job.message = "Export completed successfully"
                job.completed_at = datetime.utcnow()
                job.result = {"download_url": download_url, "filename": f"project_{project_id}_export.zip"}
                db_final.commit()

            # Get user and project info for email
            user = db_final.query(User).filter(User.id == UUID(user_id)).first()
            project = db_final.query(Project).filter(Project.id == UUID(project_id)).first()
            
            # Send email notification
            if user and user.email:
                user_name = user.full_name or user.email.split("@")[0]
                project_name = project.name if project else f"Project {project_id[:8]}"
                filename = f"{project_name.replace(' ', '_')}_export.zip"
                
                email_sent = email_service.send_export_ready_email(
                    to_email=user.email,
                    user_name=user_name,
                    project_name=project_name,
                    download_url=download_url,
                    filename=filename,
                )
                if email_sent:
                    logger.info(f"Export ready email sent to {user.email} for job {job_id}")
                else:
                    logger.warning(f"Failed to send export email to {user.email} for job {job_id}")

            job_progress_notifier.notify(
                job_id,
                {
                    "type": "completed",
                    "progress": 100,
                    "message": "Export completed successfully",
                    "download_url": download_url,
                },
            )
        finally:
            try:
                db_final.close()
            except Exception:
                pass

        logger.info(f"Job {job_id} export_project_zip completed successfully")
        return {"success": True, "job_id": job_id, "download_url": download_url}

    except Exception as e:
        import traceback
        logger.error(f"Job {job_id} failed: {e}")
        logger.error(f"Traceback: {traceback.format_exc()}")

        db_err = SessionLocal()
        try:
            job = db_err.query(Job).filter(Job.job_id == job_id).first()
            if job:
                job.status = JobStatus.FAILED
                job.completed_at = datetime.utcnow()
                job.error = str(e)
                job.message = f"Export failed: {str(e)}"
                db_err.commit()

            job_progress_notifier.notify(
                job_id,
                {"type": "failed", "error": str(e), "message": f"Export failed: {str(e)}"},
            )
        finally:
            try:
                db_err.close()
            except Exception:
                pass
        return {"success": False, "error": str(e)}
