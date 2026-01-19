import pandas as pd
import numpy as np
import json
import io
from datetime import datetime
from typing import Dict, Any, List, Optional, Tuple

import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter

class StudyAnalysisService:
    def __init__(self):
        self.rng = np.random.default_rng(123)
        
        # Constants
        self.PANEL_COL = "Panelist"
        self.RATING_COL = "Rating"
        self.RESPONSE_TIME_COL = "ResponseTime"
        self.GENDER_COL = "Gender"
        self.AGE_COL = "Age"
        self.TASK_COL = "Task"
        
        self.AGE_BINS = [
            "13-17", "18-24", "25-34", "35-44", 
            "45-54", "55-64", "65+"
        ]
        
        # Styles
        self.bold_font = Font(bold=True)
        self.header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        self.green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        self.red_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
        self.blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        self.MAX_WIDTH = 45
        self.letters = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"

    def generate_report(self, df: pd.DataFrame, study_data: Dict[str, Any]) -> io.BytesIO:
        """
        Generates the Excel report from the DataFrame and Study Data.
        Returns a BytesIO object containing the Excel file.
        """
        # 1. Preprocess Data
        # Handle NA values if not already handled
        # df is passed in, assuming it's already reasonably clean from response service
        # but we might need to handle specific NA representations if any
        
        # Extract metadata
        elements_json = study_data.get("elements", [])
        categories_json = study_data.get("categories", [])
        
        # 2. Element Metadata & Column Mapping
        element_meta = []
        for el in elements_json:
            cat_obj = el.get("category", {})
            # Fallback if category object is not fully populated but we have category_id
            if not cat_obj and el.get("category_id"):
                 # Find category in categories_json
                 found_cat = next((c for c in categories_json if c.get("id") == el.get("category_id")), {})
                 cat_obj = found_cat
            
            cat_name = cat_obj.get("name")
            el_name = el.get("name")
            
            if not cat_name or not el_name:
                continue
                
            # Construct expected column name (handling the mismatch)
            # analysis_v2 expected: f"{cat_name}_{el_name}"
            # response.py generates: f"{cat_name}-{el_name}".replace('_', '-').replace(' ', '-')
            
            # We need to find which column in df corresponds to this element
            # Let's try to match flexible
            
            # Expected "clean" name for internal logic
            internal_col_name = f"{cat_name}_{el_name}"
            
            # Try to find matching column in DF
            # 1. Try exact match (unlikely given the mismatch)
            # 2. Try hyphenated version
            # 3. Try hyphenated + safe replacements
            
            candidates = [
                internal_col_name,
                f"{cat_name}-{el_name}",
                f"{cat_name}-{el_name}".replace('_', '-').replace(' ', '-'),
                f"{cat_name}_{el_name}".replace(' ', '_')
            ]
            
            actual_col = None
            for cand in candidates:
                if cand in df.columns:
                    actual_col = cand
                    break
            
            if actual_col:
                element_meta.append({
                    "csv_col": actual_col, # Use the actual column name in DF
                    "category_name": cat_name,
                    "element_name": el_name,
                    "category_order": cat_obj.get("order", 0),
                })

        # Keep only columns that exist and deduplicate (some elements may have same name)
        element_cols_raw = [m["csv_col"] for m in element_meta]
        # Deduplicate while preserving order
        seen = set()
        element_cols = []
        for col in element_cols_raw:
            if col not in seen:
                element_cols.append(col)
                seen.add(col)
        
        # Maps
        col_to_catname = {m["csv_col"]: m["category_name"] for m in element_meta}
        col_to_eltname = {m["csv_col"]: m["element_name"] for m in element_meta}
        
        # Category ordering
        cat_order = {}
        for m in element_meta:
            name = m["category_name"]
            order = m["category_order"]
            if name not in cat_order or order < cat_order[name]:
                cat_order[name] = order
        sorted_categories = sorted(cat_order.keys(), key=lambda c: cat_order[c])

        # 3. Classification Columns
        # In response.py, classification columns are named by Question Text
        # We can identify them by excluding known columns
        known_cols = {self.PANEL_COL, self.RATING_COL, self.RESPONSE_TIME_COL, 
                      self.GENDER_COL, self.AGE_COL, self.TASK_COL}
        known_cols.update(element_cols)
        
        # Also exclude "session_id" etc if present
        classification_cols = []
        # Heuristic: Columns between Age and Task? Or just use study_data
        # Using study_data is safer
        class_qs = study_data.get("classification_questions", [])
        for q in class_qs:
            q_text = q.get("question_text")
            if q_text and q_text in df.columns:
                classification_cols.append(q_text)
                
        # 4. Run Analysis
        # 4a. Panel-level Regressions
        coef_table_T = self._run_panel_regressions(df, element_cols, "TOP")
        coef_table_B = self._run_panel_regressions(df, element_cols, "BOTTOM")
        coef_table_R = self._run_panel_regressions(df, element_cols, "RESPONSE")
        
        # 4b. Aggregations (Means & Groups)
        # Base Size
        base_size = df[self.PANEL_COL].nunique()
        
        # Means
        element_means_T = coef_table_T[element_cols].mean(axis=0).round().astype(int)
        element_means_B = coef_table_B[element_cols].mean(axis=0).round().astype(int)
        element_means_R = coef_table_R[element_cols].mean(axis=0) # Float for RT
        
        # Groups
        # Need to pass df to build groups because we need Gender/Age/Class info mapped to Panelist
        gender_groups_T = self._build_gender_groups(coef_table_T, df, element_cols)
        gender_groups_B = self._build_gender_groups(coef_table_B, df, element_cols)
        gender_groups_R = self._build_gender_groups(coef_table_R, df, element_cols)
        
        age_groups_T = self._build_age_groups(coef_table_T, df, element_cols)
        age_groups_B = self._build_age_groups(coef_table_B, df, element_cols)
        age_groups_R = self._build_age_groups(coef_table_R, df, element_cols)
        
        class_groups_T = self._build_class_groups(coef_table_T, df, element_cols, classification_cols)
        class_groups_B = self._build_class_groups(coef_table_B, df, element_cols, classification_cols)
        class_groups_R = self._build_class_groups(coef_table_R, df, element_cols, classification_cols)
        
        # 4c. Pooled Regressions (Intercepts)
        intercepts_T = self._run_pooled_regression(df, element_cols, "TOP")
        intercepts_B = self._run_pooled_regression(df, element_cols, "BOTTOM")
        intercepts_R = self._run_pooled_regression(df, element_cols, "RESPONSE")
        
        coef_threshold_T = intercepts_T.get("threshold")
        coef_threshold_B = intercepts_B.get("threshold")
        coef_threshold_R = intercepts_R.get("threshold")
        
        # 5. Excel Generation
        wb = Workbook()
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
            
        # 5a. Front Page & Info
        self._create_front_page(wb, study_data)
        self._create_info_block(wb, study_data)
        self._create_raw_data_sheet(wb, df)
        
        # 5b. Overall Sheets
        self._create_overall_sheet(wb, "(T) Overall", element_cols, sorted_categories, col_to_catname, col_to_eltname, 
                                   element_means_T, base_size, coef_threshold_T, self.green_fill, round_vals=True)
        self._create_overall_sheet(wb, "(B) Overall", element_cols, sorted_categories, col_to_catname, col_to_eltname, 
                                   element_means_B, base_size, coef_threshold_B, self.red_fill, round_vals=True)
        self._create_overall_sheet(wb, "(R) Overall", element_cols, sorted_categories, col_to_catname, col_to_eltname, 
                                   element_means_R, base_size, coef_threshold_R, self.blue_fill, round_vals=False)
                                   
        # 5c. Mindsets (Clustering)
        # Run clustering on T coefficients
        X_T = coef_table_T[element_cols].to_numpy(dtype=float)
        n_samples = X_T.shape[0]
        
        # Handle clustering based on available samples
        # Need at least k samples to create k clusters
        if n_samples >= 2:
            labels_2_T = self._custom_kmeans_pearson(X_T, k=2, seed=101)
        else:
            labels_2_T = np.zeros(n_samples, dtype=int)
            
        if n_samples >= 3:
            labels_3_T = self._custom_kmeans_pearson(X_T, k=3, seed=202)
        else:
            labels_3_T = np.zeros(n_samples, dtype=int)

        self._create_mindset_sheet(wb, "(T) Mindsets", coef_table_T, element_cols, sorted_categories, col_to_catname, col_to_eltname,
                                   base_size, coef_threshold_T, self.green_fill, labels_2_T, labels_3_T, round_vals=True)
        self._create_mindset_sheet(wb, "(B) Mindsets", coef_table_B, element_cols, sorted_categories, col_to_catname, col_to_eltname,
                                   base_size, coef_threshold_B, self.red_fill, labels_2_T, labels_3_T, round_vals=True)
        self._create_mindset_sheet(wb, "(R) Mindsets", coef_table_R, element_cols, sorted_categories, col_to_catname, col_to_eltname,
                                   base_size, coef_threshold_R, self.blue_fill, labels_2_T, labels_3_T, round_vals=False)

        # 5d. Gender Sheets
        self._create_segment_sheet(wb, "(T) Gender", element_cols, sorted_categories, col_to_catname, col_to_eltname,
                                   gender_groups_T, coef_threshold_T, self.green_fill, round_vals=True)
        self._create_segment_sheet(wb, "(B) Gender", element_cols, sorted_categories, col_to_catname, col_to_eltname,
                                   gender_groups_B, coef_threshold_B, self.red_fill, round_vals=True)
        self._create_segment_sheet(wb, "(R) Gender", element_cols, sorted_categories, col_to_catname, col_to_eltname,
                                   gender_groups_R, coef_threshold_R, self.blue_fill, round_vals=False)

        # 5e. Age Sheets
        self._create_segment_sheet(wb, "(T) Age", element_cols, sorted_categories, col_to_catname, col_to_eltname,
                                   age_groups_T, coef_threshold_T, self.green_fill, round_vals=True, segment_order=self.AGE_BINS)
        self._create_segment_sheet(wb, "(B) Age", element_cols, sorted_categories, col_to_catname, col_to_eltname,
                                   age_groups_B, coef_threshold_B, self.red_fill, round_vals=True, segment_order=self.AGE_BINS)
        self._create_segment_sheet(wb, "(R) Age", element_cols, sorted_categories, col_to_catname, col_to_eltname,
                                   age_groups_R, coef_threshold_R, self.blue_fill, round_vals=False, segment_order=self.AGE_BINS)

        # 5f. Classification Sheets
        self._create_classification_sheet(wb, "(T) Classification Questions", element_cols, sorted_categories, col_to_catname, col_to_eltname,
                                          class_groups_T, coef_threshold_T, self.green_fill, round_vals=True)
        self._create_classification_sheet(wb, "(B) Classification Questions", element_cols, sorted_categories, col_to_catname, col_to_eltname,
                                          class_groups_B, coef_threshold_B, self.red_fill, round_vals=True)
        self._create_classification_sheet(wb, "(R) Classification Questions", element_cols, sorted_categories, col_to_catname, col_to_eltname,
                                          class_groups_R, coef_threshold_R, self.blue_fill, round_vals=False)

        # 5g. Combined Sheets
        self._create_combined_sheet(wb, "(T) Combined", element_cols, sorted_categories, col_to_catname, col_to_eltname,
                                    base_size, element_means_T, gender_groups_T, age_groups_T, class_groups_T,
                                    coef_threshold_T, self.green_fill, round_vals=True)
        self._create_combined_sheet(wb, "(B) Combined", element_cols, sorted_categories, col_to_catname, col_to_eltname,
                                    base_size, element_means_B, gender_groups_B, age_groups_B, class_groups_B,
                                    coef_threshold_B, self.red_fill, round_vals=True)
        self._create_combined_sheet(wb, "(R) Combined", element_cols, sorted_categories, col_to_catname, col_to_eltname,
                                    base_size, element_means_R, gender_groups_R, age_groups_R, class_groups_R,
                                    coef_threshold_R, self.blue_fill, round_vals=False)

        # 5h. Intercepts Sheets
        self._create_intercepts_sheet(wb, "(T) Intercepts", intercepts_T["df"], coef_threshold_T, self.green_fill)
        self._create_intercepts_sheet(wb, "(B) Intercepts", intercepts_B["df"], coef_threshold_B, self.red_fill)
        self._create_intercepts_sheet(wb, "(R) Intercepts", intercepts_R["df"], coef_threshold_R, self.blue_fill)

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def generate_json_report(self, df: pd.DataFrame, study_data: Dict[str, Any]) -> Dict[str, Any]:
        """
        Generates the JSON report from the DataFrame and Study Data.
        Returns a dictionary with sheet names as keys and their data as values.
        """
        # 1. Preprocess Data (same as generate_report)
        elements_json = study_data.get("elements", [])
        categories_json = study_data.get("categories", [])
        
        # 2. Element Metadata & Column Mapping (same as generate_report)
        element_meta = []
        for el in elements_json:
            cat_obj = el.get("category", {})
            if not cat_obj and el.get("category_id"):
                found_cat = next((c for c in categories_json if c.get("id") == el.get("category_id")), {})
                cat_obj = found_cat
            
            cat_name = cat_obj.get("name")
            el_name = el.get("name")
            
            if not cat_name or not el_name:
                continue
                
            internal_col_name = f"{cat_name}_{el_name}"
            candidates = [
                internal_col_name,
                f"{cat_name}-{el_name}",
                f"{cat_name}-{el_name}".replace('_', '-').replace(' ', '-'),
                f"{cat_name}_{el_name}".replace(' ', '_')
            ]
            
            actual_col = None
            for cand in candidates:
                if cand in df.columns:
                    actual_col = cand
                    break
            
            if actual_col:
                element_meta.append({
                    "csv_col": actual_col,
                    "category_name": cat_name,
                    "element_name": el_name,
                    "category_order": cat_obj.get("order", 0),
                })

        element_cols_raw = [m["csv_col"] for m in element_meta]
        seen = set()
        element_cols = []
        for col in element_cols_raw:
            if col not in seen:
                element_cols.append(col)
                seen.add(col)
        
        col_to_catname = {m["csv_col"]: m["category_name"] for m in element_meta}
        col_to_eltname = {m["csv_col"]: m["element_name"] for m in element_meta}
        
        cat_order = {}
        for m in element_meta:
            name = m["category_name"]
            order = m["category_order"]
            if name not in cat_order or order < cat_order[name]:
                cat_order[name] = order
        sorted_categories = sorted(cat_order.keys(), key=lambda c: cat_order[c])

        # 3. Classification Columns
        known_cols = {self.PANEL_COL, self.RATING_COL, self.RESPONSE_TIME_COL, 
                      self.GENDER_COL, self.AGE_COL, self.TASK_COL}
        known_cols.update(element_cols)
        
        classification_cols = []
        class_qs = study_data.get("classification_questions", [])
        for q in class_qs:
            q_text = q.get("question_text")
            if q_text and q_text in df.columns:
                classification_cols.append(q_text)
                
        # 4. Run Analysis (same as generate_report)
        coef_table_T = self._run_panel_regressions(df, element_cols, "TOP")
        coef_table_B = self._run_panel_regressions(df, element_cols, "BOTTOM")
        coef_table_R = self._run_panel_regressions(df, element_cols, "RESPONSE")
        
        base_size = df[self.PANEL_COL].nunique()
        
        element_means_T = coef_table_T[element_cols].mean(axis=0).round().astype(int)
        element_means_B = coef_table_B[element_cols].mean(axis=0).round().astype(int)
        element_means_R = coef_table_R[element_cols].mean(axis=0)
        
        gender_groups_T = self._build_gender_groups(coef_table_T, df, element_cols)
        gender_groups_B = self._build_gender_groups(coef_table_B, df, element_cols)
        gender_groups_R = self._build_gender_groups(coef_table_R, df, element_cols)
        
        age_groups_T = self._build_age_groups(coef_table_T, df, element_cols)
        age_groups_B = self._build_age_groups(coef_table_B, df, element_cols)
        age_groups_R = self._build_age_groups(coef_table_R, df, element_cols)
        
        class_groups_T = self._build_class_groups(coef_table_T, df, element_cols, classification_cols)
        class_groups_B = self._build_class_groups(coef_table_B, df, element_cols, classification_cols)
        class_groups_R = self._build_class_groups(coef_table_R, df, element_cols, classification_cols)
        
        intercepts_T = self._run_pooled_regression(df, element_cols, "TOP")
        intercepts_B = self._run_pooled_regression(df, element_cols, "BOTTOM")
        intercepts_R = self._run_pooled_regression(df, element_cols, "RESPONSE")
        
        coef_threshold_T = intercepts_T.get("threshold")
        coef_threshold_B = intercepts_B.get("threshold")
        coef_threshold_R = intercepts_R.get("threshold")
        
        # Clustering for Mindsets
        X_T = coef_table_T[element_cols].to_numpy(dtype=float)
        n_samples = X_T.shape[0]
        
        if n_samples >= 2:
            labels_2_T = self._custom_kmeans_pearson(X_T, k=2, seed=101)
        else:
            labels_2_T = np.zeros(n_samples, dtype=int)
            
        if n_samples >= 3:
            labels_3_T = self._custom_kmeans_pearson(X_T, k=3, seed=202)
        else:
            labels_3_T = np.zeros(n_samples, dtype=int)
        
        # 5. Build JSON structure
        result = {}
        
        # 5a. Front Page
        result["Front Page"] = {
            "Title": study_data.get("title", ""),
            "Background": study_data.get("background", ""),
            "Language": study_data.get("language", ""),
            "Launched At": study_data.get("launched_at", "")
        }
        
        # 5b. Information Block
        info_block = {
            "Study Title": study_data.get("title", ""),
            "Study Type": study_data.get("study_type", ""),
            "Study Background": study_data.get("background", ""),
            "Categories": []
        }
        
        categories = study_data.get("categories", [])
        elements = study_data.get("elements", [])
        for cat in categories:
            cat_name = cat.get("name", "")
            cat_id = cat.get("id")
            cat_info = {
                "name": cat_name,
                "elements": []
            }
            c_elements = [e for e in elements if e.get("category_id") == cat_id]
            for el in c_elements:
                cat_info["elements"].append({
                    "name": el.get("name", ""),
                    "content": el.get("content", "")
                })
            info_block["Categories"].append(cat_info)
        
        result["Information Block"] = info_block
        
        # 5c. RawData - Convert DataFrame to JSON-serializable format
        raw_data_list = []
        for _, row in df.iterrows():
            raw_row = {}
            for col in df.columns:
                val = row[col]
                if pd.isna(val):
                    raw_row[col] = None
                elif isinstance(val, (np.integer, np.int64)):
                    raw_row[col] = int(val)
                elif isinstance(val, (np.floating, np.float64)):
                    raw_row[col] = float(val)
                elif isinstance(val, pd.Timestamp):
                    raw_row[col] = val.isoformat()
                else:
                    raw_row[col] = val
            raw_data_list.append(raw_row)
        result["RawData"] = raw_data_list
        
        # 5d. Overall Sheets
        result["(T) Overall"] = self._build_overall_json(
            element_cols, sorted_categories, col_to_catname, col_to_eltname,
            element_means_T, base_size, coef_threshold_T, round_vals=True
        )
        result["(B) Overall"] = self._build_overall_json(
            element_cols, sorted_categories, col_to_catname, col_to_eltname,
            element_means_B, base_size, coef_threshold_B, round_vals=True
        )
        result["(R) Overall"] = self._build_overall_json(
            element_cols, sorted_categories, col_to_catname, col_to_eltname,
            element_means_R, base_size, coef_threshold_R, round_vals=False
        )
        
        # 5e. Mindsets Sheets
        result["(T) Mindsets"] = self._build_mindset_json(
            coef_table_T, element_cols, sorted_categories, col_to_catname, col_to_eltname,
            base_size, coef_threshold_T, labels_2_T, labels_3_T, round_vals=True
        )
        result["(B) Mindsets"] = self._build_mindset_json(
            coef_table_B, element_cols, sorted_categories, col_to_catname, col_to_eltname,
            base_size, coef_threshold_B, labels_2_T, labels_3_T, round_vals=True
        )
        result["(R) Mindsets"] = self._build_mindset_json(
            coef_table_R, element_cols, sorted_categories, col_to_catname, col_to_eltname,
            base_size, coef_threshold_R, labels_2_T, labels_3_T, round_vals=False
        )
        
        # 5f. Gender Sheets
        result["(T) Gender"] = self._build_segment_json(
            element_cols, sorted_categories, col_to_catname, col_to_eltname,
            gender_groups_T, coef_threshold_T, round_vals=True
        )
        result["(B) Gender"] = self._build_segment_json(
            element_cols, sorted_categories, col_to_catname, col_to_eltname,
            gender_groups_B, coef_threshold_B, round_vals=True
        )
        result["(R) Gender"] = self._build_segment_json(
            element_cols, sorted_categories, col_to_catname, col_to_eltname,
            gender_groups_R, coef_threshold_R, round_vals=False
        )
        
        # 5g. Age Sheets
        result["(T) Age"] = self._build_segment_json(
            element_cols, sorted_categories, col_to_catname, col_to_eltname,
            age_groups_T, coef_threshold_T, round_vals=True, segment_order=self.AGE_BINS
        )
        result["(B) Age"] = self._build_segment_json(
            element_cols, sorted_categories, col_to_catname, col_to_eltname,
            age_groups_B, coef_threshold_B, round_vals=True, segment_order=self.AGE_BINS
        )
        result["(R) Age"] = self._build_segment_json(
            element_cols, sorted_categories, col_to_catname, col_to_eltname,
            age_groups_R, coef_threshold_R, round_vals=False, segment_order=self.AGE_BINS
        )
        
        # 5h. Classification Sheets
        result["(T) Classification Questions"] = self._build_classification_json(
            element_cols, sorted_categories, col_to_catname, col_to_eltname,
            class_groups_T, coef_threshold_T, round_vals=True
        )
        result["(B) Classification Questions"] = self._build_classification_json(
            element_cols, sorted_categories, col_to_catname, col_to_eltname,
            class_groups_B, coef_threshold_B, round_vals=True
        )
        result["(R) Classification Questions"] = self._build_classification_json(
            element_cols, sorted_categories, col_to_catname, col_to_eltname,
            class_groups_R, coef_threshold_R, round_vals=False
        )
        
        # 5i. Combined Sheets
        result["(T) Combined"] = self._build_combined_json(
            element_cols, sorted_categories, col_to_catname, col_to_eltname,
            base_size, element_means_T, gender_groups_T, age_groups_T, class_groups_T,
            coef_threshold_T, round_vals=True
        )
        result["(B) Combined"] = self._build_combined_json(
            element_cols, sorted_categories, col_to_catname, col_to_eltname,
            base_size, element_means_B, gender_groups_B, age_groups_B, class_groups_B,
            coef_threshold_B, round_vals=True
        )
        result["(R) Combined"] = self._build_combined_json(
            element_cols, sorted_categories, col_to_catname, col_to_eltname,
            base_size, element_means_R, gender_groups_R, age_groups_R, class_groups_R,
            coef_threshold_R, round_vals=False
        )
        
        # 5j. Intercepts Sheets
        result["(T) Intercepts"] = self._build_intercepts_json(intercepts_T["df"], coef_threshold_T)
        result["(B) Intercepts"] = self._build_intercepts_json(intercepts_B["df"], coef_threshold_B)
        result["(R) Intercepts"] = self._build_intercepts_json(intercepts_R["df"], coef_threshold_R)
        
        return result

    # --- JSON Builder Helpers ---
    def _build_overall_json(self, element_cols, sorted_cats, col_to_cat, col_to_elt, means, base, threshold, round_vals):
        result = {
            "base_size": int(base),
            "threshold": float(threshold) if threshold is not None else None,
            "categories": []
        }
        
        for i, cat_name in enumerate(sorted_cats):
            letter = self.letters[i]
            cat_data = {
                "code": letter,
                "name": cat_name,
                "elements": []
            }
            
            cols = [c for c in element_cols if col_to_cat.get(c) == cat_name]
            for j, col in enumerate(cols, 1):
                code = f"{letter}{j}"
                val = means[col]
                if round_vals:
                    val = int(val)
                else:
                    val = float(val)
                
                element_data = {
                    "code": code,
                    "name": col_to_elt.get(col, col),
                    "value": val,
                    "above_threshold": threshold is not None and val >= threshold
                }
                cat_data["elements"].append(element_data)
            
            result["categories"].append(cat_data)
        
        return result

    def _build_mindset_json(self, coef_df, element_cols, sorted_cats, col_to_cat, col_to_elt, base, threshold, l2, l3, round_vals):
        counts_2 = np.bincount(l2, minlength=2).tolist()
        counts_3 = np.bincount(l3, minlength=3).tolist()
        
        means_total = coef_df[element_cols].mean(axis=0)
        means_2 = [
            coef_df.iloc[l2 == i][element_cols].mean(axis=0) if np.any(l2 == i) 
            else pd.Series(0, index=element_cols) 
            for i in range(2)
        ]
        means_3 = [
            coef_df.iloc[l3 == i][element_cols].mean(axis=0) if np.any(l3 == i) 
            else pd.Series(0, index=element_cols) 
            for i in range(3)
        ]
        
        result = {
            "base_size": int(base),
            "threshold": float(threshold) if threshold is not None else None,
            "groups": {
                "Total": {"base_size": int(base)},
                "Mindset_2": {
                    f"Mindset_{i+1}_of_2": {"base_size": int(counts_2[i])} 
                    for i in range(2)
                },
                "Mindset_3": {
                    f"Mindset_{i+1}_of_3": {"base_size": int(counts_3[i])} 
                    for i in range(3)
                }
            },
            "categories": []
        }
        
        for i, cat_name in enumerate(sorted_cats):
            letter = self.letters[i]
            cat_data = {
                "code": letter,
                "name": cat_name,
                "elements": []
            }
            
            cols = [c for c in element_cols if col_to_cat.get(c) == cat_name]
            for j, col in enumerate(cols, 1):
                code = f"{letter}{j}"
                element_name = col_to_elt.get(col, col)
                
                def get_val(mean_series):
                    val = mean_series[col]
                    if round_vals:
                        return int(val)
                    return float(val)
                
                element_data = {
                    "code": code,
                    "name": element_name,
                    "values": {
                        "Total": get_val(means_total),
                        "Mindset_1_of_2": get_val(means_2[0]),
                        "Mindset_2_of_2": get_val(means_2[1]),
                        "Mindset_1_of_3": get_val(means_3[0]),
                        "Mindset_2_of_3": get_val(means_3[1]),
                        "Mindset_3_of_3": get_val(means_3[2])
                    },
                    "above_threshold": {}
                }
                
                # Check threshold for each value
                for key, val in element_data["values"].items():
                    element_data["above_threshold"][key] = threshold is not None and val >= threshold
                
                cat_data["elements"].append(element_data)
            
            result["categories"].append(cat_data)
        
        return result

    def _build_segment_json(self, element_cols, sorted_cats, col_to_cat, col_to_elt, groups, threshold, round_vals, segment_order=None):
        if not groups:
            return {"base_size": 0, "threshold": float(threshold) if threshold is not None else None, "segments": {}, "categories": []}
        
        result = {
            "threshold": float(threshold) if threshold is not None else None,
            "segments": {},
            "categories": []
        }
        
        keys = segment_order if segment_order else sorted(groups.keys())
        
        for k in keys:
            if k in groups:
                result["segments"][k] = {
                    "base_size": groups[k]["base"]
                }
        
        for i, cat_name in enumerate(sorted_cats):
            letter = self.letters[i]
            cat_data = {
                "code": letter,
                "name": cat_name,
                "elements": []
            }
            
            cols = [c for c in element_cols if col_to_cat.get(c) == cat_name]
            for j, col in enumerate(cols, 1):
                code = f"{letter}{j}"
                element_name = col_to_elt.get(col, col)
                
                element_data = {
                    "code": code,
                    "name": element_name,
                    "values": {},
                    "above_threshold": {}
                }
                
                for k in keys:
                    if k in groups:
                        val = groups[k]["means"][col]
                        if round_vals:
                            val = int(round(val))
                        else:
                            val = float(val)
                        element_data["values"][k] = val
                        element_data["above_threshold"][k] = threshold is not None and val >= threshold
                
                cat_data["elements"].append(element_data)
            
            result["categories"].append(cat_data)
        
        return result

    def _build_classification_json(self, element_cols, sorted_cats, col_to_cat, col_to_elt, groups, threshold, round_vals):
        if not groups:
            return {"threshold": float(threshold) if threshold is not None else None, "questions": [], "categories": []}
        
        result = {
            "threshold": float(threshold) if threshold is not None else None,
            "questions": [],
            "categories": []
        }
        
        for q_col, info in groups.items():
            question_data = {
                "question_text": info["question_text"],
                "segments": {}
            }
            
            for ans in info["answer_labels"]:
                question_data["segments"][ans] = {
                    "base_size": info["segments"][ans]["base"]
                }
            
            result["questions"].append(question_data)
        
        for i, cat_name in enumerate(sorted_cats):
            letter = self.letters[i]
            cat_data = {
                "code": letter,
                "name": cat_name,
                "elements": []
            }
            
            cols = [c for c in element_cols if col_to_cat.get(c) == cat_name]
            for j, col in enumerate(cols, 1):
                code = f"{letter}{j}"
                element_name = col_to_elt.get(col, col)
                
                element_data = {
                    "code": code,
                    "name": element_name,
                    "values": {}
                }
                
                for q_col, info in groups.items():
                    for ans in info["answer_labels"]:
                        key = f"{q_col}::{ans}"
                        val = info["segments"][ans]["means"][col]
                        if round_vals:
                            val = int(round(val))
                        else:
                            val = float(val)
                        element_data["values"][key] = {
                            "value": val,
                            "above_threshold": threshold is not None and val >= threshold
                        }
                
                cat_data["elements"].append(element_data)
            
            result["categories"].append(cat_data)
        
        return result

    def _build_combined_json(self, element_cols, sorted_cats, col_to_cat, col_to_elt, base, means, g_groups, a_groups, c_groups, threshold, round_vals):
        result = {
            "base_size": int(base),
            "threshold": float(threshold) if threshold is not None else None,
            "segments": {
                "Overall": {"base_size": int(base)},
                "Gender": {},
                "Age": {},
                "Classification": {}
            },
            "categories": []
        }
        
        # Gender segments
        for g in ["Male", "Female"]:
            if g in g_groups:
                result["segments"]["Gender"][g] = {"base_size": g_groups[g]["base"]}
        
        # Age segments
        for a in self.AGE_BINS:
            if a in a_groups:
                result["segments"]["Age"][a] = {"base_size": a_groups[a]["base"]}
        
        # Classification segments
        for q_col, info in c_groups.items():
            result["segments"]["Classification"][info["question_text"]] = {
                "base_size": 0,
                "answers": {}
            }
            for ans in info["answer_labels"]:
                result["segments"]["Classification"][info["question_text"]]["answers"][ans] = {
                    "base_size": info["segments"][ans]["base"]
                }
        
        for i, cat_name in enumerate(sorted_cats):
            letter = self.letters[i]
            cat_data = {
                "code": letter,
                "name": cat_name,
                "elements": []
            }
            
            cols = [c for c in element_cols if col_to_cat.get(c) == cat_name]
            for j, col in enumerate(cols, 1):
                code = f"{letter}{j}"
                element_name = col_to_elt.get(col, col)
                
                def get_val(v):
                    if round_vals:
                        return int(round(v))
                    return float(v)
                
                element_data = {
                    "code": code,
                    "name": element_name,
                    "values": {
                        "Overall": get_val(means[col])
                    },
                    "above_threshold": {
                        "Overall": threshold is not None and get_val(means[col]) >= threshold
                    }
                }
                
                # Gender values
                for g in ["Male", "Female"]:
                    if g in g_groups:
                        val = get_val(g_groups[g]["means"][col])
                        element_data["values"][f"Gender::{g}"] = val
                        element_data["above_threshold"][f"Gender::{g}"] = threshold is not None and val >= threshold
                
                # Age values
                for a in self.AGE_BINS:
                    if a in a_groups:
                        val = get_val(a_groups[a]["means"][col])
                        element_data["values"][f"Age::{a}"] = val
                        element_data["above_threshold"][f"Age::{a}"] = threshold is not None and val >= threshold
                
                # Classification values
                for q_col, info in c_groups.items():
                    for ans in info["answer_labels"]:
                        val = get_val(info["segments"][ans]["means"][col])
                        key = f"Classification::{q_col}::{ans}"
                        element_data["values"][key] = val
                        element_data["above_threshold"][key] = threshold is not None and val >= threshold
                
                cat_data["elements"].append(element_data)
            
            result["categories"].append(cat_data)
        
        return result

    def _build_intercepts_json(self, df, threshold):
        result = {
            "threshold": float(threshold) if threshold is not None else None,
            "data": []
        }
        
        for _, row in df.iterrows():
            row_data = {
                "element": str(row["element"]),
                "beta_no_intercept": float(row["beta_no_intercept"]),
                "beta_with_intercept": float(row["beta_with_intercept"]),
                "t_with_intercept": float(row["t_with_intercept"]),
                "t_above_2": float(row["t_with_intercept"]) >= 2.0
            }
            result["data"].append(row_data)
        
        return result

    # --- Regression Helpers ---
    def _run_panel_regressions(self, df: pd.DataFrame, element_cols: List[str], mode: str) -> pd.DataFrame:
        rows = []
        for pid, g in df.groupby(self.PANEL_COL):
            X = g[element_cols].to_numpy(dtype=float)
            
            if mode == "TOP":
                ratings = g[self.RATING_COL].to_numpy()
                Y = np.where(ratings >= 4, 100.0, 0.0)
                Y = Y + self.rng.uniform(-0.5, 0.5, size=Y.shape) * 1e-5
            elif mode == "BOTTOM":
                ratings = g[self.RATING_COL].to_numpy()
                Y = np.where(ratings <= 2, 100.0, 0.0)
                Y = Y + self.rng.uniform(-0.5, 0.5, size=Y.shape) * 1e-5
            else: # RESPONSE
                # Cap at 7s
                rt = g[self.RESPONSE_TIME_COL].clip(upper=7.0).to_numpy()
                Y = rt

            # Regress
            beta, _, _, _ = np.linalg.lstsq(X, Y, rcond=None)
            
            # Calc R2
            Y_hat = X @ beta
            sse = float(np.sum((Y - Y_hat) ** 2))
            sst = float(np.sum(Y ** 2))
            r2 = np.nan if sst == 0 else 1.0 - sse / sst
            
            row = {"Panelist": pid, f"R2_{mode}": r2}
            for col_name, b in zip(element_cols, beta):
                row[col_name] = b
            rows.append(row)
            
        return pd.DataFrame(rows)

    def _run_pooled_regression(self, df: pd.DataFrame, element_cols: List[str], mode: str) -> Dict[str, Any]:
        X_all = df[element_cols].to_numpy(dtype=float)
        
        if mode == "TOP":
            ratings_all = df[self.RATING_COL].to_numpy()
            Y_all = np.where(ratings_all >= 4, 100.0, 0.0)
            Y_all = Y_all + np.random.default_rng(123).uniform(-0.5, 0.5, size=Y_all.shape) * 1e-5
        elif mode == "BOTTOM":
            ratings_all = df[self.RATING_COL].to_numpy()
            Y_all = np.where(ratings_all <= 2, 100.0, 0.0)
            Y_all = Y_all + np.random.default_rng(456).uniform(-0.5, 0.5, size=Y_all.shape) * 1e-5
        else:
            rt_all = df[self.RESPONSE_TIME_COL].clip(upper=7.0).to_numpy()
            Y_all = rt_all

        # No intercept
        beta_no, _, _, _ = np.linalg.lstsq(X_all, Y_all, rcond=None)
        
        # With intercept (for t-values)
        n, p = X_all.shape
        X_design = np.column_stack([np.ones(n), X_all])
        beta_full, _, _, _ = np.linalg.lstsq(X_design, Y_all, rcond=None)
        
        # Stats
        Y_hat = X_design @ beta_full
        e = Y_all - Y_hat
        sse = float(np.sum(e ** 2))
        dof = n - X_design.shape[1]
        sigma2 = sse / dof
        
        try:
            XtX_inv = np.linalg.inv(X_design.T @ X_design)
            se = np.sqrt(np.diag(sigma2 * XtX_inv))
            t_vals = beta_full / se
        except:
            t_vals = np.zeros_like(beta_full)
            
        beta_with = beta_full[1:]
        t_elements = t_vals[1:]
        
        pooled_df = pd.DataFrame({
            "element": element_cols,
            "beta_no_intercept": beta_no,
            "beta_with_intercept": beta_with,
            "t_with_intercept": t_elements
        })
        
        # Threshold
        mask = t_elements >= 2.0
        threshold = float(np.min(beta_no[mask])) if np.any(mask) else None
        
        return {"df": pooled_df, "threshold": threshold}

    # --- Grouping Helpers ---
    def _normalize_gender(self, val):
        if not isinstance(val, str): return np.nan
        v = val.strip().lower()
        if v.startswith("m"): return "Male"
        if v.startswith("f"): return "Female"
        return np.nan

    def _build_gender_groups(self, coef_table, df, element_cols):
        # Map panelist to gender
        gender_map = df.dropna(subset=[self.GENDER_COL]).groupby(self.PANEL_COL)[self.GENDER_COL].first()
        
        groups = {}
        if gender_map.empty: return groups
        
        coef_with_gender = coef_table.merge(gender_map.rename("Gender"), left_on="Panelist", right_index=True, how="left")
        coef_with_gender["Gender_norm"] = coef_with_gender["Gender"].apply(self._normalize_gender)
        
        for g_name in ["Male", "Female"]:
            sub = coef_with_gender[coef_with_gender["Gender_norm"] == g_name]
            if not sub.empty:
                groups[g_name] = {
                    "base": int(sub["Panelist"].nunique()),
                    "means": sub[element_cols].mean(axis=0)
                }
        return groups

    def _normalize_age_to_bin(self, val):
        if pd.isna(val): return np.nan
        
        # Helper to check bin
        def check_bin(v_str):
            clean = v_str.replace(" ", "")
            for b in self.AGE_BINS:
                if clean == b.replace(" ", ""): return b
            return None

        if isinstance(val, (int, float)):
            age = int(val)
            if 13 <= age <= 17: return "13-17"
            if 18 <= age <= 24: return "18-24"
            if 25 <= age <= 34: return "25-34"
            if 35 <= age <= 44: return "35-44"
            if 45 <= age <= 54: return "45-54"
            if 55 <= age <= 64: return "55-64"
            if age >= 65: return "65+"
            return np.nan
            
        if isinstance(val, str):
            b = check_bin(val)
            if b: return b
            # Try parsing number
            digits = "".join(ch if ch.isdigit() else " " for ch in val)
            parts = digits.split()
            if parts:
                try:
                    return self._normalize_age_to_bin(int(parts[0]))
                except: pass
        return np.nan

    def _build_age_groups(self, coef_table, df, element_cols):
        age_map = df.dropna(subset=[self.AGE_COL]).groupby(self.PANEL_COL)[self.AGE_COL].first()
        groups = {}
        if age_map.empty: return groups
        
        age_bin_map = age_map.apply(self._normalize_age_to_bin).dropna()
        coef_with_age = coef_table.merge(age_bin_map.rename("AgeBin"), left_on="Panelist", right_index=True, how="left")
        
        for bin_label in self.AGE_BINS:
            sub = coef_with_age[coef_with_age["AgeBin"] == bin_label]
            if not sub.empty:
                groups[bin_label] = {
                    "base": int(sub["Panelist"].nunique()),
                    "means": sub[element_cols].mean(axis=0)
                }
        return groups

    def _build_class_groups(self, coef_table, df, element_cols, class_cols):
        groups = {}
        for col_name in class_cols:
            if col_name not in df.columns: continue
            
            ans_series = df.dropna(subset=[col_name]).groupby(self.PANEL_COL)[col_name].first()
            coef_with_ans = coef_table.merge(ans_series.rename("Answer"), left_on="Panelist", right_index=True, how="left")
            
            segs = {}
            answer_labels = []
            unique_opts = ans_series.dropna().unique()
            
            for opt in unique_opts:
                sub = coef_with_ans[coef_with_ans["Answer"] == opt]
                if not sub.empty:
                    segs[opt] = {
                        "base": int(sub["Panelist"].nunique()),
                        "means": sub[element_cols].mean(axis=0)
                    }
                    answer_labels.append(opt)
            
            if segs:
                groups[col_name] = {
                    "question_text": col_name,
                    "answer_labels": sorted(answer_labels, key=str), # Sort for consistency
                    "segments": segs
                }
        return groups

    def _custom_kmeans_pearson(self, data, k, max_iters=100, seed=42):
        # Center
        row_means = data.mean(axis=1, keepdims=True)
        centered = data - row_means
        # Normalize
        norms = np.linalg.norm(centered, axis=1, keepdims=True)
        norms[norms == 0] = 1e-9
        normalized = centered / norms
        
        n_samples = normalized.shape[0]
        rng = np.random.default_rng(seed)
        
        indices = rng.choice(n_samples, k, replace=False)
        centroids = normalized[indices]
        labels = np.zeros(n_samples, dtype=int)
        
        for _ in range(max_iters):
            dots = normalized @ centroids.T
            new_labels = np.argmax(dots, axis=1)
            if np.array_equal(new_labels, labels): break
            labels = new_labels
            
            new_centroids = np.zeros_like(centroids)
            for i in range(k):
                points = normalized[labels == i]
                if len(points) > 0:
                    mean_vec = points.mean(axis=0)
                    norm = np.linalg.norm(mean_vec)
                    if norm > 1e-9:
                        new_centroids[i] = mean_vec / norm
                    else:
                        new_centroids[i] = normalized[rng.choice(n_samples)]
                else:
                    new_centroids[i] = normalized[rng.choice(n_samples)]
            centroids = new_centroids
            
        return labels

    # --- Excel Helpers ---
    def _autofit_all_cols(self, ws, max_width=30):
        for col_idx in range(1, ws.max_column + 1):
            max_len = 0
            for row_idx in range(1, ws.max_row + 1):
                v = ws.cell(row=row_idx, column=col_idx).value
                if v is not None:
                    max_len = max(max_len, len(str(v)))
            if max_len > 0:
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)

    def _wrap_col_B(self, ws):
        max_len = 0
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=2).value
            if v: max_len = max(max_len, len(str(v)))
        ws.column_dimensions["B"].width = min(max_len + 2, self.MAX_WIDTH)
        for r in range(1, ws.max_row + 1):
            cell = ws.cell(row=r, column=2)
            cell.alignment = Alignment(horizontal=cell.alignment.horizontal or "general", vertical="top", wrap_text=True)

    def _create_front_page(self, wb, data):
        ws = wb.create_sheet("Front Page", 0)
        title = data.get("title", "")
        background = data.get("background", "")
        language = data.get("language", "")
        launched_at = data.get("launched_at", "")
        
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 50
        
        labels = ["Title:", "Background:", "Language:", "Launched At:"]
        values = [title, background, language, launched_at]
        
        thick = Side(border_style="thick", color="000000")
        thin = Side(border_style="thin", color="000000")
        
        for i, (lab, val) in enumerate(zip(labels, values)):
            r = 2 + i
            c_lab = ws.cell(row=r, column=2, value=lab)
            c_val = ws.cell(row=r, column=3, value=val)
            c_lab.font = self.bold_font
            
            top = thick if i == 0 else thin
            bottom = thick if i == len(labels) - 1 else thin
            c_lab.border = Border(top=top, left=thick, right=thin, bottom=bottom)
            c_val.border = Border(top=top, left=thin, right=thick, bottom=bottom)

    def _create_info_block(self, wb, data):
        ws = wb.create_sheet("Information Block", 1)
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 60
        
        row = 1
        def w(l, v, bold=True):
            nonlocal row
            c1 = ws.cell(row=row, column=1, value=l)
            c2 = ws.cell(row=row, column=2, value=v)
            if bold: c1.font = self.bold_font
            c2.alignment = Alignment(wrap_text=True)
            row += 1
            
        w("Study Title", data.get("title", ""))
        w("Study Type", data.get("study_type", ""))
        w("Study Background", data.get("background", ""))
        
        # Categories & Elements
        w("", "")
        categories = data.get("categories", [])
        elements = data.get("elements", [])
        
        for cat in categories:
            cat_name = cat.get("name", "")
            cat_id = cat.get("id")
            w("Study Category", cat_name)
            
            c_elements = [e for e in elements if e.get("category_id") == cat_id]
            for el in c_elements:
                w(f"{cat_name} element", el.get("name", ""))
                w("Element Content", el.get("content", ""))
            w("", "")

    def _create_raw_data_sheet(self, wb, df):
        ws = wb.create_sheet("RawData")
        # Headers
        for c_idx, col in enumerate(df.columns, 1):
            ws.cell(row=1, column=c_idx, value=col).font = self.bold_font
        
        # Data
        for r_idx, row in enumerate(df.itertuples(index=False), 2):
            for c_idx, val in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=val)

    def _create_overall_sheet(self, wb, name, element_cols, sorted_cats, col_to_cat, col_to_elt, means, base, threshold, fill, round_vals):
        ws = wb.create_sheet(name)
        ws["B1"] = "Group"; ws["B1"].font = self.bold_font
        ws["D1"] = "Total"; ws["D1"].font = self.bold_font
        ws["B2"] = "Base Size"; ws["B2"].font = self.bold_font
        ws["D2"] = base
        
        row = 5
        first_val = None
        
        for i, cat_name in enumerate(sorted_cats):
            letter = self.letters[i]
            ws.cell(row=row, column=2, value=f"{letter}. {cat_name}").font = self.bold_font
            ws.cell(row=row, column=2).fill = self.header_fill
            row += 1
            
            cols = [c for c in element_cols if col_to_cat.get(c) == cat_name]
            for j, col in enumerate(cols, 1):
                code = f"{letter}{j}"
                val = means[col]
                if round_vals: val = int(val)
                
                ws.cell(row=row, column=1, value=code)
                ws.cell(row=row, column=2, value=col_to_elt.get(col, col))
                ws.cell(row=row, column=4, value=val)
                
                if first_val is None: first_val = row
                row += 1
            row += 1
            
        self._wrap_col_B(ws)
        
        if threshold is not None and first_val is not None:
            last_val = row - 2
            formula = f'AND(D{first_val}<>"",D{first_val}>={threshold})'
            ws.conditional_formatting.add(f"D{first_val}:D{last_val}", FormulaRule(formula=[formula], fill=fill))

    def _create_mindset_sheet(self, wb, name, coef_df, element_cols, sorted_cats, col_to_cat, col_to_elt, base, threshold, fill, l2, l3, round_vals):
        ws = wb.create_sheet(name)
        ws["B1"] = "Group"; ws["B1"].font = self.bold_font
        ws["B2"] = "Base Size"; ws["B2"].font = self.bold_font
        
        col_idx = 4
        col_map = {}
        
        # Total
        ws.cell(row=1, column=col_idx, value="Total").font = self.bold_font
        ws.cell(row=2, column=col_idx, value=base)
        col_map["Total"] = col_idx
        col_idx += 2
        
        # 2 Clusters
        counts_2 = np.bincount(l2, minlength=2)
        for i in range(2):
            ws.cell(row=1, column=col_idx, value=f"Mindset {i+1} of 2").font = self.bold_font
            ws.cell(row=2, column=col_idx, value=counts_2[i])
            col_map[f"2_{i}"] = col_idx
            col_idx += 1
        col_idx += 1
        
        # 3 Clusters
        counts_3 = np.bincount(l3, minlength=3)
        for i in range(3):
            ws.cell(row=1, column=col_idx, value=f"Mindset {i+1} of 3").font = self.bold_font
            ws.cell(row=2, column=col_idx, value=counts_3[i])
            col_map[f"3_{i}"] = col_idx
            col_idx += 1
            
        # Calc means
        means_total = coef_df[element_cols].mean(axis=0)
        means_2 = [coef_df.iloc[l2 == i][element_cols].mean(axis=0) if np.any(l2 == i) else pd.Series(0, index=element_cols) for i in range(2)]
        means_3 = [coef_df.iloc[l3 == i][element_cols].mean(axis=0) if np.any(l3 == i) else pd.Series(0, index=element_cols) for i in range(3)]
        
        row = 5
        first_val = None
        
        for i, cat_name in enumerate(sorted_cats):
            letter = self.letters[i]
            ws.cell(row=row, column=2, value=f"{letter}. {cat_name}").font = self.bold_font
            ws.cell(row=row, column=2).fill = self.header_fill
            row += 1
            
            cols = [c for c in element_cols if col_to_cat.get(c) == cat_name]
            for j, col in enumerate(cols, 1):
                ws.cell(row=row, column=1, value=f"{letter}{j}")
                ws.cell(row=row, column=2, value=col_to_elt.get(col, col))
                
                def put(k, v):
                    if round_vals: v = int(round(v))
                    ws.cell(row=row, column=col_map[k], value=v)
                
                put("Total", means_total[col])
                for k in range(2): put(f"2_{k}", means_2[k][col])
                for k in range(3): put(f"3_{k}", means_3[k][col])
                
                if first_val is None: first_val = row
                row += 1
            row += 1
            
        self._wrap_col_B(ws)
        
        if threshold is not None and first_val is not None:
            last_val = row - 2
            for c_idx in col_map.values():
                col_let = get_column_letter(c_idx)
                f = f'AND({col_let}{first_val}<>"",{col_let}{first_val}>={threshold})'
                ws.conditional_formatting.add(f"{col_let}{first_val}:{col_let}{last_val}", FormulaRule(formula=[f], fill=fill))

    def _create_segment_sheet(self, wb, name, element_cols, sorted_cats, col_to_cat, col_to_elt, groups, threshold, fill, round_vals, segment_order=None):
        if not groups: return
        ws = wb.create_sheet(name)
        ws["B1"] = "Group"; ws["B1"].font = self.bold_font
        ws["B2"] = "Base Size"; ws["B2"].font = self.bold_font
        
        col_idx = 4
        col_map = {}
        keys = segment_order if segment_order else sorted(groups.keys())
        
        for k in keys:
            if k in groups:
                ws.cell(row=1, column=col_idx, value=k).font = self.bold_font
                ws.cell(row=2, column=col_idx, value=groups[k]["base"])
                col_map[k] = col_idx
                col_idx += 1
                
        row = 5
        first_val = None
        
        for i, cat_name in enumerate(sorted_cats):
            letter = self.letters[i]
            ws.cell(row=row, column=2, value=f"{letter}. {cat_name}").font = self.bold_font
            ws.cell(row=row, column=2).fill = self.header_fill
            row += 1
            
            cols = [c for c in element_cols if col_to_cat.get(c) == cat_name]
            for j, col in enumerate(cols, 1):
                ws.cell(row=row, column=1, value=f"{letter}{j}")
                ws.cell(row=row, column=2, value=col_to_elt.get(col, col))
                
                for k in keys:
                    if k in groups:
                        val = groups[k]["means"][col]
                        if round_vals: val = int(round(val))
                        ws.cell(row=row, column=col_map[k], value=val)
                
                if first_val is None: first_val = row
                row += 1
            row += 1
            
        self._wrap_col_B(ws)
        
        if threshold is not None and first_val is not None:
            last_val = row - 2
            for c_idx in col_map.values():
                col_let = get_column_letter(c_idx)
                f = f'AND({col_let}{first_val}<>"",{col_let}{first_val}>={threshold})'
                ws.conditional_formatting.add(f"{col_let}{first_val}:{col_let}{last_val}", FormulaRule(formula=[f], fill=fill))

    def _create_classification_sheet(self, wb, name, element_cols, sorted_cats, col_to_cat, col_to_elt, groups, threshold, fill, round_vals):
        if not groups: return
        ws = wb.create_sheet(name)
        ws["B1"] = "Group"; ws["B1"].font = self.bold_font
        ws["B2"] = "Base Size"; ws["B2"].font = self.bold_font
        
        col_idx = 4
        col_map = {}
        
        for q_col, info in groups.items():
            ws.cell(row=1, column=col_idx, value=info["question_text"]).font = self.bold_font
            ws.cell(row=1, column=col_idx).fill = self.header_fill
            col_idx += 1
            
            for ans in info["answer_labels"]:
                ws.cell(row=1, column=col_idx, value=ans).font = self.bold_font
                ws.cell(row=2, column=col_idx, value=info["segments"][ans]["base"])
                col_map[(q_col, ans)] = col_idx
                col_idx += 1
            col_idx += 1 # Spacer
            
        row = 5
        first_val = None
        
        for i, cat_name in enumerate(sorted_cats):
            letter = self.letters[i]
            ws.cell(row=row, column=2, value=f"{letter}. {cat_name}").font = self.bold_font
            ws.cell(row=row, column=2).fill = self.header_fill
            row += 1
            
            cols = [c for c in element_cols if col_to_cat.get(c) == cat_name]
            for j, col in enumerate(cols, 1):
                ws.cell(row=row, column=1, value=f"{letter}{j}")
                ws.cell(row=row, column=2, value=col_to_elt.get(col, col))
                
                for (q_col, ans), c_idx in col_map.items():
                    val = groups[q_col]["segments"][ans]["means"][col]
                    if round_vals: val = int(round(val))
                    ws.cell(row=row, column=c_idx, value=val)
                
                if first_val is None: first_val = row
                row += 1
            row += 1
            
        self._wrap_col_B(ws)
        
        if threshold is not None and first_val is not None:
            last_val = row - 2
            for c_idx in col_map.values():
                col_let = get_column_letter(c_idx)
                f = f'AND({col_let}{first_val}<>"",{col_let}{first_val}>={threshold})'
                ws.conditional_formatting.add(f"{col_let}{first_val}:{col_let}{last_val}", FormulaRule(formula=[f], fill=fill))

    def _create_combined_sheet(self, wb, name, element_cols, sorted_cats, col_to_cat, col_to_elt, base, means, g_groups, a_groups, c_groups, threshold, fill, round_vals):
        ws = wb.create_sheet(name)
        ws["B1"] = "Group"; ws["B1"].font = self.bold_font
        ws["B2"] = "Base Size"; ws["B2"].font = self.bold_font
        
        col_idx = 4
        col_map = {}
        
        # Overall
        ws.cell(row=1, column=col_idx, value="Overall").font = self.bold_font
        ws.cell(row=2, column=col_idx, value=base)
        col_map["Overall"] = col_idx
        col_idx += 1
        
        # Gender
        for g in ["Male", "Female"]:
            if g in g_groups:
                ws.cell(row=1, column=col_idx, value=g).font = self.bold_font
                ws.cell(row=2, column=col_idx, value=g_groups[g]["base"])
                col_map[f"G_{g}"] = col_idx
                col_idx += 1
                
        # Age
        for a in self.AGE_BINS:
            if a in a_groups:
                ws.cell(row=1, column=col_idx, value=a).font = self.bold_font
                ws.cell(row=2, column=col_idx, value=a_groups[a]["base"])
                col_map[f"A_{a}"] = col_idx
                col_idx += 1
                
        # Classification
        for q_col, info in c_groups.items():
            ws.cell(row=1, column=col_idx, value=info["question_text"]).font = self.bold_font
            ws.cell(row=1, column=col_idx).fill = self.header_fill
            col_idx += 1
            for ans in info["answer_labels"]:
                ws.cell(row=1, column=col_idx, value=ans).font = self.bold_font
                ws.cell(row=2, column=col_idx, value=info["segments"][ans]["base"])
                col_map[f"C_{q_col}_{ans}"] = col_idx
                col_idx += 1
            col_idx += 1
            
        row = 5
        first_val = None
        
        for i, cat_name in enumerate(sorted_cats):
            letter = self.letters[i]
            ws.cell(row=row, column=2, value=f"{letter}. {cat_name}").font = self.bold_font
            ws.cell(row=row, column=2).fill = self.header_fill
            row += 1
            
            cols = [c for c in element_cols if col_to_cat.get(c) == cat_name]
            for j, col in enumerate(cols, 1):
                ws.cell(row=row, column=1, value=f"{letter}{j}")
                ws.cell(row=row, column=2, value=col_to_elt.get(col, col))
                
                def p(k, v):
                    if round_vals: v = int(round(v))
                    ws.cell(row=row, column=col_map[k], value=v)
                
                p("Overall", means[col])
                for g in ["Male", "Female"]:
                    if g in g_groups: p(f"G_{g}", g_groups[g]["means"][col])
                for a in self.AGE_BINS:
                    if a in a_groups: p(f"A_{a}", a_groups[a]["means"][col])
                for q_col, info in c_groups.items():
                    for ans in info["answer_labels"]:
                        p(f"C_{q_col}_{ans}", info["segments"][ans]["means"][col])
                        
                if first_val is None: first_val = row
                row += 1
            row += 1
            
        self._wrap_col_B(ws)
        
        if threshold is not None and first_val is not None:
            last_val = row - 2
            for c_idx in col_map.values():
                col_let = get_column_letter(c_idx)
                f = f'AND({col_let}{first_val}<>"",{col_let}{first_val}>={threshold})'
                ws.conditional_formatting.add(f"{col_let}{first_val}:{col_let}{last_val}", FormulaRule(formula=[f], fill=fill))

    def _create_intercepts_sheet(self, wb, name, df, threshold, fill):
        ws = wb.create_sheet(name)
        
        # Headers
        for i, col in enumerate(df.columns, 1):
            ws.cell(row=1, column=i, value=col).font = self.bold_font
            
        # Data
        for r_idx, row in enumerate(df.itertuples(index=False), 2):
            for c_idx, val in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=val)
                
        self._autofit_all_cols(ws)
        
        if threshold is not None:
            last = ws.max_row
            ws.cell(row=last+2, column=1, value="Threshold").font = self.bold_font
            ws.cell(row=last+2, column=2, value=threshold)
            
            # Conditional formatting on t-value column (assumed index 4)
            # Actually logic in v2 was on t_with_intercept >= 2.0
            # Let's just highlight rows where t >= 2?
            # v2 logic: ws_T_int.conditional_formatting.add(f"D2:D{last_T_raw}", rule_t_T)
            # D is 4th column.
            rule = CellIsRule(operator="greaterThanOrEqual", formula=["2.0"], fill=fill)
            ws.conditional_formatting.add(f"D2:D{ws.max_row}", rule)
