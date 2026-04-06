# app/api/v1/project.py
from __future__ import annotations

import csv
import io
import zipfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import List, Dict, Any, Optional, Tuple

from uuid import UUID

import pandas as pd

from fastapi import APIRouter, Depends, Query, status
from fastapi.responses import StreamingResponse, Response
from sqlalchemy.orm import Session

from app.core.dependencies import get_current_active_user
from app.core.cache import RedisCache
from app.db.session import get_db, SessionLocal
from app.models.user_model import User
from app.schemas.project_schema import (
    ProjectCreate, ProjectUpdate, ProjectOut, ProjectListItem,
    ProjectMemberInvite, ProjectMemberOut, ProjectMemberUpdate,
    ValidateProductRequest, ValidateProductResponse,
    AssignStudyRequest, AssignStudyResponse,
    ExportCompletedPanelistsRequest,
    ExportAbandonedResponsesRequest,
)
from app.services import project_service
from app.services.project_member_service import project_member_service
from app.services.study import build_study_data_for_analysis
from app.services.response import StudyResponseService
from app.services.analysis import StudyAnalysisService

router = APIRouter()


def _invalidate_public_project_studies_cache(project_id: UUID) -> None:
    RedisCache.delete(f"project_public_studies:{project_id}")


@router.post("/{project_id}/members/invite", response_model=ProjectMemberOut)
def invite_project_member_endpoint(
    project_id: UUID,
    payload: ProjectMemberInvite,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """
    Invite a user to a project by email.
    User will automatically get access to all studies in the project.
    """
    member = project_member_service.invite_member(
        db=db,
        project_id=project_id,
        inviter=current_user,
        payload=payload
    )
    
    # Enrich with user details if available
    member_dict = {
        "id": member.id,
        "project_id": member.project_id,
        "user_id": member.user_id,
        "email": member.invited_email,
        "role": member.role,
        "created_at": member.created_at,
        "updated_at": member.updated_at,
        "name": None,
        "is_registered": False
    }
    
    if hasattr(member, 'user') and member.user:
        member_dict["name"] = member.user.name
        member_dict["is_registered"] = True
    
    return ProjectMemberOut(**member_dict)


@router.get("/{project_id}/members", response_model=List[ProjectMemberOut])
def list_project_members_endpoint(
    project_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """
    List all members of a project, including the project creator.
    """
    # Ensure current user has access to this project (owner or member)
    project_service.get_project(db=db, project_id=project_id, user_id=current_user.id)
    
    members = project_member_service.list_members(db=db, project_id=project_id)
    
    # Enrich members with user details
    enriched_members = []
    for member in members:
        member_dict = {
            "id": member.id,
            "project_id": member.project_id,
            "user_id": member.user_id,
            "email": member.invited_email,
            "role": member.role,
            "created_at": member.created_at,
            "updated_at": member.updated_at,
            "name": None,
            "is_registered": False
        }
        
        if hasattr(member, 'user') and member.user:
            member_dict["name"] = member.user.name
            member_dict["is_registered"] = True
        
        enriched_members.append(member_dict)
    
    return enriched_members


@router.patch("/{project_id}/members/{member_id}", response_model=ProjectMemberOut)
def update_project_member_endpoint(
    project_id: UUID,
    member_id: UUID,
    payload: ProjectMemberUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """
    Update a project member's role.
    Role change will automatically sync to all studies in the project.
    """
    member = project_member_service.update_member_role(
        db=db,
        project_id=project_id,
        member_id=member_id,
        current_user=current_user,
        payload=payload
    )
    
    return ProjectMemberOut(
        id=member.id,
        project_id=member.project_id,
        user_id=member.user_id,
        email=member.invited_email,
        role=member.role,
        created_at=member.created_at,
        updated_at=member.updated_at
    )


@router.delete("/{project_id}/members/{member_id}", status_code=status.HTTP_204_NO_CONTENT)
def remove_project_member_endpoint(
    project_id: UUID,
    member_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """
    Remove a member from the project.
    Member will also be removed from all studies in the project.
    """
    project_member_service.remove_member(
        db=db,
        project_id=project_id,
        member_id=member_id,
        current_user=current_user
    )
    return None


@router.post("", response_model=ProjectOut, status_code=status.HTTP_201_CREATED)
def create_project_endpoint(
    payload: ProjectCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """
    Create a new project (requires: name, optional: description).
    Ultra-fast endpoint optimized for <100ms response time.
    """
    project = project_service.create_project(
        db=db,
        creator_id=current_user.id,
        payload=payload
    )
    
    # Return with study count (0 for new projects)
    return ProjectOut(
        id=project.id,
        name=project.name,
        description=project.description,
        creator_id=project.creator_id,
        created_at=project.created_at,
        updated_at=project.updated_at,
        study_count=0
    )


@router.get("", response_model=List[ProjectOut])
def get_projects_endpoint(
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """
    Fetch all projects for the authenticated user.
    Optimized for <200ms response time with study counts.
    """
    projects = project_service.get_projects_for_user(
        db=db,
        user_id=current_user.id,
        page=page,
        per_page=per_page
    )
    return projects


@router.get("/{project_id}", response_model=ProjectOut)
def get_project_endpoint(
    project_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """
    Get a single project by ID.
    """
    from sqlalchemy import select, func
    from app.models.study_model import Study
    from app.models.project_model import Project
    
    project = project_service.get_project(
        db=db,
        project_id=project_id,
        user_id=current_user.id
    )
    
    # Get study count for this project
    study_count = db.scalar(
        select(func.count(Study.id)).where(Study.project_id == project_id)
    ) or 0
    
    return ProjectOut(
        id=project.id,
        name=project.name,
        description=project.description,
        creator_id=project.creator_id,
        created_at=project.created_at,
        updated_at=project.updated_at,
        study_count=int(study_count)
    )


@router.get("/{project_id}/studies")
def get_project_studies_endpoint(
    project_id: UUID,
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """
    Fetch all studies for a specific project.
    Optimized for <200ms response time.
    """
    studies = project_service.get_project_studies(
        db=db,
        project_id=project_id,
        user_id=current_user.id,
        page=page,
        per_page=per_page
    )
    return studies


@router.post("/validate-product", response_model=ValidateProductResponse)
def validate_product_endpoint(
    payload: ValidateProductRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """
    Validate that product_id and key combination are unique within the study's project.
    Sends product_id, study_id, product_keys in body. Project is derived from study.
    If study has no project, returns valid=True. Optimized for double-digit ms.
    """
    return project_service.validate_product_by_study(
        db=db,
        user_id=current_user.id,
        payload=payload,
    )


@router.post("/{project_id}/assign-study", response_model=AssignStudyResponse)
def assign_study_to_project_endpoint(
    project_id: UUID,
    payload: AssignStudyRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """
    Assign a standalone study to a project. Ultra-fast (~10ms).

    - Study must not already be in any project
    - Study viewer cannot assign; editor and admin can
    - Project viewer cannot assign; editor and admin can
    - Project creator becomes study admin; study creator demoted per project logic
    """
    result = project_service.assign_study_to_project(
        db=db,
        project_id=project_id,
        user_id=current_user.id,
        payload=payload,
    )
    _invalidate_public_project_studies_cache(project_id)
    return result


@router.put("/{project_id}", response_model=ProjectOut)
def update_project_endpoint(
    project_id: UUID,
    payload: ProjectUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """
    Update a project.
    """
    from sqlalchemy import select, func
    from app.models.study_model import Study
    
    project = project_service.update_project(
        db=db,
        project_id=project_id,
        user_id=current_user.id,
        payload=payload
    )
    
    # Get study count for this project
    study_count = db.scalar(
        select(func.count(Study.id)).where(Study.project_id == project_id)
    ) or 0
    _invalidate_public_project_studies_cache(project_id)
    
    return ProjectOut(
        id=project.id,
        name=project.name,
        description=project.description,
        creator_id=project.creator_id,
        created_at=project.created_at,
        updated_at=project.updated_at,
        study_count=int(study_count)
    )


@router.post("/{project_id}/flattened-project-csv")
def flattened_project_csv_endpoint(
    project_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """
    Export an Excel file (.xlsx) with two sheets:
    1. Product Data: one row per study (product_id, keys %, classification averages, element T Overall scores, Rating, ResponseTime).
    2. Range Sheet: Dependent Variable, Low, High, Average for each classification question and category-image column.
    Studies are ordered by product_id. Any project member (viewer/editor/admin) can download.
    """
    studies = project_service.get_project_studies_for_export(
        db=db,
        project_id=project_id,
        user_id=current_user.id,
    )
    if not studies:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Product Data"
        ws.append(["product_id", "Rating", "ResponseTime"])
        range_ws = wb.create_sheet("Range Sheet")
        range_ws.append(["Dependent Variable", "Low", "High", "Average"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="project_{project_id}_flattened.xlsx"'},
        )

    # Build canonical header from all studies
    all_key_names = []
    seen_keys = set()
    for s in studies:
        for k in (getattr(s, "product_keys", None) or []):
            if isinstance(k, dict) and k.get("name") and k["name"] not in seen_keys:
                seen_keys.add(k["name"])
                all_key_names.append(k["name"])

    class_q_names = []
    if studies and studies[0].classification_questions:
        for q in sorted(studies[0].classification_questions, key=lambda x: getattr(x, "order", 0)):
            class_q_names.append(q.question_text or "")

    element_cols_ordered = []
    seen_el = set()
    for s in studies:
        study_data = build_study_data_for_analysis(s)
        for el in study_data.get("elements", []):
            cat = el.get("category", {}) or {}
            cat_name = cat.get("name", "")
            el_name = el.get("name", "")
            if not cat_name or not el_name:
                continue
            col = f"{cat_name}-{el_name}".replace("_", "-").replace(" ", "-")
            if col not in seen_el:
                seen_el.add(col)
                element_cols_ordered.append(col)

    header = ["product_id"] + all_key_names + class_q_names + element_cols_ordered + ["Rating", "ResponseTime"]

    def _round6(val):
        """Format numeric values to at most 6 decimal places for export."""
        if val is None or val == "":
            return ""
        try:
            f = float(val)
            return round(f, 6)
        except (TypeError, ValueError):
            return val

    def compute_row(
        study_id: UUID,
        product_id_val: Optional[str],
        product_keys_list: List[Dict[str, Any]],
        study_data_dict: Dict[str, Any],
        index: int,
    ):
        db_session = SessionLocal()
        try:
            response_svc = StudyResponseService(db_session)
            df = response_svc.get_study_dataframe(
                study_id,
                unilever_format=True,
                completed_only=True,
            )
            analysis_svc = StudyAnalysisService()
            t_scores = analysis_svc.get_t_overall_scores(df, study_data_dict) if not df.empty else {}

            key_map = {}
            for k in product_keys_list or []:
                if isinstance(k, dict) and k.get("name") is not None:
                    pct = k.get("percentage")
                    key_map[k["name"]] = _round6(pct) if pct is not None else ""

            row = {"product_id": product_id_val or ""}
            for kn in all_key_names:
                row[kn] = key_map.get(kn, "")

            for qn in class_q_names:
                if qn and not df.empty and qn in df.columns:
                    try:
                        row[qn] = _round6(float(df[qn].mean()))
                    except (TypeError, ValueError):
                        row[qn] = ""
                else:
                    row[qn] = ""

            for ec in element_cols_ordered:
                row[ec] = _round6(t_scores.get(ec, ""))

            if not df.empty and "Rating" in df.columns:
                try:
                    row["Rating"] = _round6(float(df["Rating"].mean()))
                except (TypeError, ValueError):
                    row["Rating"] = ""
            else:
                row["Rating"] = ""
            if not df.empty and "ResponseTime" in df.columns:
                try:
                    row["ResponseTime"] = _round6(float(df["ResponseTime"].mean()))
                except (TypeError, ValueError):
                    row["ResponseTime"] = ""
            else:
                row["ResponseTime"] = ""

            return (index, row)
        finally:
            db_session.close()

    max_workers = min(10, max(1, len(studies)))
    index_to_row = {}
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {}
        for i, s in enumerate(studies):
            study_data_dict = build_study_data_for_analysis(s)
            product_keys_list = getattr(s, "product_keys", None) or []
            if not isinstance(product_keys_list, list):
                product_keys_list = []
            fut = executor.submit(
                compute_row,
                s.id,
                getattr(s, "product_id", None),
                product_keys_list,
                study_data_dict,
                i,
            )
            futures[fut] = i
        for fut in as_completed(futures):
            idx, row = fut.result()
            index_to_row[idx] = row

    def format_cell(v):
        """Ensure numeric cells are written with at most 6 decimal places."""
        if v is None or v == "":
            return ""
        if isinstance(v, (int, float)):
            return round(float(v), 6)
        try:
            return round(float(v), 6)
        except (TypeError, ValueError):
            return v

    # Range sheet: dependent variables = classification questions first, then categoryname-imagename
    dependent_vars = class_q_names + element_cols_ordered
    ordered_rows = [index_to_row[i] for i in range(len(studies)) if index_to_row.get(i) is not None]
    range_rows = []
    for col in dependent_vars:
        values = []
        for row in ordered_rows:
            v = row.get(col, "")
            if v is None or v == "":
                continue
            try:
                values.append(float(v))
            except (TypeError, ValueError):
                continue
        if values:
            low = round(min(values), 6)
            high = round(max(values), 6)
            avg = round(sum(values) / len(values), 6)
            range_rows.append((col, low, high, avg))
        else:
            range_rows.append((col, "", "", ""))

    from openpyxl import Workbook

    wb = Workbook()
    ws_product = wb.active
    ws_product.title = "Product Data"
    ws_product.append(header)
    for i in range(len(studies)):
        row = index_to_row.get(i)
        if row is None:
            continue
        cells = [format_cell(row.get(h, "")) for h in header]
        ws_product.append(cells)

    ws_range = wb.create_sheet("Range Sheet")
    ws_range.append(["Dependent Variable", "Low", "High", "Average"])
    for dep_var, low, high, avg in range_rows:
        ws_range.append([dep_var, low, high, avg])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="project_{project_id}_flattened.xlsx"'},
    )


@router.post("/{project_id}/export-zip")
def export_project_zip_endpoint(
    project_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """
    Start a background job to export project as ZIP containing per-study Excel reports and mega_sheet.xlsx.
    Returns a job_id that can be polled for status and download URL.
    """
    import uuid as uuid_module
    from app.models.job_model import Job, JobStatus
    from app.tasks.celery_jobs import export_project_zip_celery

    # Create a job record
    job_id = str(uuid_module.uuid4())
    job = Job(
        job_id=job_id,
        study_id=str(project_id),
        user_id=str(current_user.id),
        status=JobStatus.PENDING,
        progress=0.0,
        message="Export queued...",
    )
    db.add(job)
    db.commit()

    # Trigger Celery task
    export_project_zip_celery.delay(job_id, str(project_id), str(current_user.id))

    return {"job_id": job_id, "status": "pending", "message": "Export started. Poll /projects/export-job/{job_id} for status."}


@router.post("/{project_id}/export-zip-sync")
def export_project_zip_sync_endpoint(
    project_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """
    [Legacy/Debug] Synchronous export - may timeout for large projects.
    Export a ZIP containing per-study Excel reports and mega_sheet.xlsx.
    Per-study files: full Excel report (same as /export/study/{id}/flattened-csv) - Raw Data, T Overall, B Overall, Info Block, etc. Named {product_id}.xlsx.
    mega_sheet.xlsx: Raw Data (sorted by product_id), Product Data, Range Sheet.
    Target: <30s for 100 studies, <60s for 200 studies.
    """
    studies = project_service.get_project_studies_for_export(
        db=db,
        project_id=project_id,
        user_id=current_user.id,
    )

    def _round6(val):
        if val is None or val == "":
            return ""
        try:
            return round(float(val), 6)
        except (TypeError, ValueError):
            return ""

    def _format_cell(v):
        if v is None or v == "":
            return ""
        if isinstance(v, (int, float)):
            return round(float(v), 6)
        try:
            return round(float(v), 6)
        except (TypeError, ValueError):
            return v

    # Build canonical header from all studies (same as flattened-project-csv)
    all_key_names = []
    seen_keys = set()
    for s in studies:
        for k in (getattr(s, "product_keys", None) or []):
            if isinstance(k, dict) and k.get("name") and k["name"] not in seen_keys:
                seen_keys.add(k["name"])
                all_key_names.append(k["name"])

    class_q_names = []
    if studies and getattr(studies[0], "classification_questions", None):
        for q in sorted(studies[0].classification_questions, key=lambda x: getattr(x, "order", 0)):
            class_q_names.append(q.question_text or "")

    element_cols_ordered = []
    seen_el = set()
    for s in studies:
        study_data = build_study_data_for_analysis(s)
        for el in study_data.get("elements", []):
            cat = el.get("category", {}) or {}
            cat_name = cat.get("name", "")
            el_name = el.get("name", "")
            if not cat_name or not el_name:
                continue
            col = f"{cat_name}-{el_name}".replace("_", "-").replace(" ", "-")
            if col not in seen_el:
                seen_el.add(col)
                element_cols_ordered.append(col)

    product_header = ["product_id"] + all_key_names + class_q_names + element_cols_ordered + ["Rating", "ResponseTime"]

    def _worker(
        study_id: UUID,
        product_id_val: Optional[str],
        product_keys_list: List[Dict[str, Any]],
        study_data_dict: Dict[str, Any],
        index: int,
    ) -> Tuple[int, UUID, str, bytes, Dict[str, Any], pd.DataFrame]:
        """Worker: returns (index, study_id, product_id, xlsx_bytes, product_row_dict, raw_df). Per-study file = full Excel report (same as flattened-csv). Robust: never raises."""
        xlsx_bytes = b""
        product_row = {"product_id": product_id_val or ""}
        for kn in all_key_names:
            product_row[kn] = ""
        for qn in class_q_names:
            product_row[qn] = ""
        for ec in element_cols_ordered:
            product_row[ec] = ""
        product_row["Rating"] = ""
        product_row["ResponseTime"] = ""
        raw_df = pd.DataFrame()
        db_session = None

        try:
            db_session = SessionLocal()
            response_svc = StudyResponseService(db_session)
            df = response_svc.get_study_dataframe(
                study_id,
                unilever_format=True,
                completed_only=True,
            )

            # Per-study file: full Excel report (same as /export/study/{id}/flattened-csv) - Raw Data, T Overall, B Overall, Info Block, etc.
            analysis_svc = StudyAnalysisService()
            try:
                excel_io = analysis_svc.generate_report(df, study_data_dict)
                xlsx_bytes = excel_io.getvalue() if excel_io else b""
            except Exception:
                # Fallback: minimal Excel with headers if report generation fails
                from openpyxl import Workbook
                wb = Workbook()
                ws = wb.active
                ws.title = "Raw Data"
                ws.append(list(df.columns) if not df.empty else ["product_id"])
                buf = io.BytesIO()
                wb.save(buf)
                xlsx_bytes = buf.getvalue()

            # Product row (aggregated)
            t_scores = analysis_svc.get_t_overall_scores(df, study_data_dict) if not df.empty else {}

            key_map = {}
            for k in product_keys_list or []:
                if isinstance(k, dict) and k.get("name") is not None:
                    pct = k.get("percentage")
                    key_map[k["name"]] = _round6(pct) if pct is not None else ""
            for kn in all_key_names:
                product_row[kn] = key_map.get(kn, "")

            for qn in class_q_names:
                if qn and not df.empty and qn in df.columns:
                    try:
                        product_row[qn] = _round6(float(df[qn].mean()))
                    except (TypeError, ValueError):
                        pass
                else:
                    product_row[qn] = ""

            for ec in element_cols_ordered:
                product_row[ec] = _round6(t_scores.get(ec, ""))

            if not df.empty and "Rating" in df.columns:
                try:
                    product_row["Rating"] = _round6(float(df["Rating"].mean()))
                except (TypeError, ValueError):
                    pass
            if not df.empty and "ResponseTime" in df.columns:
                try:
                    product_row["ResponseTime"] = _round6(float(df["ResponseTime"].mean()))
                except (TypeError, ValueError):
                    pass

            # Raw df: add product_id as first column; drop duplicate "Product ID" if present
            pid = product_id_val or "unknown"
            raw_df = df.copy()
            if "Product ID" in raw_df.columns:
                raw_df = raw_df.drop(columns=["Product ID"])
            raw_df.insert(0, "product_id", pid)
        except Exception:
            # Robust: return safe fallbacks so worker never fails the export
            pass
        finally:
            if db_session is not None:
                try:
                    db_session.close()
                except Exception:
                    pass

        return (index, study_id, product_id_val or "", xlsx_bytes, product_row, raw_df)

    if not studies:
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Raw Data"
            ws.append(["product_id"])
            ws2 = wb.create_sheet("Product Data")
            ws2.append(product_header)
            ws3 = wb.create_sheet("Range Sheet")
            ws3.append(["Dependent Variable", "Low", "High", "Average"])
            xl_buf = io.BytesIO()
            wb.save(xl_buf)
            xl_buf.seek(0)
            zf.writestr("mega_sheet.xlsx", xl_buf.getvalue())
        buf.seek(0)
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="application/zip",
            headers={"Content-Disposition": f'attachment; filename="project_{project_id}_export.zip"'},
        )

    max_workers = min(20, max(1, len(studies)))
    results: List[Tuple[int, UUID, str, bytes, Dict[str, Any]]] = []
    raw_chunks: List[pd.DataFrame] = []
    BATCH_SIZE = 20
    current_batch: List[pd.DataFrame] = []

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {}
        for i, s in enumerate(studies):
            study_data_dict = build_study_data_for_analysis(s)
            product_keys_list = getattr(s, "product_keys", None) or []
            if not isinstance(product_keys_list, list):
                product_keys_list = []
            fut = executor.submit(
                _worker,
                s.id,
                getattr(s, "product_id", None),
                product_keys_list,
                study_data_dict,
                i,
            )
            futures[fut] = i
        for fut in as_completed(futures):
            try:
                idx, sid, pid, xlsx_bytes, product_row, raw_df = fut.result()
                results.append((idx, sid, pid, xlsx_bytes, product_row))
                if raw_df is not None and not raw_df.empty:
                    current_batch.append(raw_df)
                    if len(current_batch) >= BATCH_SIZE:
                        raw_chunks.append(pd.concat(current_batch, join="outer", ignore_index=True))
                        current_batch = []
            except Exception:
                pass

        if current_batch:
            raw_chunks.append(pd.concat(current_batch, join="outer", ignore_index=True))

    # Sort by index so order matches studies
    results.sort(key=lambda r: r[0])

    # Build product_id -> filename map (handle duplicates); sanitize for filesystem
    def _safe_filename(s: str) -> str:
        safe = "".join(c if c.isalnum() or c in "._-" else "_" for c in (s or ""))
        return safe.strip() or "unknown"

    pid_counts: Dict[str, int] = {}
    index_to_filename: Dict[int, str] = {}
    for idx, _sid, pid, _xlsx, _row in results:
        base = _safe_filename(pid or "unknown")
        pid_counts[base] = pid_counts.get(base, 0) + 1
        n = pid_counts[base]
        filename = f"{base}.xlsx" if n == 1 else f"{base}_{n}.xlsx"
        index_to_filename[idx] = filename

    # Build ZIP
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for idx, _sid, _pid, xlsx_bytes, _row in results:
            zf.writestr(index_to_filename[idx], xlsx_bytes)

        # Mega-sheet: Raw Data, Product Data, Range Sheet (raw_chunks built during executor loop)
        raw_combined = pd.concat(raw_chunks, join="outer", ignore_index=True) if raw_chunks else pd.DataFrame(columns=["product_id"])
        # Sort Raw Data by product_id (401, 402, ...); preserve panelist/row order within each product_id
        if not raw_combined.empty and "product_id" in raw_combined.columns:
            raw_combined["_sort_key"] = raw_combined["product_id"].fillna("").astype(str).str.strip()
            raw_combined.loc[raw_combined["_sort_key"] == "", "_sort_key"] = "zzzz"
            raw_combined["_order"] = range(len(raw_combined))
            raw_combined = raw_combined.sort_values(by=["_sort_key", "_order"], kind="mergesort").drop(
                columns=["_sort_key", "_order"]
            ).reset_index(drop=True)

        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows

        wb = Workbook()
        ws_raw = wb.active
        ws_raw.title = "Raw Data"
        if not raw_combined.empty:
            for r in dataframe_to_rows(raw_combined, index=False, header=True):
                ws_raw.append(r)
        else:
            ws_raw.append(["product_id"])

        ws_product = wb.create_sheet("Product Data")
        ws_product.append(product_header)
        for r in results:
            row_dict = r[4]
            cells = [_format_cell(row_dict.get(h, "")) for h in product_header]
            ws_product.append(cells)

        ws_range = wb.create_sheet("Range Sheet")
        ws_range.append(["Dependent Variable", "Low", "High", "Average"])
        product_rows = [r[4] for r in results]
        # Only classification questions and task elements; exclude product_id, keys, Rating, ResponseTime
        range_cols = class_q_names + element_cols_ordered
        for col in range_cols:
            vals = []
            for row in product_rows:
                v = row.get(col, "")
                if v is None or v == "":
                    continue
                try:
                    vals.append(float(v))
                except (TypeError, ValueError):
                    continue
            if vals:
                ws_range.append([col, round(min(vals), 6), round(max(vals), 6), round(sum(vals) / len(vals), 6)])
            else:
                ws_range.append([col, "", "", ""])

        xl_buf = io.BytesIO()
        wb.save(xl_buf)
        xl_buf.seek(0)
        zf.writestr("mega_sheet.xlsx", xl_buf.getvalue())

    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="project_{project_id}_export.zip"'},
    )


@router.delete("/{project_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_project_endpoint(
    project_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """
    Delete a project (studies will have their project_id set to NULL).
    """
    project_service.delete_project(
        db=db,
        project_id=project_id,
        user_id=current_user.id
    )
    _invalidate_public_project_studies_cache(project_id)
    return None


@router.get("/export-job/{job_id}")
def get_export_job_status(
    job_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """
    Get the status of an export job.
    Returns status, progress, and download_url when completed.
    """
    from app.models.job_model import Job, JobStatus
    from fastapi import HTTPException

    job = db.query(Job).filter(Job.job_id == job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    
    if job.user_id != str(current_user.id):
        raise HTTPException(status_code=403, detail="Not authorized to access this job")

    response = {
        "job_id": job.job_id,
        "status": job.status.value if job.status else "unknown",
        "progress": job.progress or 0,
        "message": job.message or "",
    }

    if job.status == JobStatus.COMPLETED and job.result:
        response["download_url"] = job.result.get("download_url")
        response["filename"] = job.result.get("filename")
    elif job.status == JobStatus.FAILED:
        response["error"] = job.error

    return response


@router.post("/{project_id}/export-completed-panelists")
def export_completed_panelists_endpoint(
    project_id: UUID,
    payload: ExportCompletedPanelistsRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """
    Export completed panelists from all studies in a project as CSV.
    
    Optimized for fast response using a single SQL query with joins.
    
    Args:
        project_id: UUID of the project
        payload: Optional time filters (after_utc, before_utc)
    
    Returns:
        StreamingResponse with CSV data
    """
    from fastapi import HTTPException
    from sqlalchemy import text, and_
    from datetime import datetime
    
    # Verify project exists and user has access
    project = project_service.get_project(db=db, project_id=project_id, user_id=current_user.id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    # Build optimized single query with all joins
    query = """
        SELECT 
            s.title as study_name,
            s.id as study_id,
            sr.panelist_id,
            sr.respondent_id,
            sr.session_id,
            sr.is_completed,
            sr.status,
            sr.completion_percentage,
            sr.session_start_time,
            sr.session_end_time,
            sr.total_study_duration
        FROM study_responses sr
        INNER JOIN studies s ON sr.study_id = s.id
        WHERE s.project_id = :project_id
          AND (sr.is_completed = true OR sr.status = 'completed')
    """
    params: Dict[str, Any] = {"project_id": str(project_id)}
    
    if payload.after_utc:
        query += " AND sr.session_end_time >= :after_utc"
        params["after_utc"] = payload.after_utc
    
    if payload.before_utc:
        query += " AND sr.session_end_time < :before_utc"
        params["before_utc"] = payload.before_utc
    
    query += " ORDER BY sr.session_end_time"
    
    # Execute query
    results = db.execute(text(query), params).fetchall()
    
    # Generate CSV using streaming for memory efficiency
    def generate_csv():
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow([
            "study_name",
            "study_id", 
            "panelist_id",
            "respondent_id",
            "session_id",
            "is_completed",
            "status",
            "completion_percentage",
            "session_start_time",
            "session_end_time",
            "total_duration_seconds"
        ])
        yield output.getvalue()
        output.seek(0)
        output.truncate(0)
        
        # Write data rows in chunks for better streaming
        for row in results:
            writer.writerow([
                row.study_name,
                str(row.study_id),
                row.panelist_id or "N/A",
                row.respondent_id,
                row.session_id,
                row.is_completed,
                row.status,
                row.completion_percentage,
                row.session_start_time.isoformat() if row.session_start_time else "",
                row.session_end_time.isoformat() if row.session_end_time else "",
                row.total_study_duration or 0
            ])
            yield output.getvalue()
            output.seek(0)
            output.truncate(0)
    
    # Generate filename with timestamp
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    safe_project_name = "".join(c if c.isalnum() or c in " _-" else "_" for c in project.name)
    safe_project_name = safe_project_name.replace(" ", "_")[:50]
    filename = f"completed_panelists_{safe_project_name}_{timestamp}.csv"
    
    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"',
        "X-Total-Records": str(len(results))
    }
    
    return StreamingResponse(
        generate_csv(),
        media_type="text/csv",
        headers=headers
    )


@router.post("/{project_id}/export-abandoned-responses")
def export_abandoned_responses_endpoint(
    project_id: UUID,
    payload: ExportAbandonedResponsesRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """
    Export abandoned responses from all studies in a project as CSV.
    
    Includes panelist details, demographics, classification status, and task progress.
    Optimized for fast response using a single SQL query with joins.
    
    Args:
        project_id: UUID of the project
        payload: Optional time filters (after_utc, before_utc)
    
    Returns:
        StreamingResponse with CSV data
    """
    from fastapi import HTTPException
    from sqlalchemy import text
    from datetime import datetime
    
    # Verify project exists and user has access
    project = project_service.get_project(db=db, project_id=project_id, user_id=current_user.id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    # Build optimized single query with classification count subquery
    query = """
        SELECT 
            s.title as study_name,
            s.id as study_id,
            sr.panelist_id,
            sr.product_id,
            sr.respondent_id,
            sr.session_id,
            sr.personal_info,
            sr.completed_tasks_count,
            sr.total_tasks_assigned,
            sr.completion_percentage,
            sr.session_start_time,
            sr.abandonment_timestamp,
            sr.abandonment_reason,
            sr.total_study_duration,
            sr.last_activity,
            (SELECT COUNT(*) FROM classification_answers ca WHERE ca.study_response_id = sr.id) as classification_count
        FROM study_responses sr
        INNER JOIN studies s ON sr.study_id = s.id
        WHERE s.project_id = :project_id
          AND sr.is_abandoned = true
    """
    params: Dict[str, Any] = {"project_id": str(project_id)}
    
    if payload.after_utc:
        query += " AND sr.abandonment_timestamp >= :after_utc"
        params["after_utc"] = payload.after_utc
    
    if payload.before_utc:
        query += " AND sr.abandonment_timestamp < :before_utc"
        params["before_utc"] = payload.before_utc
    
    query += " ORDER BY sr.abandonment_timestamp DESC"
    
    # Execute query
    results = db.execute(text(query), params).fetchall()
    
    # Generate CSV using streaming for memory efficiency
    def generate_csv():
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow([
            "study_name",
            "study_id",
            "panelist_id",
            "product_id",
            "respondent_id",
            "session_id",
            "age",
            "gender",
            "did_classification",
            "classification_questions_answered",
            "tasks_completed",
            "total_tasks",
            "completion_percentage",
            "session_start_time",
            "abandonment_timestamp",
            "abandonment_reason",
            "total_duration_seconds",
            "last_activity"
        ])
        yield output.getvalue()
        output.seek(0)
        output.truncate(0)
        
        # Write data rows
        for row in results:
            personal_info = row.personal_info or {}
            age = personal_info.get("age", "N/A")
            gender = personal_info.get("gender", "N/A")
            classification_count = row.classification_count or 0
            did_classification = "Yes" if classification_count > 0 else "No"
            
            writer.writerow([
                row.study_name,
                str(row.study_id),
                row.panelist_id or "N/A",
                row.product_id or "N/A",
                row.respondent_id,
                row.session_id,
                age,
                gender,
                did_classification,
                classification_count,
                row.completed_tasks_count or 0,
                row.total_tasks_assigned or 0,
                row.completion_percentage or 0,
                row.session_start_time.isoformat() if row.session_start_time else "",
                row.abandonment_timestamp.isoformat() if row.abandonment_timestamp else "",
                row.abandonment_reason or "",
                row.total_study_duration or 0,
                row.last_activity.isoformat() if row.last_activity else ""
            ])
            yield output.getvalue()
            output.seek(0)
            output.truncate(0)
    
    # Generate filename with timestamp
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    safe_project_name = "".join(c if c.isalnum() or c in " _-" else "_" for c in project.name)
    safe_project_name = safe_project_name.replace(" ", "_")[:50]
    filename = f"abandoned_responses_{safe_project_name}_{timestamp}.csv"
    
    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"',
        "X-Total-Records": str(len(results))
    }
    
    return StreamingResponse(
        generate_csv(),
        media_type="text/csv",
        headers=headers
    )


@router.get("/public/{project_id}/studies")
def get_public_project_studies_endpoint(
    project_id: UUID,
    db: Session = Depends(get_db),
):
    """
    Get public studies for a project.
    Returns project name and a list of active studies.
    No authentication required.
    """
    from app.models.project_model import Project
    from app.models.study_model import Study
    from app.models.user_model import User
    from fastapi import HTTPException
    from sqlalchemy import select

    cache_key = f"project_public_studies:{project_id}"
    cached_data = RedisCache.get(cache_key)
    if cached_data:
        return cached_data

    # Verify project exists and get its name + creator email
    project_row = db.execute(
        select(Project.name, User.email)
        .select_from(Project)
        .join(User, User.id == Project.creator_id, isouter=True)
        .where(Project.id == project_id)
    ).first()

    if not project_row:
        raise HTTPException(status_code=404, detail="Project not found")

    # Get active studies for this project
    studies = db.execute(
        select(Study.id, Study.title, Study.study_type, Study.product_id)
        .where(Study.project_id == project_id, Study.status == 'active')
        .order_by(Study.created_at.desc())
    ).all()

    result = {
        "project_name": project_row.name,
        "creator_email": project_row.email,
        "studies": [
            {
                "id": str(study.id),
                "title": study.title,
                "study_type": study.study_type,
                "product_id": study.product_id,
            }
            for study in studies
        ]
    }
    RedisCache.set(cache_key, result, ttl_seconds=7)
    return result
